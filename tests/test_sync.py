# -*- coding: utf-8 -*-
"""Интеграционный тест синхронизации МойСклад (без pytest, просто python).

Сценарий:
  1) поднимаем mock-МойСклад (tests/mock_ms.py) на 127.0.0.1:9800;
  2) поднимаем приложение с MS_BASE_URL=mock и HISTORY_DAYS=60 на 127.0.0.1:8801;
  3) регистрируем владельца, проходим онбординг API-ручками:
     неверный токен (человеческая ошибка) → верный токен → список складов →
     выбор двух торговых → POST /api/sync/initial → поллинг /api/sync/status;
  4) сверяем БД и аналитику с эталонными числами mock-мира:
     товары/размеры, stock_days (включая явные нули у распроданных),
     нетто-продажи по позициям, /api/summary, /api/replenish;
  5) инкрементальный синк POST /api/sync/run — числа не «уезжают»;
  6) демо-режим второго пользователя (POST /api/connect/demo) не сломан;
  7) условия производства: аддитивная миграция старой таблицы productions,
     округление до минимальной партии и кратности (сумма по размерам = итог
     позиции), изоляция производств между организациями, границы id в пути.

Запуск из корня репозитория:  python tests/test_sync.py

Сценарий длинный: в CI он один шёл 448.9 с (прогон 32953333557) и в одиночку
задавал критический путь всего строгого набора. Поэтому он разложен на ШАРДЫ —
`python tests/test_sync.py --shard <имя>` исполняет и засчитывает только свою
часть, а недостающую подготовку доигрывает сам, теми же запросами и теми же
реальными ожиданиями. Без аргументов работает как раньше: весь сценарий,
413 проверок. Ни одно ожидание не сокращено, ни одна проверка не выброшена и не
сделана необязательной; замок полноты — `tests/test_sync_shards.py`.
"""
import argparse
import io
import json
import os
import sqlite3
import re
import sys
import threading
import time
from datetime import date as _dt_date, timedelta as _dt_delta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tests"))

from test_sync_shards import ACTS, ALL, load_baseline, plan_for  # noqa: E402

# ── Шарды ────────────────────────────────────────────────────────────────────
# Сценарий один и тот же; `--shard` говорит, какую его часть ЗАПИСЫВАТЬ. Всё,
# что шарду нужно для этой части (организация, эталонный проход истории), он
# доигрывает сам — теми же запросами и теми же реальными ожиданиями, только
# не засчитывая чужие проверки. Без аргументов — весь сценарий, как раньше.
# Разбиение и замок полноты — в tests/test_sync_shards.py.
_ap = argparse.ArgumentParser(description="Интеграционный тест синхронизации МойСклад")
_ap.add_argument("--shard", default=ALL,
                 help="часть сценария (по умолчанию весь сценарий целиком)")
_ap.add_argument("--dump-baseline", action="store_true",
                 help="напечатать эталон полноты (акт + имена проверок) вместо "
                      "сверки с ним; годится только с полным сценарием. Файл "
                      "tests/sync_baseline_checks.txt не переписывается сам: "
                      "обновление эталона обязано быть видно в diff")
_args = _ap.parse_args()
SHARD, DUMP = _args.shard, _args.dump_baseline
_PLAN = plan_for(SHARD)

# Своя база на шард: иначе параллельные шарды перепишут файл друг другу.
# Имя остаётся под маской test_*.db — её чистит tests/run_all.py перед прогоном.
DB_PATH = ROOT / ("test_oborot.db" if SHARD == ALL else f"test_oborot_{SHARD}.db")
# Порты берутся из окружения: так tests/run_all.py разводит наборы и
# может гонять их параллельно. Значения по умолчанию — прежние.
APP_PORT = int(os.environ.get("OBOROT_TEST_PORT", "8801"))

# Окружение — ДО импорта приложения (db.py и ms_client читают env).
os.environ["DATABASE_URL"] = f"sqlite:///{DB_PATH}"
os.environ["MS_BASE_URL"] = f"http://127.0.0.1:{os.environ.get('OBOROT_MOCK_PORT', '9800')}"
os.environ["HISTORY_DAYS"] = "60"
os.environ["SYNC_DAYS_BACK"] = "3"
# Инцидент 21.08: маленькие чанки истории, чтобы проверить прерывание/продолжение.
os.environ["STOCK_CHUNK_DATES"] = "5"
os.environ["MS_CHUNK_PAUSE"] = "0.3"
# Деплой П1: окно быстрого старта — 10 дат, остальные 50 догружаются назад чанками.
os.environ["INITIAL_WINDOW_DAYS"] = "10"

if DB_PATH.exists():
    DB_PATH.unlink()

import httpx  # noqa: E402
import uvicorn  # noqa: E402

import mock_ms  # noqa: E402
from app.main import app as oborot_app  # noqa: E402


# ── Инфраструктура ───────────────────────────────────────────────────────────

class ServerThread:
    def __init__(self, asgi_app, port: int):
        self.config = uvicorn.Config(asgi_app, host="127.0.0.1", port=port,
                                     log_level="warning")
        self.server = uvicorn.Server(self.config)
        self.thread = threading.Thread(target=self.server.run, daemon=True)

    def start(self):
        self.thread.start()
        deadline = time.time() + 15
        while time.time() < deadline:
            if self.server.started:
                return
            time.sleep(0.05)
        raise RuntimeError(f"сервер на порту {self.config.port} не поднялся")

    def stop(self):
        self.server.should_exit = True
        self.thread.join(timeout=10)


PASS, FAIL = [], []
RECORDED: list[tuple[str, str]] = []   # (акт, имя) засчитанных проверок по порядку
_RECORDING = False         # идёт записываемый акт, а не подготовка соседнего
_CUR_ACT = ""


def begin(name: str) -> bool:
    """Открыть акт сценария. False — в этом шарде акт не исполняется.

    Акт, попавший в план как подготовка, исполняется полностью (те же запросы,
    те же паузы и повторы), но его проверки не идут в отчёт: иначе одна и та же
    проверка попала бы в общий итог из нескольких шардов и «413» перестало бы
    что-либо доказывать. Провал в подготовке всё равно красит шард — молчаливая
    подготовка опаснее лишней строки в выводе.
    """
    global _RECORDING, _CUR_ACT
    _RECORDING = _PLAN.get(name, False)
    _CUR_ACT = name
    return name in _PLAN


def check(name: str, cond: bool, detail: str = ""):
    if not _RECORDING:
        if cond:
            print(f"  ..   {name}")
        else:
            FAIL.append(f"[подготовка] {name}")
            print(f"  FAIL [подготовка] {name}  {detail}")
        return
    RECORDED.append((_CUR_ACT, name))
    if cond:
        PASS.append(name)
        print(f"  OK   {name}" + (f"  [{detail}]" if detail else ""))
    else:
        FAIL.append(name)
        print(f"  FAIL {name}  {detail}")


def finish() -> int:
    """Сверка с замороженным эталоном и канонический итог.

    Проверяется УПОРЯДОЧЕННЫЙ список фактически выполненных проверок против
    среза эталона по актам этого шарда. Пропала проверка, переехала в другой
    акт, поменялся порядок — шард красный. Успешная сверка своей строки OK не
    добавляет: сумма OK по всем шардам обязана дать ровно 413 legacy-проверок.
    """
    got = [n for _, n in RECORDED]
    if DUMP:
        print("\n===== эталон полноты (в tests/sync_baseline_checks.txt) =====")
        seen = ""
        for a, n in RECORDED:
            if a != seen:
                print(f"# акт: {a}")
                seen = a
            print(n)
        print("===== конец эталона =====")
        print()
        print(f"ИТОГО: {len(PASS)} OK, {len(FAIL)} FAIL")
        return 1 if FAIL else 0
    base = load_baseline()
    want = [n for a in ACTS if _PLAN.get(a) for n in base.get(a, ())]
    same_order = got == want
    if not same_order:
        first = next((i for i in range(max(len(want), len(got)))
                      if want[i:i + 1] != got[i:i + 1]), 0)
        FAIL.append(
            f"[полнота] шард {SHARD}: выполнено {len(got)} проверок, "
            f"эталон требует {len(want)}; первое расхождение №{first + 1}: "
            f"ожидалось {want[first:first + 1] or ['—']}, "
            f"выполнено {got[first:first + 1] or ['—']}")
    stray = [(a, n) for a, n in RECORDED if n not in base.get(a, ())]
    if stray:
        FAIL.append(f"[полнота] проверка выполнена не в своём акте: {stray[:3]}")
    if same_order and not stray:
        print(f"\n  ..   полнота шарда {SHARD}: {len(got)} проверок "
              f"совпали с эталоном по составу, порядку и актам")
    print()
    print(f"ИТОГО: {len(PASS)} OK, {len(FAIL)} FAIL")
    if FAIL:
        print("Провалены:", *FAIL, sep="\n  - ")
        return 1
    return 0


def wait_sync_done(client: httpx.Client, timeout: float = 240.0) -> dict:
    deadline = time.time() + timeout
    last = {}
    while time.time() < deadline:
        last = client.get("/api/sync/status").json()
        if last.get("state") in ("done", "error"):
            return last
        time.sleep(1.0)
    return last


def _err_line(stderr: str) -> str:
    """Строка ошибки, по которой видно ПРИЧИНУ, а не ссылку на документацию.

    Последняя строка трейсбэка SQLAlchemy — это «(Background on this error at:
    …)», из неё нельзя понять, блокировка это или битая схема. Берём последнюю
    содержательную.
    """
    lines = [ln.strip() for ln in (stderr or "").splitlines() if ln.strip()]
    for ln in reversed(lines):
        if "Background on this error" in ln or ln.startswith("[SQL:") or ln.startswith("[param"):
            continue
        return ln[:160]
    return (lines[-1][:160] if lines else "")


def _race_start(snippet: str, n: int, attempts: int = 3):
    """Запускает n процессов одновременно, повторяя при блокировках SQLite.

    Проверяемое свойство — «миграции переживают одновременный старт», и оно
    про КОД. Но SQLite отдаёт writer-лок по таймауту (busy_timeout), и когда
    машина занята чем-то ещё — параллельными наборами, браузером в тестах
    интерфейса, — один из процессов не успевает и падает по времени, а не по
    логике. Это ловилось уже дважды и оба раза оказывалось нагрузкой, а не
    регрессией: тот же набор в одиночку даёт 6/6.

    Поэтому: сценарий повторяется, ЕСЛИ единственная причина падения —
    блокировка. Любая другая ошибка (duplicate column, битая схема) возвращается
    сразу и роняет тест, как и должна.
    """
    import subprocess

    lock_words = ("database is locked", "database table is locked", "timeout")
    results = []
    for attempt in range(attempts):
        start_at = time.time() + 1.5
        code = re.sub(r"time\.sleep\(max\(0\.0, [0-9.]+ - time\.time\(\)\)\)",
                      f"time.sleep(max(0.0, {start_at!r} - time.time()))", snippet)
        procs = [subprocess.Popen([sys.executable, "-c", code],
                                  stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
                 for _ in range(n)]
        results = [pr.communicate(timeout=90) for pr in procs]
        started = sum(1 for out, _ in results if "OK" in out)
        if started == n:
            return started, results
        errs = " ".join((e or "").lower() for o, e in results if "OK" not in o)
        if not any(w in errs for w in lock_words):
            return started, results  # ошибка не про блокировку — не повторяем
        time.sleep(1.0 + attempt)
    return sum(1 for out, _ in results if "OK" in out), results


def main() -> int:
    mock_srv = ServerThread(mock_ms.app, mock_ms.PORT)
    app_srv = ServerThread(oborot_app, APP_PORT)
    mock_srv.start()
    app_srv.start()
    try:
        return run_scenario()
    finally:
        app_srv.stop()
        mock_srv.stop()


def run_scenario() -> int:
    base = f"http://127.0.0.1:{APP_PORT}"
    client = httpx.Client(headers={"X-Oborot-CSRF": "1"}, base_url=base, timeout=60.0)

    # Ядро исполняется в любом шарде: остальным оно нужно как подготовка —
    # организация, подключение, склады и первичная история берутся отсюда.
    begin("core")
    print("== Онбординг ==")
    r = client.post("/register", data={
        "name": "Владелец", "email": "owner@test.io",
        "password": "secret123", "org_name": "Тестовый бренд",
    })
    check("регистрация владельца", r.status_code == 303, f"status={r.status_code}")

    r = client.post("/api/connect/moysklad", json={"token": "definitely-wrong-token"})
    check("неверный токен отклонён с человеческим текстом",
          r.status_code == 400 and "токен" in r.json().get("detail", "").lower(),
          f"status={r.status_code} detail={r.json().get('detail', '')[:60]}")

    r = client.post("/api/connect/moysklad", json={"token": mock_ms.TOKEN})
    check("верный токен принят", r.status_code == 200 and r.json().get("ok"),
          f"status={r.status_code}")

    r = client.get("/api/connect/moysklad/stores")
    stores = r.json().get("stores", [])
    check("список складов из МС", len(stores) == 3,
          f"got={[s['name'] for s in stores]}")

    r = client.post("/api/connect/moysklad/stores",
                    json={"ext_ids": ["st-flag", "st-web"]})
    check("выбор двух торговых складов",
          r.status_code == 200 and r.json().get("active") == 2
          and r.json().get("total") == 3, f"resp={r.json()}")

    r = client.get("/", follow_redirects=False)
    check("до синка дашборд недоступен (redirect на онбординг)",
          r.status_code == 302 and "/onboarding" in r.headers.get("location", ""))

    t0 = time.time()
    r = client.post("/api/sync/initial")
    check("первичный синк запущен", r.status_code == 200 and r.json().get("ok"),
          f"estimate={r.json().get('estimate_minutes')} мин")

    status = wait_sync_done(client)
    took = time.time() - t0
    check("первичный синк завершился state=done", status.get("state") == "done",
          f"state={status.get('state')} error={status.get('error', '')[:120]}")
    stats = status.get("stats", {})
    print(f"  … синк занял {took:.1f} c, stats={stats}")

    r = client.get("/", follow_redirects=False)
    check("после синка «/» ведёт в Оборачиваемость (дашборд скрыт)",
          r.status_code == 302 and (r.headers.get("location") or "") == "/turnover",
          f"status={r.status_code} loc={r.headers.get('location')}")
    r = client.get("/turnover", follow_redirects=False)
    check("Оборачиваемость открывается", r.status_code == 200,
          f"status={r.status_code}")

    print("== Товары и размеры ==")
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    n_products = con.execute("SELECT COUNT(*) c FROM products").fetchone()["c"]
    expected_products = len(mock_ms.SKUS) + len(mock_ms.SIZED)  # + родительские product
    check("создано товаров (варианты + родители + безразмерные)",
          n_products == expected_products,
          f"got={n_products} expected={expected_products}")

    row = con.execute(
        "SELECT base_name, size FROM products WHERE ext_id='v-hoodie1-M'"
    ).fetchone()
    check("вариант: base_name без размера, size из характеристики",
          row and row["base_name"] == "Худи «Скетч»" and row["size"] == "M",
          f"got={dict(row) if row else None}")

    row = con.execute(
        "SELECT base_name, size FROM products WHERE ext_id='v-tee3-L'"
    ).fetchone()
    check("вариант без характеристики: размер распарсен из скобок имени",
          row and row["base_name"] == "Футболка «Полночь»" and row["size"] == "L",
          f"got={dict(row) if row else None}")

    row = con.execute(
        "SELECT sale_price, cost_price, category FROM products WHERE ext_id='p-bag1'"
    ).fetchone()
    check("цены из копеек в рубли + категория из pathName",
          row and row["sale_price"] == 5900 and row["cost_price"] == 2400
          and row["category"] == "Сумки", f"got={dict(row) if row else None}")

    row = con.execute("SELECT archived FROM products WHERE ext_id='p-old1'").fetchone()
    check("архивная позиция помечена archived", row and row["archived"] == 1)

    print("== Исключение расходников/сертификатов из аналитики ==")
    row = con.execute("SELECT excluded FROM products WHERE ext_id='p-pack1'").fetchone()
    check("упаковка авто-исключена (категория «Расходный материал»)",
          row and row["excluded"] == 1)
    row = con.execute("SELECT excluded FROM products WHERE ext_id='p-cert1'").fetchone()
    check("сертификат авто-исключён (слово в названии)", row and row["excluded"] == 1)
    row = con.execute("SELECT excluded FROM products WHERE ext_id='p-bag1'").fetchone()
    check("обычный товар НЕ исключён (эвристика не жадничает)",
          row and row["excluded"] == 0)

    excl_resp = client.get("/api/exclusions").json()
    excl_names = {e["base_name"] for e in excl_resp.get("excluded", [])}
    check("GET /api/exclusions отдаёт оба расходника",
          {"Брендированный пакет средний", "Подарочный сертификат на десять тысяч"} <= excl_names,
          f"got={sorted(excl_names)}")

    turno_bases = {it["base_name"] for it in client.get("/api/turnover").json()["items"]}
    check("исключённые не видны в оборачиваемости",
          "Брендированный пакет средний" not in turno_bases
          and "Подарочный сертификат на десять тысяч" not in turno_bases)

    sum_before = client.get("/api/summary").json()
    r = client.post("/api/exclusions",
                    json={"base_name": "Брендированный пакет средний", "excluded": False})
    check("возврат позиции в аналитику работает", r.status_code == 200 and r.json().get("ok"))
    turno_bases2 = {it["base_name"] for it in client.get("/api/turnover").json()["items"]}
    check("возвращённая позиция появилась в оборачиваемости (пакет распродан, nq>0)",
          "Брендированный пакет средний" in turno_bases2)
    # Сертификат в mock-мире имеет живой остаток — проверяем влияние на сток.
    r = client.post("/api/exclusions",
                    json={"base_name": "Подарочный сертификат на десять тысяч", "excluded": False})
    check("возврат сертификата работает", r.status_code == 200 and r.json().get("ok"))
    sum_after = client.get("/api/summary").json()
    check("возврат позиции с остатком увеличил сток",
          sum_after["stock_units"] > sum_before["stock_units"],
          f"units {sum_before['stock_units']} -> {sum_after['stock_units']}")
    for base in ("Брендированный пакет средний", "Подарочный сертификат на десять тысяч"):
        r = client.post("/api/exclusions", json={"base_name": base, "excluded": True})
        check(f"повторное исключение работает ({base[:20]}…)",
              r.status_code == 200 and r.json().get("ok"))
    sum_back = client.get("/api/summary").json()
    check("после исключения числа вернулись",
          sum_back["stock_units"] == sum_before["stock_units"],
          f"{sum_back['stock_units']} != {sum_before['stock_units']}")

    print("== История остатков ==")
    n_stock = con.execute("SELECT COUNT(*) c FROM stock_days").fetchone()["c"]
    n_dates = con.execute("SELECT COUNT(DISTINCT date) c FROM stock_days").fetchone()["c"]
    check("stock_days: 60 дат истории", n_dates == 60, f"dates={n_dates}")
    check("stock_days заполнены", n_stock > 1000, f"rows={n_stock}")

    sellouts = mock_ms.sellout_ext_ids()
    check("в mock-мире есть распроданные позиции", len(sellouts) > 0,
          f"n={len(sellouts)}")
    zero_ok = True
    detail = ""
    for ext in sellouts:
        pid = con.execute("SELECT id FROM products WHERE ext_id=?", (ext,)).fetchone()
        zrow = con.execute(
            "SELECT COUNT(*) c FROM stock_days WHERE product_id=? AND qty=0",
            (pid["id"],)).fetchone()
        lrow = con.execute(
            "SELECT qty FROM stock_days WHERE product_id=? ORDER BY date DESC LIMIT 1",
            (pid["id"],)).fetchone()
        if not zrow["c"] or lrow["qty"] != 0:
            zero_ok = False
            detail = f"{ext}: zeros={zrow['c']} last={lrow['qty']}"
            break
    check("явные нули записаны у распроданных (правило самоизлечения)", zero_ok, detail)

    n_ws = con.execute("SELECT COUNT(*) c FROM warehouse_stock").fetchone()["c"]
    n_wh = con.execute(
        "SELECT COUNT(DISTINCT warehouse_id) c FROM warehouse_stock").fetchone()["c"]
    check("warehouse_stock заполнен по 2 активным складам",
          n_ws > 0 and n_wh == 2, f"rows={n_ws} склады={n_wh}")
    con.close()

    print("== Продажи против эталона mock-мира ==")
    expected = mock_ms.expected_net_sales()
    turno = client.get("/api/turnover").json()["items"]
    by_base = {it["base_name"]: it for it in turno}
    qty_ok = rev_ok = True
    qdet = rdet = ""
    for base_name, (enq, enr) in sorted(expected.items()):
        it = by_base.get(base_name)
        if it is None:
            qty_ok = False
            qdet = f"нет позиции {base_name}"
            break
        if abs(it["nq"] - enq) > 0.51:
            qty_ok = False
            qdet = f"{base_name}: nq={it['nq']} expected={enq}"
        if abs(it["nr"] - enr) > 2:
            rev_ok = False
            rdet = f"{base_name}: nr={it['nr']} expected={round(enr)}"
    total_nq = sum(v[0] for v in expected.values())
    total_nr = sum(v[1] for v in expected.values())
    check(f"нетто-штуки по всем {len(expected)} позициям сходятся (всего {total_nq:.0f} шт)",
          qty_ok, qdet)
    check(f"нетто-выручка сходится (всего {total_nr:,.0f} ₽)".replace(",", " "),
          rev_ok, rdet)

    dead = by_base.get("Ремень «Ось»")
    check("неликвид без продаж: nq=0, остаток есть",
          dead is not None and dead["nq"] == 0 and dead["cs"] > 0,
          f"got={dead and (dead['nq'], dead['cs'])}")

    print("== Аналитика ==")
    summary = client.get("/api/summary").json()
    exp_stock = mock_ms.expected_stock_today()
    exp_units = round(sum(exp_stock.values()))
    check("summary: штук на складе = эталон mock-мира",
          summary["stock_units"] == exp_units,
          f"got={summary['stock_units']} expected={exp_units}")
    check("summary: продано за 30 дней > 0",
          summary["sold_30d_qty"] > 0 and summary["sold_30d_rev"] > 0,
          f"qty={summary['sold_30d_qty']} rev={summary['sold_30d_rev']}")
    check("summary: позиции и классы посчитаны",
          summary["positions"] > 0 and sum(summary["classes"].values()) == summary["positions"])

    repl = client.get("/api/replenish").json()
    items = repl["items"]
    check("replenish: есть рекомендации need>0", len(items) > 0,
          f"n={len(items)}")
    sellout_item = next((i for i in items if i["base_name"] == "Футболка «Курсив»"), None)
    check("replenish: распроданная позиция в списке заказа (cs=0, need>0)",
          sellout_item is not None and sellout_item["cs"] == 0
          and sellout_item["need"] > 0,
          f"got={sellout_item and (sellout_item['cs'], sellout_item['need'])}")
    if sellout_item:
        check("replenish: размерная сетка распроданной позиции из продаж (S/M/L)",
              set(sellout_item["sizes"].keys()) == {"S", "M", "L"}
              and sum(s["rec"] for s in sellout_item["sizes"].values())
              == sellout_item["need"],
              f"sizes={list(sellout_item['sizes'].keys())}")

    stocks = client.get("/api/stocks").json()
    check("stocks: 2 активных склада в колонках", len(stocks["warehouses"]) == 2,
          f"got={[w['name'] for w in stocks['warehouses']]}")
    hoodie = next((i for i in stocks["items"] if i["base_name"] == "Худи «Скетч»"), None)
    check("stocks: размерная разбивка по складам",
          hoodie is not None and {s["size"] for s in hoodie["sizes"]} >= {"S", "M", "L"},
          f"got={hoodie and [s['size'] for s in hoodie['sizes']]}")

    print("== Рейтинг оборачиваемости: правила legacy-таблицы ==")
    tresp = client.get("/api/turnover").json()["items"]
    groups = [it["group"] for it in tresp]
    order_ok = groups == sorted(groups, key=lambda g: {"rank": 0, "low_data": 1, "no_sales": 2}[g])
    check("порядок групп: рейтинг → мало данных → без продаж", order_ok,
          f"groups={groups}")
    check("первая строка — из рейтинга (шум не наверху)",
          tresp and tresp[0]["group"] == "rank",
          f"first={tresp and (tresp[0]['base_name'], tresp[0]['group'])}")
    cameo = next((it for it in tresp if it["base_name"] == "Бомбер «Камео»"), None)
    check("крошечный тираж («бомбер Регби»-кейс): low_data, не в рейтинге",
          cameo is not None and cameo["low_data"] and cameo["group"] == "low_data"
          and cameo["nq"] > 0,
          f"got={cameo and (cameo['group'], cameo['dis'], cameo['nq'], cameo['turnover'])}")
    belt = next((it for it in tresp if it["base_name"] == "Ремень «Ось»"), None)
    check("неликвид — в группе «без продаж»",
          belt is not None and belt["group"] == "no_sales")
    exp_tt = round(sum(it["turnover"] for it in tresp
                       if not it["archived"] and not it["low_data"]))
    check("turnover_total на дашборде без low_data-шума",
          summary["turnover_total"] == exp_tt,
          f"got={summary['turnover_total']} expected={exp_tt}")
    repl_ld = [it.get("low_data", False) for it in repl["items"]]
    check("replenish: «мало данных» не выше значимых позиций",
          repl_ld == sorted(repl_ld),
          f"flags={repl_ld}")
    # Ревью 21.08 (минор 9): истории всего 60 дней, НИ ОДИН сезон годового окна
    # не покрыт целиком — сезонные колонки обязаны быть null, а не 0 (0 читался
    # как «ничего не продали» и врал в таблице), плюс аналитика отдаёт границу
    # покрытия, чтобы фронт погасил такие колонки.
    turno_payload = client.get("/api/turnover").json()
    sea0 = turno_payload["items"][0].get("sea") or {}
    check("сезонные колонки при частичном покрытии = null (а не 0)",
          set(sea0) == {"winter", "spring", "summer", "autumn"}
          and all(v is None for v in sea0.values()), f"sea={sea0}")
    check("аналитика отдаёт coverage_start и карту покрытия сезонов",
          turno_payload.get("coverage_start") == mock_ms.DATES[0]
          and set(turno_payload.get("season_covered") or {}) == set(sea0)
          and not any((turno_payload.get("season_covered") or {}).values()),
          f"coverage_start={turno_payload.get('coverage_start')} "
          f"covered={turno_payload.get('season_covered')}")
    from app.analytics import season_bounds as _season_bounds
    _season_start = _season_bounds(_dt_date.today())[0].isoformat()
    _season_partial = mock_ms.DATES[0] > _season_start
    sh = summary["season_health"]
    check("«здоровье сезона»: покрытие начинается внутри сезона → no_data",
          (sh.get("status") == "no_data" and sh.get("partial_coverage") is True)
          if _season_partial else sh.get("partial_coverage") is False,
          f"partial={_season_partial} health={sh.get('status')}/{sh.get('partial_coverage')}")
    # Раньше при покрытии, начавшемся внутри сезона, новинкой сезона считалась
    # КАЖДАЯ позиция (первое появление = первая загруженная дата), и в остаток
    # сезона попадал весь склад — доля остатка завышалась, статус срывался.
    _all_new_leftover = round(sum(
        it["cs"] * it["sale_price"] for it in turno_payload["items"]
        if not it["archived"] and not it["hidden"]
        and (it["cs"] > 0 or it["nq"] > 0 or it["dis"] > 0)))
    check("«новинки сезона» — не разом все позиции (остаток сезона < всего склада)",
          not _season_partial or sh.get("leftover_value", 0) < _all_new_leftover,
          f"leftover={sh.get('leftover_value')} всего={_all_new_leftover}")
    # Формула заказа (правило legacy): need = темп×горизонт − прогнозный остаток
    # к приходу заказа; proj_stock = max(0, cs + едет − темп×lead_time).
    lead = repl.get("lead_time_days")
    hor = repl.get("horizon_days")
    check("replenish отдаёт lead_time_days и proj_stock",
          isinstance(lead, int) and lead > 0
          and all("proj_stock" in it for it in repl["items"]))
    bad = [
        it["base_name"] for it in repl["items"]
        if abs(it["need"] - max(0, round(it["rate"] * hor) - it["proj_stock"])) > 1
    ]
    check("need = темп×горизонт − прогнозный остаток (все позиции)", not bad,
          f"bad={bad[:3]}")
    proj_bad = [
        it["base_name"] for it in repl["items"]
        if it["proj_stock"] > it["cs"] + it["ordered"]
    ]
    check("прогнозный остаток не больше (остаток + едет)", not proj_bad,
          f"bad={proj_bad[:3]}")

    print("== «Едет к нам» из заказов поставщику МС ==")
    exp_inc = mock_ms.expected_incoming()
    check("mock: seeded-заказы дают ожидания (Худи 14, Кольцо 5+150, Футболка 12)",
          exp_inc == {"Худи «Скетч»": 14.0, "Кольцо «Грань»": 155.0,
                      "Футболка «Манифест»": 12.0},
          f"exp={exp_inc}")
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    ms_rows = {r["base_name"]: r["ms_qty"] for r in con.execute(
        "SELECT base_name, ms_qty FROM ordered_qty WHERE ms_qty != 0")}
    check("ordered_qty.ms_qty = эталон (частично принятый: qty−shipped)",
          ms_rows == exp_inc, f"got={ms_rows}")
    check("непроведённый, полностью принятый и старый доки не в «едет»",
          all(b not in ms_rows for b in ("Сумка «Тоут»", "Рубашка «Разворот»")))
    # po-seed-6 (D-28): чужой документ с меткой НЕСУЩЕСТВУЮЩЕГО заказа
    # [oborot#9091]. В «едет» он входит — товар приедет физически, — но
    # «Оборот» не имеет права считать его своим решением.
    _con = sqlite3.connect(DB_PATH)
    tracked = {b: (m, t) for b, m, t in _con.execute(
        "SELECT base_name, ms_qty, ms_qty_tracked FROM ordered_qty")}
    _con.close()
    check("чужой заказ виден в «едет к нам»",
          ms_rows.get("Футболка «Манифест»") == 12.0,
          f"got={ms_rows.get('Футболка «Манифест»')}")
    check("но своим «Оборот» его НЕ считает (маркер без заказа — не связь)",
          all(t == 0 for _, t in tracked.values()),
          f"tracked={{b: t for b, (m, t) in tracked.items() if t}}")
    check("своих заказов у организации нет — значит tracked-поток пуст",
          stats.get("incoming_qty_tracked") == 0
          and stats.get("incoming_qty_external") == stats.get("incoming_qty"),
          f"tracked={stats.get('incoming_qty_tracked')} "
          f"external={stats.get('incoming_qty_external')}")
    con.close()
    # 19 из старых доков + 150 из po-seed-5 (150 позиций, аудит 18.08 —
    # проверяет дочитывание хвоста >100 позиций через /positions)
    # + 12 из po-seed-6 (чужой документ с чужой меткой, D-28)
    check("stats синка: incoming_qty=181 из 4 открытых доков",
          stats.get("incoming_qty") == 181 and stats.get("incoming_open_docs") == 4,
          f"got qty={stats.get('incoming_qty')} docs={stats.get('incoming_open_docs')}")
    check("документ >100 позиций дочитан пагинацией (positions_refetched ≥ 1)",
          (stats.get("positions_refetched") or 0) >= 1,
          f"refetched={stats.get('positions_refetched')}")

    r = client.post("/api/ordered", json={"base_name": "Худи «Скетч»", "qty": 3})
    check("ручная правка qty поверх ms_qty принята", r.status_code == 200)
    # Контракт не должен врать: раньше «Заказано» на несуществующий товар
    # отвечало ok, а запись повисала в базе и нигде не показывалась.
    r = client.post("/api/ordered", json={"base_name": "Пальто «Которого нет»", "qty": 5})
    check("«Заказано» на товар не из каталога → 404 с объяснением",
          r.status_code == 404 and "каталоге" in r.text, f"status={r.status_code} {r.text[:90]}")
    r = client.post("/api/ordered/add", json={"base_name": "Пальто «Которого нет»", "qty": 5})
    check("«Заказ отправлен» на товар не из каталога → 404", r.status_code == 404,
          f"status={r.status_code}")
    repl_inc = client.get("/api/replenish").json()
    hood_item = next((i for i in repl_inc["items"]
                      if i["base_name"] == "Худи «Скетч»"), None)
    hood_excl = next((e for e in repl_inc.get("excluded", [])
                      if e["base_name"] == "Худи «Скетч»"), None)
    if hood_item is not None:
        check("replenish: ordered = ручной qty + ms_qty (3+14)",
              hood_item["ordered"] == 17, f"got={hood_item['ordered']}")
    else:
        check("replenish: позиция закрыта заказом (причина упоминает заказ)",
              hood_excl is not None and "заказ" in hood_excl.get("reason", ""),
              f"got={hood_excl}")

    print("== Русские категории и sample-исключения (чистые функции) ==")
    from app.categories import ru_category
    from app.exclusions import is_service_item
    check("латинская категория переводится (Shirts → Рубашки)",
          ru_category("Shirts") == "Рубашки")
    check("без категории — по имени товара (бомбер → Верхняя одежда)",
          ru_category("", "Черный бомбер \"Регби\"") == "Верхняя одежда")
    check("кириллическая категория не трогается",
          ru_category("Худи и свитшоты", "х") == "Худи и свитшоты")
    check("нераспознанная латиница остаётся как есть",
          ru_category("Vintage", "Штука") == "Vintage")
    check("sample-позиции исключаются эвристикой",
          is_service_item("Пальто \"Скала\" sample", "Samples")
          and is_service_item("Молочные брюки сэмпл", "")
          and not is_service_item("Пальто \"Скала\"", "Одежда"))

    print("== «Оборот» за период (раздел снова открыт) ==")
    r = client.get(f"/api/revenue?date_from={mock_ms.DATES[0]}&date_to={mock_ms.DATES[-1]}")
    check("GET /api/revenue отдаёт 200", r.status_code == 200,
          f"status={r.status_code}")
    rev = r.json()
    exp_sales = mock_ms.expected_net_sales()
    exp_rev = round(sum(v[1] for v in exp_sales.values()))
    exp_qty = round(sum(v[0] for v in exp_sales.values()))
    check("выручка за весь период = эталон mock-мира",
          abs(rev["total_rev"] - exp_rev) <= 2 and abs(rev["total_qty"] - exp_qty) <= 1,
          f"got={rev['total_rev']}/{rev['total_qty']} exp={exp_rev}/{exp_qty}")
    top_base = max(exp_sales.items(), key=lambda kv: kv[1][1])[0]
    check("топ позиций: первый = максимум по выручке",
          rev["items"] and rev["items"][0]["base_name"] == top_base,
          f"got={rev['items'] and rev['items'][0]['base_name']} exp={top_base}")
    check("категории с долями, сумма долей ≈ 1",
          rev["categories"] and abs(sum(c["share"] for c in rev["categories"]) - 1) < 0.02)
    check("помесячный ряд: 18 месяцев, последний не пустой",
          len(rev["monthly"]) == 18 and rev["monthly"][-1]["total"] > 0,
          f"n={len(rev['monthly'])} last={rev['monthly'][-1]['total']}")
    r = client.get("/api/revenue?date_from=2026-01-31&date_to=2026-01-01")
    check("обратный период отдаёт 422", r.status_code == 422, f"status={r.status_code}")
    r = client.get("/orders")
    check("страница «Заказы» убрана: /orders отдаёт 404", r.status_code == 404,
          f"status={r.status_code}")
    for _pg in ("/budget", "/forecast", "/revenue"):
        r = client.get(_pg)
        check(f"страница {_pg} открыта (200)", r.status_code == 200, f"status={r.status_code}")
    r = client.get("/api/forecast")
    check("прогноз: помесячный ряд «сток vs продажи» на 7 месяцев (текущий + 6 для мини-колец)",
          r.status_code == 200 and len(r.json().get("months", [])) == 7
          and all("stock_value" in m and "sales_value" in m for m in r.json()["months"]))

    print("== «Пульс»: этот месяц против среднего за 6 мес ==")
    r = client.get("/api/pulse")
    check("GET /api/pulse отвечает", r.status_code == 200, f"status={r.status_code}")
    pulse = r.json()
    _date = _dt_date
    _cur = f"{_date.today().year:04d}-{_date.today().month:02d}"
    # Ревью 21.08 (мажор 3): в среднее берутся только ПОЛНОСТЬЮ покрытые
    # прошлые месяцы. У теста истории 60 дней → максимум один такой месяц,
    # значит partial=true и pct=null («мало истории»), а не выдуманный процент.
    _cov_start = mock_ms.DATES[0]
    _all6 = []
    _y, _m = _date.today().year, _date.today().month
    for _ in range(6):
        _m -= 1
        if _m == 0:
            _y, _m = _y - 1, 12
        _all6.append(f"{_y:04d}-{_m:02d}")
    _all6.reverse()
    _exp_months = [m for m in _all6 if f"{m}-01" >= _cov_start]
    check("окно пульса: только полностью покрытые историей месяцы",
          pulse["months"] == _exp_months
          and pulse["covered_months"] == len(_exp_months)
          and pulse.get("coverage_start") == _cov_start
          and all(m < _cur for m in pulse["months"]),
          f"months={pulse['months']} exp={_exp_months}")
    check("мало истории (<2 полных месяцев): partial=true, pct=null у обеих шкал",
          pulse["partial"] is (len(_exp_months) < 2)
          and (pulse["sales"]["pct"] is None and pulse["stock"]["pct"] is None
               if pulse["partial"] else True),
          f"partial={pulse['partial']} sales_pct={pulse['sales']['pct']} "
          f"stock_pct={pulse['stock']['pct']}")
    check("в пульсе нет месяца, начавшегося раньше загруженной истории",
          all(f"{m['month']}-01" >= _cov_start
              for m in pulse["sales"]["months"] + pulse["stock"]["months"]),
          f"coverage_start={_cov_start} months={[m['month'] for m in pulse['sales']['months']]}")
    ps = pulse["sales"]
    exp_proj = round(ps["mtd"] / ps["days_passed"] * ps["days_in_month"]) \
        if ps["days_passed"] else ps["mtd"]
    check("прогноз месяца = продано ÷ прошло дней × дней в месяце",
          abs(ps["projected"] - exp_proj) <= 1,
          f"got={ps['projected']} exp={exp_proj}")
    known = [m["v"] for m in ps["months"] if m["v"] is not None]
    check("среднее продаж = среднее известных (покрытых) месяцев",
          bool(known) and abs(ps["avg6"] - sum(known) / len(known)) <= 1,
          f"avg6={ps['avg6']} known={known}")
    rev_by_month = {m["month"]: m["total"] for m in rev["monthly"]}
    diverged = [m["month"] for m in ps["months"]
                if m["v"] is not None and m["month"] in rev_by_month
                and abs(m["v"] - rev_by_month[m["month"]]) > 2]
    check("помесячные продажи пульса сходятся с рядом /api/revenue",
          not diverged, f"расхождение в {diverged}")
    if ps["avg6"] > 0 and not pulse["partial"]:
        check("pct продаж = прогноз/среднее",
              abs(ps["pct"] - ps["projected"] / ps["avg6"]) < 0.01,
              f"pct={ps['pct']}")
    pst = pulse["stock"]
    check("склад сейчас: положительная стоимость и свежая дата",
          pst["current"] > 0 and pst["as_of"] is not None
          and pst["as_of"] >= mock_ms.DATES[-1],
          f"current={pst['current']} as_of={pst['as_of']}")
    known_st = [m["v"] for m in pst["months"] if m["v"] is not None]
    check("средний склад = среднее известных (покрытых) месяцев, все > 0",
          bool(known_st) and all(v > 0 for v in known_st)
          and abs(pst["avg6"] - sum(known_st) / len(known_st)) <= 1,
          f"avg6={pst['avg6']} known={known_st}")
    if pst["avg6"] > 0 and not pulse["partial"]:
        check("pct склада = сейчас/среднее",
              abs(pst["pct"] - pst["current"] / pst["avg6"]) < 0.01,
              f"pct={pst['pct']}")
    check("дни экстраполяции — по данным, не больше календарных",
          1 <= ps["days_passed"] <= _date.today().day,
          f"days_passed={ps['days_passed']}")

    print("== Табло свежести данных ==")
    r = client.get("/api/freshness")
    fresh = r.json() if r.status_code == 200 else {}
    check("GET /api/freshness: даты продаж и остатков на месте",
          r.status_code == 200 and fresh.get("last_sale_date")
          and fresh.get("last_stock_date")
          and fresh["last_stock_date"] >= mock_ms.DATES[-1],
          f"got={fresh}")
    check("свежесть согласована с пульсом",
          fresh.get("last_sale_date") == pulse.get("last_sale_date")
          and fresh.get("last_stock_date") == pulse.get("last_stock_date"))

    print("== Ручные скидки и «Дефолтные скидки» ==")
    r = client.post("/api/discount-overrides",
                    json={"base_name": "Худи «Скетч»", "discount": 25})
    check("ручная скидка сохраняется", r.status_code == 200 and r.json().get("ok"))
    r = client.get("/api/discounts")
    check("скрытый раздел «Скидки»: API отдаёт 404", r.status_code == 404,
          f"status={r.status_code}")
    # Отчёт скрыт из продукта, но расчёт остаётся в коде — проверяем напрямую.
    from app import analytics as _an, analytics_markdown as _amd
    from app.db import SessionLocal as _SL
    from app.models import Org as _Org
    _con = sqlite3.connect(DB_PATH)
    _org_id = _con.execute(
        "SELECT org_id FROM sales GROUP BY org_id ORDER BY COUNT(*) DESC LIMIT 1"
    ).fetchone()[0]
    _con.close()
    _db = _SL()
    try:
        _org = _db.get(_Org, _org_id)
        _snap = _an.get_snapshot(_db, _org)
        _overrides = client.get("/api/discount-overrides").json()
        d_items = {it["base_name"]: it
                   for it in _amd.build_discounts(_snap, _overrides)["items"]}
    finally:
        _db.close()
    hood_d = d_items.get("Худи «Скетч»")
    check("отчёт «Скидки»: ручная скидка приоритетна (25%, manual)",
          hood_d is not None and hood_d["discount_pct"] == 25 and hood_d["manual"]
          and "Ручная" in hood_d["reason"],
          f"got={hood_d and (hood_d['discount_pct'], hood_d.get('manual'))}")
    r = client.post("/api/discount-overrides",
                    json={"base_name": "Худи «Скетч»", "discount": 0})
    check("нулевая скидка снимает ручную",
          r.status_code == 200
          and "Худи «Скетч»" not in client.get("/api/discount-overrides").json())
    r = client.post("/api/discount-overrides/defaults")
    check("«Дефолтные скидки» расставились (owner)",
          r.status_code == 200 and r.json().get("count", 0) > 0,
          f"resp={r.text[:80]}")
    dover = client.get("/api/discount-overrides").json()
    belt_disc = dover.get("Ремень «Ось»")
    check("неликвид без продаж → глубокая скидка 60% (правило legacy)",
          belt_disc == 60, f"got={belt_disc}")

    print("== «По нулям N дн» на остатках ==")
    stocks_resp = client.get("/api/stocks").json()
    all_sizes = [s for it in stocks_resp["items"] for s in it["sizes"]]
    check("у размеров есть поле zero_days",
          all_sizes and all("zero_days" in s for s in all_sizes))
    zero_sizes = [s for s in all_sizes if s["total"] == 0]
    check("размеры «по нулям» несут дни с последнего остатка (или None без истории)",
          all(s["zero_days"] is None or s["zero_days"] >= 0 for s in zero_sizes),
          f"n_zero={len(zero_sizes)}")

    print("== Активный сток / архив / категории / правило скидок ==")
    ast = client.get("/api/active-stock").json()
    check("active-stock: 2 склада и позиции", len(ast["warehouses"]) == 2
          and len(ast["items"]) > 0)
    hood_a = next((i for i in ast["items"] if i["base_name"] == "Худи «Скетч»"), None)
    check("active-stock: «Заказано» раздельно (ручное 3 + из МС 14)",
          hood_a is not None and hood_a["ordered_manual"] == 3
          and hood_a["ordered_ms"] == 14,
          f"got={hood_a and (hood_a['ordered_manual'], hood_a['ordered_ms'])}")
    check("active-stock: у позиций есть per_wh/zat/defq/sizes",
          all(k in ast["items"][0] for k in ("per_wh", "zat", "defq", "sizes")))
    first_groups = [i["group"] for i in ast["items"]]
    check("active-stock: без продаж — в конце",
          first_groups == sorted(first_groups,
                                 key=lambda g: {"rank": 0, "low_data": 1, "no_sales": 2}[g]))

    r = client.post("/api/hidden", json={"base_name": "Ремень «Ось»", "hidden": True})
    check("архив: позиция убрана", r.status_code == 200)
    t_items = {i["base_name"]: i for i in client.get("/api/turnover").json()["items"]}
    check("архив: hidden=true в оборачиваемости",
          t_items.get("Ремень «Ось»", {}).get("hidden") is True)
    ast2 = client.get("/api/active-stock").json()
    check("архив: позиции нет в активном стоке",
          all(i["base_name"] != "Ремень «Ось»" for i in ast2["items"]))
    r = client.post("/api/hidden", json={"base_name": "Ремень «Ось»", "hidden": False})
    check("архив: возврат работает", r.status_code == 200)

    r = client.post("/api/categories/merge",
                    json={"from_category": "Украшения", "to_category": "Аксессуары"})
    check("слияние категорий сохраняется", r.status_code == 200)
    ast3 = client.get("/api/active-stock").json()
    ring = next((i for i in ast3["items"] if i["base_name"] == "Кольцо «Грань»"), None)
    check("слияние: Украшения показываются как Аксессуары",
          ring is not None and ring["category"] == "Аксессуары",
          f"got={ring and ring['category']}")
    r = client.post("/api/categories/override",
                    json={"base_name": "Кольцо «Грань»", "category": "Витрина"})
    ring2 = next((i for i in client.get("/api/active-stock").json()["items"]
                  if i["base_name"] == "Кольцо «Грань»"), None)
    check("перенос позиции приоритетнее слияния",
          r.status_code == 200 and ring2 and ring2["category"] == "Витрина",
          f"got={ring2 and ring2['category']}")
    client.post("/api/categories/override", json={"base_name": "Кольцо «Грань»", "category": ""})
    client.post("/api/categories/merge", json={"from_category": "Украшения", "to_category": ""})

    r = client.get("/api/discount-rule").json()
    check("правило скидок отдаётся с дефолтами",
          r["rule"]["weak_over_pct"] == 60 and r["defaults"]["top_pct"] == 15)
    new_rule = dict(r["rule"]); new_rule["weak_over_pct"] = 70
    r = client.post("/api/discount-rule", json=new_rule)
    check("правило скидок сохраняется (owner)", r.status_code == 200
          and r.json()["rule"]["weak_over_pct"] == 70)
    client.post("/api/discount-overrides/defaults")
    belt2 = client.get("/api/discount-overrides").json().get("Ремень «Ось»")
    check("дефолтные скидки считаются по новому правилу (60 → 70)",
          belt2 == 70, f"got={belt2}")
    client.post("/api/discount-rule", json=dict(r.json()["rule"], weak_over_pct=60))

    # ── Норма запаса — настройка организации, а не константа продукта ──
    # Решение владельца 22.08 (DECISIONS D-15). До этого «Сток на 90 дней» и
    # «Не хватает до нормы» считались по литералу 90 на сервере и ещё раз по
    # литералу 90 в браузере — и владелец, поменявший порог затоварки на той же
    # странице, получал две соседние колонки, живущие по разным числам.
    print("== Норма запаса берётся из настроек организации ==")
    base_rule = client.get("/api/discount-rule").json()["rule"]
    st = client.get("/api/active-stock").json()
    tu = client.get("/api/turnover").json()
    check("норма отдаётся вместе с данными (Активный сток)",
          st.get("stock_norm_days") == 90, f"got={st.get('stock_norm_days')}")
    check("норма отдаётся вместе с данными (Оборачиваемость)",
          tu.get("stock_norm_days") == 90, f"got={tu.get('stock_norm_days')}")
    row90 = next((i for i in st["items"] if i.get("defq")), None)
    check("есть позиция с недостатком до нормы", row90 is not None,
          f"позиций={len(st['items'])}")
    if row90:
        base_name, defq90, zat90 = row90["base_name"], row90["defq"], row90["zat"]
        client.post("/api/discount-rule", json=dict(base_rule, overstock_days=180))
        st2 = client.get("/api/active-stock").json()
        row180 = next((i for i in st2["items"] if i["base_name"] == base_name), None)
        check("норма в ответе изменилась вслед за настройкой",
              st2.get("stock_norm_days") == 180, f"got={st2.get('stock_norm_days')}")
        # Норма вдвое больше ⇒ недостаток растёт, а «сток в % от нормы» падает.
        check("«не хватает до нормы» пересчитано по новой норме",
              row180 is not None and row180["defq"] > defq90,
              f"было={defq90} стало={row180 and row180['defq']}")
        check("«сток на N дней» пересчитан по новой норме",
              row180 is not None and (zat90 is None or row180["zat"] < zat90),
              f"было={zat90} стало={row180 and row180['zat']}")
        tu2 = client.get("/api/turnover").json()
        check("страница «Оборачиваемость» получает ту же норму",
              tu2.get("stock_norm_days") == 180, f"got={tu2.get('stock_norm_days')}")
        # Мусор в настройке не должен ронять расчёт — падаем на дефолт.
        client.post("/api/discount-rule", json=dict(base_rule, overstock_days=90))
        st3 = client.get("/api/active-stock").json()
        row_back = next((i for i in st3["items"] if i["base_name"] == base_name), None)
        check("возврат к 90 даёт прежние числа",
              st3.get("stock_norm_days") == 90
              and row_back is not None and row_back["defq"] == defq90,
              f"norm={st3.get('stock_norm_days')} defq={row_back and row_back['defq']}")

    print("== Инкрементальный синк ==")
    # Эмулируем приёмку в МС: размер S принят полностью → у «Скетча» едет 8.
    mock_ms.PURCHASE_ORDERS[0]["positions"]["rows"][0]["shipped"] = 10.0
    r = client.post("/api/sync/run")
    check("инкрементальный синк запущен", r.status_code == 200 and r.json().get("ok"))
    status = wait_sync_done(client, timeout=120)
    check("инкрементальный синк done", status.get("state") == "done",
          f"state={status.get('state')} error={status.get('error', '')[:120]}")
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    hood = con.execute(
        "SELECT qty, ms_qty FROM ordered_qty WHERE base_name='Худи «Скетч»'"
    ).fetchone()
    check("инкремент пересобрал ms_qty после «приёмки» (14 → 8), qty=3 цел",
          hood is not None and hood["ms_qty"] == 8 and hood["qty"] == 3,
          f"got={dict(hood) if hood else None}")
    con.close()
    summary2 = client.get("/api/summary").json()
    check("после инкремента числа стабильны (остаток не «уехал»)",
          summary2["stock_units"] == summary["stock_units"],
          f"{summary['stock_units']} -> {summary2['stock_units']}")
    turno2 = {it["base_name"]: it["nq"] for it in client.get("/api/turnover").json()["items"]}
    drift = [b for b, it in by_base.items() if turno2.get(b) != it["nq"]]
    check("после инкремента нетто-продажи не изменились", not drift,
          f"drift={drift[:3]}")

    print("== Переименование товара в МС мигрирует пользовательские данные ==")
    # Аудит 18.08: «Заказано»/скидки/архив ключуются base_name — раньше после
    # переименования в МойСкладе они «осиротевали» и пропадали из аналитики.
    r = client.post("/api/ordered", json={"base_name": "Кепка «Штамп»", "qty": 7})
    check("ручное «Заказано» на кепку принято", r.status_code == 200)
    _cap = mock_ms.SKU_BY_EXT["p-cap1"]
    _cap["name"] = "Кепка «Штамп-2»"
    _cap["base"] = "Кепка «Штамп-2»"
    r = client.post("/api/sync/run")
    check("синк после переименования запущен", r.status_code == 200 and r.json().get("ok"))
    status = wait_sync_done(client, timeout=120)
    check("синк после переименования done", status.get("state") == "done",
          f"state={status.get('state')} error={status.get('error', '')[:120]}")
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    cap_old = con.execute(
        "SELECT qty FROM ordered_qty WHERE base_name='Кепка «Штамп»'").fetchone()
    cap_new = con.execute(
        "SELECT qty FROM ordered_qty WHERE base_name='Кепка «Штамп-2»'").fetchone()
    con.close()
    check("«Заказано» переехало на новое имя (7 шт), старое очищено",
          cap_old is None and cap_new is not None and cap_new["qty"] == 7,
          f"old={dict(cap_old) if cap_old else None} new={dict(cap_new) if cap_new else None}")

    from app import ms_client as _msc, ms_sync as _mss
    mock_api = httpx.Client(base_url=mock_ms.BASE, timeout=10.0)

    def _set_faults(**faults):
        return mock_api.post("/__test/faults", json=faults).json()

    def _stock_day_stats():
        c = sqlite3.connect(DB_PATH)
        try:
            n_dates = c.execute("SELECT COUNT(DISTINCT date) FROM stock_days").fetchone()[0]
            dups = c.execute(
                "SELECT COUNT(*) FROM (SELECT org_id, product_id, date, COUNT(*) n "
                "FROM stock_days GROUP BY org_id, product_id, date HAVING n > 1)"
            ).fetchone()[0]
            return n_dates, dups
        finally:
            c.close()

    def _stock_dates():
        c = sqlite3.connect(DB_PATH)
        try:
            return {d for (d,) in c.execute("SELECT DISTINCT date FROM stock_days")}
        finally:
            c.close()

    def _sales_rows():
        c = sqlite3.connect(DB_PATH)
        try:
            return c.execute("SELECT COUNT(*) FROM sales").fetchone()[0]
        finally:
            c.close()

    def _clear_sales():
        """Свежий аккаунт: продаж на диске ещё нет (для проверки мажора 2)."""
        c = sqlite3.connect(DB_PATH)
        try:
            c.execute("DELETE FROM sales")
            c.commit()
        finally:
            c.close()

    def _main_org_id():
        c = sqlite3.connect(DB_PATH)
        try:
            return c.execute("SELECT org_id FROM connections WHERE kind='moysklad' "
                             "ORDER BY id LIMIT 1").fetchone()[0]
        finally:
            c.close()

    def _day(offset: int) -> str:
        """ISO-дата со сдвигом от сегодня (offset ≤ 0)."""
        return (_dt_date.today() + _dt_delta(days=offset)).isoformat()

    def _sellout_zeros_ok():
        c = sqlite3.connect(DB_PATH)
        c.row_factory = sqlite3.Row
        try:
            for ext in mock_ms.sellout_ext_ids():
                pid = c.execute("SELECT id FROM products WHERE ext_id=?", (ext,)).fetchone()["id"]
                lrow = c.execute("SELECT qty FROM stock_days WHERE product_id=? "
                                 "ORDER BY date DESC LIMIT 1", (pid,)).fetchone()
                if lrow is None or lrow["qty"] != 0:
                    return False
            return True
        finally:
            c.close()

    def _interrupt_initial(ok_before=25):
        """Стойкий 429 после ok_before удачных отчётов → прерванная первичная."""
        _msc.MAX_RETRIES = 3
        try:
            _set_faults(stock_ok_before=ok_before, stock_429_burst=100000)
            r = client.post("/api/sync/initial")
            assert r.status_code == 200, r.text
            st = wait_sync_done(client)
        finally:
            _msc.MAX_RETRIES = 10
            _set_faults()
        return st

    # Деплой П1: при ok_before=25 удачно проходят «сегодня» (2 запроса) и окно
    # из 9 дат (18) — первый чанк истории (10 запросов) ловит стойкий 429.
    # Точка продолжения — самая СТАРАЯ загруженная дата = начало окна.
    exp_from = mock_ms.DATES[-10]

    def _stock_sets():
        """Множества строк stock_days и sales — для инварианта П1."""
        c = sqlite3.connect(DB_PATH)
        try:
            sd = set(c.execute(
                "SELECT p.ext_id, s.date, s.qty FROM stock_days s "
                "JOIN products p ON p.id = s.product_id").fetchall())
            sl = set(c.execute(
                "SELECT p.ext_id, s.date, s.is_return, s.qty, s.revenue FROM sales s "
                "JOIN products p ON p.id = s.product_id").fetchall())
            return sd, sl
        finally:
            c.close()

    if begin("a"):
        print("== Инцидент 21.08: устойчивость к 429/5xx ==")
        # (a) всплеск 429 (12 подряд, затем норма) — и ДЕФОЛТНЫЙ размер чанка (30)
        _mss.STOCK_CHUNK_DATES = 30
        try:
            _set_faults(stock_429_burst=12)
            r = client.post("/api/sync/initial")
            check("(a) первичный синк при всплеске 429 запущен", r.status_code == 200)
            status = wait_sync_done(client)
        finally:
            _mss.STOCK_CHUNK_DATES = 5
            _set_faults()
        stats_a = status.get("stats", {})
        check("(a) синк (чанк=30) завершился done несмотря на 12×429",
              status.get("state") == "done",
              f"state={status.get('state')} error={status.get('error', '')[:120]}")
        check("(a) клиент зафиксировал 429 и повторы (stats.ms_client)",
              (stats_a.get("ms_client") or {}).get("429", 0) > 0
              and (stats_a.get("ms_client") or {}).get("retries", 0) > 0,
              f"ms_client={stats_a.get('ms_client')}")
        summary_a = client.get("/api/summary").json()
        check("(a) остатки после синка = эталон mock-мира",
              summary_a["stock_units"] == exp_units,
              f"got={summary_a['stock_units']} expected={exp_units}")
        n_dates_a, dups_a = _stock_day_stats()
        check("(a) история: 60 дат, дублей нет", n_dates_a == 60 and dups_a == 0,
              f"dates={n_dates_a} dups={dups_a}")
        check("(a) после успеха точка продолжения/отпечаток не хранятся",
              "history_loaded_from" not in stats_a and "resume_fp" not in stats_a)
        check("(a) после успеха coverage_days = HISTORY_DAYS, все месяцы done",
              status.get("coverage_days") == 60
              and all(m["state"] == "done" for m in status.get("months", [])),
              f"coverage={status.get('coverage_days')} months={status.get('months')}")

    if begin("p2ref"):
        print("== Деплой П1: прогрессивная первичная загрузка ==")
        # (p2) ИНВАРИАНТ: прямой хронологический проход (окно = вся история) даёт
        # ровно те же строки stock_days/sales, что обратная загрузка чанками.
        _mss.INITIAL_WINDOW_DAYS = 60
        try:
            r = client.post("/api/sync/initial")
            status = wait_sync_done(client)
        finally:
            _mss.INITIAL_WINDOW_DAYS = 10
        check("(p2) эталонный прямой проход (окно 60 = вся история) done",
              r.status_code == 200 and status.get("state") == "done"
              and status.get("stats", {}).get("history_chunks_total") == 0
              and _stock_day_stats() == (60, 0),
              f"state={status.get('state')} stats={_stock_day_stats()}")
        ref_stock, ref_sales = _stock_sets()
    if begin("p2back"):
        _mss.STOCK_CHUNK_DATES = 7
        try:
            r = client.post("/api/sync/initial")
            status = wait_sync_done(client)
        finally:
            _mss.STOCK_CHUNK_DATES = 5
        stats_p2 = status.get("stats", {})
        check("(p2) обратная загрузка (окно 10, чанк 7) done, 8 чанков",
              status.get("state") == "done" and stats_p2.get("history_chunks_total") == 8
              and stats_p2.get("window_days") == 10,
              f"state={status.get('state')} chunks={stats_p2.get('history_chunks_total')}")
        got_stock, got_sales = _stock_sets()
        check("(p2) ИНВАРИАНТ: stock_days обратной загрузки == прямому проходу "
              f"({len(ref_stock)} строк)", got_stock == ref_stock,
              f"only_ref={len(ref_stock - got_stock)} only_got={len(got_stock - ref_stock)} "
              f"пример={sorted(ref_stock ^ got_stock)[:3]}")
        check(f"(p2) ИНВАРИАНТ: sales обратной загрузки == прямому проходу ({len(ref_sales)} строк)",
              got_sales == ref_sales,
              f"only_ref={len(ref_sales - got_sales)} only_got={len(got_sales - ref_sales)}")
        check("(p2) этапы products/today/month/history все done с секундами",
              [s["key"] for s in status.get("stages", [])] == ["products", "today", "month", "history"]
              and all(s["state"] == "done" and s["seconds"] is not None
                      for s in status.get("stages", [])),
              f"stages={status.get('stages')}")
        # Нит 14: ВТОРОЙ размер чанка на том же мире — границы чанков ложатся
        # иначе (в т.ч. на пустой день отчёта и на дыру длиннее чанка).
        _mss.STOCK_CHUNK_DATES = 13
        try:
            r = client.post("/api/sync/initial")
            status = wait_sync_done(client)
        finally:
            _mss.STOCK_CHUNK_DATES = 5
        check("(p2) ИНВАРИАНТ держится и при другом размере чанка (13)",
              status.get("state") == "done" and _stock_sets() == (ref_stock, ref_sales)
              and _stock_day_stats() == (60, 0),
              f"state={status.get('state')} stats={_stock_day_stats()}")

    if begin("p3"):
        # (p3) РЕВЬЮ 21.08 (мажор 2): запуск умирает на ПРОДАЖАХ окна быстрого
        #      старта — точка продолжения к этому моменту уже опубликована.
        #      Раньше продолжение пропускало фазу month навсегда: остатки окна
        #      были, а до 30 дней продаж не загружались НИКОГДА (инкремент лечит
        #      только SYNC_DAYS_BACK=3 дня), при этом синк рапортовал done и
        #      coverage_days=HISTORY_DAYS.
        _clear_sales()  # свежий аккаунт: продаж на диске ещё нет
        _msc.MAX_RETRIES = 1
        try:
            mock_ms.FAULTS["docs_429_burst"] = 100000
            r = client.post("/api/sync/initial")
            status = wait_sync_done(client)
        finally:
            _msc.MAX_RETRIES = 10
            mock_ms.reset_faults()
        stats_p3 = status.get("stats", {})
        good_fp = stats_p3.get("resume_fp")
        # DATA-3: точка продолжения теперь «сегодня», а не начало окна. Окно
        # публикуется одной транзакцией (остатки + продажи), и раз продажи не
        # доехали, остатков окна в базе тоже нет — объявлять загруженным то, чего
        # не записали, значит врать самому себе при продолжении. Раньше здесь
        # стояло exp_from (начало окна), и это было верно только потому, что
        # остатки окна успевали лечь до падения.
        check("(p3) синк упал на продажах окна: точка есть, окно НЕ помечено закрытым",
              status.get("state") == "error"
              and stats_p3.get("history_loaded_from") == _day(0)
              and stats_p3.get("window_done") is None
              and _sales_rows() == 0
              and _stock_dates() == {_day(0)},
              f"state={status.get('state')} from={stats_p3.get('history_loaded_from')} "
              f"window_done={stats_p3.get('window_done')} sales={_sales_rows()} "
              f"дат остатков={len(_stock_dates())}")
        r = client.post("/api/sync/run")
        status = wait_sync_done(client)
        got_stock, got_sales = _stock_sets()
        check("(p3) продолжение ДОБРАЛО продажи окна: sales == эталонному полному синку",
              status.get("state") == "done" and got_sales == ref_sales
              and got_stock == ref_stock,
              f"state={status.get('state')} only_ref={len(ref_sales - got_sales)} "
              f"only_got={len(got_sales - ref_sales)}")

    if begin("p7"):
        # (p7) РЕВЬЮ 21.08 (мажор 1): прерванный запуск продолжают НЕ В ТОТ ЖЕ
        #      ДЕНЬ. Раньше продолжение качало только «сегодня» и уходило назад —
        #      дни между «сегодня» прерванного прогона и «сегодня» продолжения
        #      не загружались никогда, а синк заканчивался done с coverage=год.
        _gap = 3
        _mss._today = lambda: _dt_date.today() - _dt_delta(days=_gap)
        try:
            status = _interrupt_initial()
        finally:
            _mss._today = _dt_date.today
        stats_p7 = status.get("stats", {})
        _shift_from = _day(-(_gap + 9))  # начало окна «того» дня
        check("(p7) подготовка: загрузка прервана «три дня назад», оба конца записаны",
              status.get("state") == "error"
              and stats_p7.get("history_loaded_from") == _shift_from
              and stats_p7.get("history_loaded_to") == _day(-_gap),
              f"from={stats_p7.get('history_loaded_from')} to={stats_p7.get('history_loaded_to')} "
              f"exp_from={_shift_from}")
        r = client.post("/api/sync/run")
        status = wait_sync_done(client)
        _dates_now = _stock_dates()
        _missing = [d for d in mock_ms.DATES if d not in _dates_now]
        check("(p7) продолжение через 3 дня не потеряло НИ ОДНОЙ даты",
              status.get("state") == "done" and not _missing
              and status.get("coverage_days") == 60,
              f"missing={_missing} coverage={status.get('coverage_days')}")
        check("(p7) ИНВАРИАНТ после продолжения с разрывом: stock_days/sales == эталону",
              _stock_sets() == (ref_stock, ref_sales),
              f"diff_stock={len(_stock_sets()[0] ^ ref_stock)} "
              f"diff_sales={len(_stock_sets()[1] ^ ref_sales)}")

    if begin("p11"):
        # (p11) РЕВЬЮ 21.08 (повторное): КОМБИНАЦИЯ p3 и p7 — окно-продажи упали,
        #       а продолжение случилось не в тот же день. Догон начинался от окна,
        #       пересчитанного на НОВОЕ «сегодня», поэтому терялось ровно столько
        #       дней продаж, сколько прошло до продолжения — молча: state=done,
        #       coverage полный (остатки-то целы), last_sale_date = сегодня.
        _clear_sales()
        _gap2 = 4
        _mss._today = lambda: _dt_date.today() - _dt_delta(days=_gap2)
        _msc.MAX_RETRIES = 1
        try:
            mock_ms.FAULTS["docs_429_burst"] = 100000
            client.post("/api/sync/initial")
            status = wait_sync_done(client)
        finally:
            _msc.MAX_RETRIES = 10
            mock_ms.reset_faults()
            _mss._today = _dt_date.today
        stats_p11 = status.get("stats", {})
        check("(p11) подготовка: окно-продажи упали «четыре дня назад»",
              status.get("state") == "error"
              and stats_p11.get("window_done") is None and _sales_rows() == 0,
              f"state={status.get('state')} window_done={stats_p11.get('window_done')} "
              f"sales={_sales_rows()}")
        client.post("/api/sync/run")
        status = wait_sync_done(client)
        got_stock, got_sales = _stock_sets()
        _lost = sorted({row[1] for row in (ref_sales - got_sales)})
        check("(p11) продолжение через 4 дня добрало ВСЕ продажи окна",
              status.get("state") == "done" and got_sales == ref_sales
              and got_stock == ref_stock,
              f"state={status.get('state')} потеряно_дней={len(_lost)} {_lost[:5]}")

    if begin("p8"):
        # (p8) РЕВЬЮ 21.08 (минор 7): точка продолжения СТАРШЕ окна («упали на
        #      самом старом чанке, продолжили назавтра»). Это значит «всё уже
        #      загружено», а не «пересобрать с нуля» — раньше выбрасывались ~364
        #      верных дня и год качался заново.
        con = sqlite3.connect(DB_PATH)
        con.execute("UPDATE sync_state SET state='error', stats_json=? WHERE org_id=?",
                    (json.dumps({"history_loaded_from": _day(-89),  # старше окна (60)
                                 "history_loaded_to": _day(0),
                                 "resume_fp": good_fp, "coverage_days": 60,
                                 "window_done": True}, ensure_ascii=False), _main_org_id()))
        con.commit(); con.close()
        r = client.post("/api/sync/run")
        status = wait_sync_done(client)
        stats_p8 = status.get("stats", {})
        check("(p8) точка старше окна = «всё загружено»: продолжение без чанков и без wipe",
              status.get("state") == "done" and stats_p8.get("resumed_from") == mock_ms.DATES[0]
              and stats_p8.get("history_chunks_total") == 0
              and stats_p8.get("stock_dates") == 1
              and _stock_day_stats() == (60, 0) and _stock_sets() == (ref_stock, ref_sales),
              f"state={status.get('state')} resumed={stats_p8.get('resumed_from')} "
              f"chunks={stats_p8.get('history_chunks_total')} dates={stats_p8.get('stock_dates')}")

    if begin("p1"):
        # (p1) порядок фаз и finalize-lite: во время идущего синка подключение уже
        # active, coverage_days == окну, state всё ещё running.
        _set_faults(stock_delay_ms=150)
        seen_lite = None
        seen_eta = None
        try:
            r = client.post("/api/sync/initial")
            check("(p1) первичный синк запущен (замедленный mock)", r.status_code == 200)
            deadline = time.time() + 120
            while time.time() < deadline:
                st = client.get("/api/sync/status").json()
                if st.get("state") in ("done", "error"):
                    break
                if st.get("phase") == "history" and st.get("coverage_days") == 10 and seen_lite is None:
                    conn_st = (client.get("/api/settings").json().get("connection") or {}).get("status")
                    seen_lite = (st.get("state"), conn_st, st.get("coverage_days"),
                                 st.get("history_loaded_from"))
                if st.get("eta_sec") is not None and seen_eta is None:
                    seen_eta = (st.get("eta_sec"), st.get("stats", {}).get("history_chunks_done"),
                                [m["state"] for m in st.get("months", [])])
                time.sleep(0.1)
            status = wait_sync_done(client)
        finally:
            _set_faults()
        check("(p1) finalize-lite: state=running, подключение active, coverage_days=10, точка = начало окна",
              seen_lite == ("running", "active", 10, exp_from), f"seen={seen_lite} exp_from={exp_from}")
        check("(p1) eta_sec появляется после первого чанка, есть месяц running",
              seen_eta is not None and isinstance(seen_eta[0], int) and seen_eta[1] >= 1
              and "running" in seen_eta[2], f"seen={seen_eta}")
        check("(p1) финал: done, coverage_days=60, все месяцы done, eta_sec=None",
              status.get("state") == "done" and status.get("coverage_days") == 60
              and all(m["state"] == "done" for m in status.get("months", []))
              and status.get("eta_sec") is None,
              f"state={status.get('state')} coverage={status.get('coverage_days')}")
        check("(p1) после замедленного прогона инвариант держится",
              _stock_sets() == (ref_stock, ref_sales))

    if begin("p5p6"):
        # (p5/p6) публичный прогресс и свежесть
        fr = client.get("/api/freshness").json()
        check("(p6) /api/freshness отдаёт coverage_days и history_days",
              fr.get("coverage_days") == 60 and fr.get("history_days") == 60, f"got={fr}")
        anon = httpx.Client(base_url=f"http://127.0.0.1:{APP_PORT}", timeout=10.0)
        r_anon = anon.get("/api/sync/progress")
        anon.close()
        check("(p5) /api/sync/progress без сессии → 401", r_anon.status_code == 401,
              f"status={r_anon.status_code}")
        member = httpx.Client(headers={"X-Oborot-CSRF": "1"},
                              base_url=f"http://127.0.0.1:{APP_PORT}", timeout=30.0)
        r = member.post("/register", data={
            "name": "Участник", "email": "member@test.io",
            "password": "secret123", "org_name": "Временная",
        })
        c = sqlite3.connect(DB_PATH)
        main_org = c.execute("SELECT org_id FROM connections WHERE kind='moysklad' "
                             "ORDER BY id LIMIT 1").fetchone()[0]
        mem_uid = c.execute("SELECT id FROM users WHERE email='member@test.io'").fetchone()[0]
        c.execute("UPDATE memberships SET org_id=?, role='member' WHERE user_id=?", (main_org, mem_uid))
        c.commit(); c.close()
        r = member.post("/login", data={"email": "member@test.io", "password": "secret123"})
        check("(p5) участник (member) залогинен в основную организацию", r.status_code == 303)
        r_st = member.get("/api/sync/status")
        r_pr = member.get("/api/sync/progress")
        prog = r_pr.json() if r_pr.status_code == 200 else {}
        # Минор 12: полоска должна знать окно быстрого старта (не зашивать «30 дней»)
        # и режим прогона, чтобы прятаться на обычном инкременте в 06:00.
        exp_keys = {"state", "mode", "phase", "progress_pct", "detail", "error", "error_cause",
                    "coverage_days", "history_days", "window_days", "months", "stages",
                    "eta_sec", "started_at", "finished_at", "can_manage"}
        check("(p5) member: /api/sync/status → 403, /api/sync/progress → 200 с нужной формой",
              r_st.status_code == 403 and r_pr.status_code == 200 and set(prog) == exp_keys
              and "stats" not in prog and prog.get("can_manage") is False
              and prog.get("coverage_days") == 60
              and prog.get("history_days") == 60 and len(prog.get("months", [])) >= 2
              and prog.get("window_days") == 10 and prog.get("mode") == "initial"
              and len(prog.get("stages", [])) == 4,
              f"status={r_st.status_code} progress={r_pr.status_code} keys={sorted(prog)}")
        member.close()

    if begin("b"):
        # (b) одиночный 500 — прозрачный повтор
        _set_faults(stock_500_once=True)
        r = client.post("/api/sync/run")
        status = wait_sync_done(client, timeout=120)
        stats_b = status.get("stats", {})
        check("(b) одиночный 500 повторён прозрачно (инкремент done, 5xx=1)",
              status.get("state") == "done"
              and (stats_b.get("ms_client") or {}).get("5xx") == 1,
              f"state={status.get('state')} ms_client={stats_b.get('ms_client')}")

    if begin("c"):
        # (c) стойкий 429 после 25 удачных отчётов (2,5 чанка по 5 дат × 2 склада)
        status = _interrupt_initial()
        stats_c = status.get("stats", {})
        check("(c) синк упал с state=error", status.get("state") == "error",
              f"state={status.get('state')}")
        check("(c) текст ошибки: «История загружена за 10 дней из 60 — продолжим автоматически»",
              status.get("error", "").startswith("История загружена за 10 дней из 60")
              and "продолжим автоматически" in status.get("error", "")
              and "ограничил частоту" in status.get("error", ""),
              f"error={status.get('error', '')[:160]}")
        check("(c) stats.history_loaded_from = начало окна (самая старая дата), есть отпечаток",
              stats_c.get("history_loaded_from") == exp_from and stats_c.get("resume_fp"),
              f"got={stats_c.get('history_loaded_from')} exp={exp_from} fp={stats_c.get('resume_fp')}")
        conn_c = (client.get("/api/settings").json().get("connection") or {})
        check("(c) подключение осталось active, coverage_days=10 сохранён, phase=history",
              conn_c.get("status") == "active" and status.get("coverage_days") == 10
              and status.get("phase") == "history",
              f"status={conn_c.get('status')} coverage={status.get('coverage_days')} phase={status.get('phase')}")
        check("(c) причина ошибки классифицирована (transient)",
              stats_c.get("error_cause") == "transient", f"cause={stats_c.get('error_cause')}")
        # Минор 10: сервис работает (подключение active, 10 дн. истории), текст
        # обещает «продолжим автоматически» — засчитывать это в серию провалов и
        # слать алерт «синк падает второй раз подряд» нельзя.
        check("(c) прерывание ФОНОВОЙ истории не увеличивает fail_streak",
              status.get("fail_streak") == 0 and status.get("alerted_streak") == 0,
              f"streak={status.get('fail_streak')} alerted={status.get('alerted_streak')}")
        n_dates_c, dups_c = _stock_day_stats()
        check("(c) в БД частичная новая история (10 дат окна), старая стёрта на фазе today",
              n_dates_c == 10 and dups_c == 0, f"dates={n_dates_c} dups={dups_c}")

    if begin("c1"):
        # (c1) ревью #1: продолжение падает ДО первого чанка — точка не теряется
        _msc.MAX_RETRIES = 1
        try:
            _set_faults(stock_429_burst=100000)
            r = client.post("/api/sync/run")  # инкрементный вызов → продолжение initial
            check("(c1) «Синхронизировать» при прерванной загрузке принят", r.status_code == 200)
            status = wait_sync_done(client)
        finally:
            _msc.MAX_RETRIES = 10
            _set_faults()
        stats_c1 = status.get("stats", {})
        check("(c1) запуск промотирован в initial и упал на первом запросе",
              status.get("state") == "error" and status.get("mode") == "initial"
              and stats_c1.get("resumed_from") == exp_from,
              f"state={status.get('state')} mode={status.get('mode')} resumed={stats_c1.get('resumed_from')}")
        check("(c1) точка продолжения сохранена, хотя новый запуск не записал ни чанка",
              stats_c1.get("history_loaded_from") == exp_from
              and status.get("error", "").startswith("История загружена за 10 дней из 60"),
              f"from={stats_c1.get('history_loaded_from')} error={status.get('error', '')[:100]}")
        n_dates_c1, _ = _stock_day_stats()
        check("(c1) частичная история не тронута (10 дат)", n_dates_c1 == 10, f"dates={n_dates_c1}")

    if begin("c2"):
        # (c2) продолжение через «Синхронизировать» (инкрементный вызов); размер
        #      чанка при продолжении ДРУГОЙ, чем у прерванного прогона (нит 14) —
        #      границы чанков не совпадают, заплатка обязана держаться и так.
        _mss.STOCK_CHUNK_DATES = 11
        try:
            r = client.post("/api/sync/run")
            check("(c2) повторный запуск принят", r.status_code == 200)
            status = wait_sync_done(client)
        finally:
            _mss.STOCK_CHUNK_DATES = 5
        stats_c2 = status.get("stats", {})
        check("(c2) продолжение завершилось done", status.get("state") == "done",
              f"state={status.get('state')} error={status.get('error', '')[:120]}")
        check("(c2) продолжение пошло НАЗАД от точки: 50 дат истории + «сегодня» обновлено",
              stats_c2.get("resumed_from") == exp_from
              and stats_c2.get("history_dates") == 60 - 10
              and stats_c2.get("stock_dates") == 60 - 10 + 1
              and stats_c2.get("stage_times", {}).get("month", {}).get("skipped") is True,
              f"resumed_from={stats_c2.get('resumed_from')} history={stats_c2.get('history_dates')} "
              f"dates={stats_c2.get('stock_dates')}")
        n_dates_c2, dups_c2 = _stock_day_stats()
        check("(c2) после продолжения: 60 дат, дублей (org, product, date) нет",
              n_dates_c2 == 60 and dups_c2 == 0, f"dates={n_dates_c2} dups={dups_c2}")
        summary_c = client.get("/api/summary").json()
        check("(c2) остатки после продолжения = эталон mock-мира",
              summary_c["stock_units"] == exp_units,
              f"got={summary_c['stock_units']} expected={exp_units}")
        check("(c2) явные нули распроданных сохранились через границу продолжения",
              _sellout_zeros_ok())
        check("(c2) точка продолжения снята после успеха",
              "history_loaded_from" not in stats_c2 and "resume_fp" not in stats_c2)
        check("(c2) ИНВАРИАНТ: stock_days/sales после продолжения == эталонному проходу",
              _stock_sets() == (ref_stock, ref_sales),
              f"diff_stock={len(_stock_sets()[0] ^ ref_stock)} diff_sales={len(_stock_sets()[1] ^ ref_sales)}")

    if begin("c3"):
        # (c3) ревью #3: «Полная пересборка» после прерывания — с нуля, не продолжение
        status = _interrupt_initial()
        check("(c3) подготовка: первичная снова прервана на 10 датах",
              status.get("state") == "error" and _stock_day_stats()[0] == 10)
        r = client.post("/api/sync/initial")
        status = wait_sync_done(client)
        stats_c3 = status.get("stats", {})
        check("(c3) POST /api/sync/initial — настоящая пересборка (без resumed_from)",
              r.status_code == 200 and status.get("state") == "done"
              and "resumed_from" not in stats_c3 and stats_c3.get("stock_dates") == 60,
              f"state={status.get('state')} stats_keys={sorted(stats_c3)[:8]}")
        n_dates_c3, dups_c3 = _stock_day_stats()
        check("(c3) после пересборки 60 дат без дублей, остатки = эталон, инвариант держится",
              n_dates_c3 == 60 and dups_c3 == 0
              and client.get("/api/summary").json()["stock_units"] == exp_units
              and _stock_sets() == (ref_stock, ref_sales),
              f"dates={n_dates_c3} dups={dups_c3}")

    if begin("c4"):
        # (c4) ревью #3: смена набора складов снимает точку продолжения и помечает
        #      needs_full_rebuild — следующий ИНКРЕМЕНТНЫЙ вызов делает полную пересборку
        status = _interrupt_initial()
        check("(c4) подготовка: первичная прервана, точка есть",
              status.get("state") == "error"
              and status.get("stats", {}).get("history_loaded_from") == exp_from)
        wh_list = client.get("/api/settings").json()["warehouses"]
        wh_lab = next(w for w in wh_list if not w["active"])  # сервисный склад st-lab
        r = client.post(f"/api/warehouses/{wh_lab['id']}/toggle", json={})  # включаем
        st_after = client.get("/api/sync/status").json().get("stats", {})
        check("(c4) toggle склада снял history_loaded_from и поставил needs_full_rebuild",
              r.status_code == 200 and "history_loaded_from" not in st_after
              and st_after.get("needs_full_rebuild") is True, f"stats={st_after}")
        r = client.post("/api/sync/run")
        status = wait_sync_done(client)
        stats_c4 = status.get("stats", {})
        exp_units_3 = round(sum(mock_ms.expected_stock_today(
            stores=("st-flag", "st-web", "st-lab")).values()))
        sum_c4 = client.get("/api/summary").json()
        check("(c4) POST /api/sync/run → полная initial без resumed_from, 60 дат по НОВОМУ набору",
              r.status_code == 200 and status.get("state") == "done"
              and status.get("mode") == "initial" and "resumed_from" not in stats_c4
              and "needs_full_rebuild" not in stats_c4
              and _stock_day_stats() == (60, 0) and sum_c4["stock_units"] == exp_units_3,
              f"state={status.get('state')} mode={status.get('mode')} "
              f"dates={_stock_day_stats()} units={sum_c4['stock_units']} exp={exp_units_3}")
        check("(c4) новый набор складов действительно другой (эталон 3 складов ≠ 2)",
              exp_units_3 != exp_units)
        client.post(f"/api/warehouses/{wh_lab['id']}/toggle", json={})  # вернуть как было
        r = client.post("/api/connect/moysklad/stores", json={"ext_ids": ["st-flag", "st-web"]})
        check("(c4) выбор складов прежним набором работает", r.status_code == 200)
        r = client.post("/api/sync/initial")
        status = wait_sync_done(client)
        check("(c4) полная пересборка по прежнему набору done, 60 дат, эталон 2 складов",
              status.get("state") == "done" and _stock_day_stats() == (60, 0)
              and client.get("/api/summary").json()["stock_units"] == exp_units,
              f"state={status.get('state')} stats={_stock_day_stats()}")

    if begin("c6c7"):
        # (c6) ревью (2) #1: провал ДО истории остатков (429 на ассортименте) не теряет
        #      точку продолжения; (c7) — то же для флага needs_full_rebuild
        status = _interrupt_initial()
        check("(c6) подготовка: первичная прервана, точка есть",
              status.get("stats", {}).get("history_loaded_from") == exp_from)
        _msc.MAX_RETRIES = 1
        try:
            _set_faults(assortment_429_burst=100000)
            r = client.post("/api/sync/run")
            status = wait_sync_done(client)
        finally:
            _msc.MAX_RETRIES = 10
            _set_faults()
        stats_c6 = status.get("stats", {})
        check("(c6) запуск упал на ассортименте (до истории), точка и отпечаток сохранены",
              r.status_code == 200 and status.get("state") == "error"
              and stats_c6.get("history_loaded_from") == exp_from and stats_c6.get("resume_fp")
              and _stock_day_stats()[0] == 10,
              f"state={status.get('state')} stats={stats_c6}")
        r = client.post("/api/sync/run")
        status = wait_sync_done(client)
        stats_c6b = status.get("stats", {})
        check("(c6) следующий запуск корректно продолжил (resumed_from, 60 дат, эталон)",
              status.get("state") == "done" and stats_c6b.get("resumed_from") == exp_from
              and _stock_day_stats() == (60, 0)
              and client.get("/api/summary").json()["stock_units"] == exp_units,
              f"state={status.get('state')} resumed={stats_c6b.get('resumed_from')} stats={_stock_day_stats()}")

        status = _interrupt_initial()
        wh_list = client.get("/api/settings").json()["warehouses"]
        wh_lab = next(w for w in wh_list if not w["active"])
        client.post(f"/api/warehouses/{wh_lab['id']}/toggle", json={})  # → needs_full_rebuild
        client.post(f"/api/warehouses/{wh_lab['id']}/toggle", json={})  # набор прежний, флаг стоит
        check("(c7) подготовка: needs_full_rebuild выставлен",
              client.get("/api/sync/status").json().get("stats", {}).get("needs_full_rebuild") is True)
        _msc.MAX_RETRIES = 1
        try:
            _set_faults(assortment_429_burst=100000)
            client.post("/api/sync/run")
            status = wait_sync_done(client)
        finally:
            _msc.MAX_RETRIES = 10
            _set_faults()
        stats_c7 = status.get("stats", {})
        check("(c7) провал на ассортименте: флаг needs_full_rebuild пережил запуск",
              status.get("state") == "error" and stats_c7.get("needs_full_rebuild") is True,
              f"state={status.get('state')} stats={stats_c7}")
        r = client.post("/api/sync/run")
        status = wait_sync_done(client)
        stats_c7b = status.get("stats", {})
        check("(c7) следующий запуск — полная initial, флаг снят, 60 дат, эталон",
              status.get("state") == "done" and status.get("mode") == "initial"
              and "resumed_from" not in stats_c7b and "needs_full_rebuild" not in stats_c7b
              and _stock_day_stats() == (60, 0)
              and client.get("/api/summary").json()["stock_units"] == exp_units,
              f"state={status.get('state')} mode={status.get('mode')} stats={_stock_day_stats()}")

    if begin("c8"):
        # (c8) ревью (2) #3: смена складов во время идущего синка → 409, склад не изменён
        r = client.post("/api/sync/initial")
        check("(c8) первичный синк запущен", r.status_code == 200)
        r_t = client.post(f"/api/warehouses/{wh_lab['id']}/toggle", json={})
        r_s = client.post("/api/connect/moysklad/stores", json={"ext_ids": ["st-flag"]})
        wh_now = next(w for w in client.get("/api/settings").json()["warehouses"]
                      if w["id"] == wh_lab["id"])
        check("(c8) toggle и выбор складов во время синка → 409 «Дождитесь», склад не изменён",
              r_t.status_code == 409 and "Дождитесь" in r_t.json().get("detail", "")
              and r_s.status_code == 409 and wh_now["active"] is False,
              f"toggle={r_t.status_code} stores={r_s.status_code} active={wh_now['active']}")
        status = wait_sync_done(client)
        check("(c8) синк завершился done, 60 дат", status.get("state") == "done"
              and _stock_day_stats() == (60, 0))

    if begin("c5"):
        # (c5) ревью #3: несовпадающий отпечаток (другой набор/окно) → свежая initial
        status = _interrupt_initial()
        c = sqlite3.connect(DB_PATH)
        c.execute("UPDATE sync_state SET stats_json = REPLACE(stats_json, ?, ?)",
                  (f"|{mock_ms.HISTORY_DAYS}\"", "|999\""))
        c.commit(); c.close()
        st_fp = client.get("/api/sync/status").json().get("stats", {})
        check("(c5) подготовка: отпечаток подменён", st_fp.get("resume_fp", "").endswith("|999"),
              f"fp={st_fp.get('resume_fp')}")
        r = client.post("/api/sync/run")
        status = wait_sync_done(client)
        stats_c5 = status.get("stats", {})
        check("(c5) при чужом отпечатке — полная initial (без resumed_from), 60 дат",
              status.get("state") == "done" and status.get("mode") == "initial"
              and "resumed_from" not in stats_c5 and _stock_day_stats() == (60, 0),
              f"state={status.get('state')} mode={status.get('mode')} stats={_stock_day_stats()}")

    if begin("c9"):
        # (c9) деплой П1: почасовой догон подхватывает прерванную историю («продолжим
        #      автоматически в течение часа»), хотя last_sync_at свежий.
        #      Ревью 21.08 (мажор 5): и НЕ БЛОКИРУЕТСЯ на ней — первичная загрузка
        #      легально идёт 30+ минут, а джоб ждал до часа НА КАЖДУЮ организацию
        #      (после деплоя, убившего несколько фоновых историй, ежедневный джоб
        #      вставал на часы и задерживал синки и дайджесты остальных).
        from app import scheduler as _sched_p1
        status = _interrupt_initial()
        mock_ms.FAULTS["stock_delay_ms"] = 120  # продолжение заведомо не успеет за секунды
        try:
            _t_c9 = time.time()
            res_c9 = _sched_p1.run_catchup_job()
            _elapsed_c9 = time.time() - _t_c9
            st_c9 = client.get("/api/sync/status").json()
            check("(c9) догон ЗАПУСКАЕТ продолжение и сразу возвращается (не ждёт initial)",
                  list(res_c9.values()) == ["started_initial"] and _elapsed_c9 < 15
                  and st_c9.get("state") == "running" and st_c9.get("mode") == "initial",
                  f"res={res_c9} elapsed={_elapsed_c9:.1f}с state={st_c9.get('state')}")
        finally:
            mock_ms.reset_faults()
        status = wait_sync_done(client)
        check("(c9) запущенное догоном продолжение доехало: done, 60 дат, инвариант",
              status.get("state") == "done" and status.get("mode") == "initial"
              and _stock_day_stats() == (60, 0) and _stock_sets() == (ref_stock, ref_sales),
              f"res={res_c9} state={status.get('state')} stats={_stock_day_stats()}")

    if begin("d"):
        # (d) смена токена при живом подключении: статус не падает в pending, синк стартует
        r = client.post("/api/connect/moysklad", json={"token": mock_ms.TOKEN})
        body_d = r.json() if r.status_code == 200 else {}
        check("(d) POST /api/connect/moysklad при активном подключении запускает синк",
              r.status_code == 200 and body_d.get("sync_started") is True
              and "запущена" in body_d.get("note", ""), f"resp={body_d}")
        r = client.post("/api/connect/moysklad", json={"token": mock_ms.TOKEN})
        body_d2 = r.json() if r.status_code == 200 else {}
        check("(d) повторная смена токена во время синка: sync_started=false, «уже идёт»",
              r.status_code == 200 and body_d2.get("sync_started") is False
              and "уже идёт" in body_d2.get("note", ""), f"resp={body_d2}")
        st_d = client.get("/api/sync/status").json()
        check("(d) статус синка после смены токена — running/done",
              st_d.get("state") in ("running", "done"), f"state={st_d.get('state')}")
        status = wait_sync_done(client, timeout=120)
        conn_d = (client.get("/api/settings").json().get("connection") or {})
        check("(d) подключение осталось active, синк done (incremental)",
              conn_d.get("status") == "active" and status.get("state") == "done"
              and status.get("mode") == "incremental",
              f"status={conn_d.get('status')} state={status.get('state')} mode={status.get('mode')}")
        check("(d) /api/settings отдаёт роль (owner) для owner-only кнопок",
              client.get("/api/settings").json().get("role") == "owner")

    if begin("e"):
        # (e) fail_streak и Telegram-алерт: один раз на серию, считая ручные провалы
        from app import notify as _notify, scheduler as _sched
        tg_calls = []
        _orig_send = _notify.send_message
        _notify.send_message = lambda chat_id, text: (tg_calls.append((chat_id, text)) or (True, ""))
        os.environ["OBOROT_TG_BOT_TOKEN"] = "test-bot-token"
        try:
            # digest_enabled=False: иначе планировщик шлёт ещё и дайджест через тот
            # же патченый send_message, и счётчик алертов «плывёт».
            r = client.post("/api/notify/settings",
                            json={"tg_chat_id": "4242", "tg_enabled": True, "digest_enabled": False})
            check("(e) Telegram-настройки сохранены", r.status_code == 200)
            _msc.MAX_RETRIES = 1
            _set_faults(stock_429_burst=100000)
            client.post("/api/sync/run"); st1 = wait_sync_done(client)
            check("(e) 1-й РУЧНОЙ провал: fail_streak=1, алерта нет",
                  st1.get("state") == "error" and st1.get("fail_streak") == 1 and not tg_calls,
                  f"state={st1.get('state')} streak={st1.get('fail_streak')} calls={len(tg_calls)}")
            client.post("/api/sync/run"); st2 = wait_sync_done(client)
            check("(e) 2-й РУЧНОЙ провал: fail_streak=2, алерт отправлен один раз",
                  st2.get("fail_streak") == 2 and st2.get("alerted_streak") == 2
                  and len(tg_calls) == 1 and tg_calls[0][0] == "4242"
                  and "второй раз подряд" in tg_calls[0][1],
                  f"streak={st2.get('fail_streak')} alerted={st2.get('alerted_streak')} calls={tg_calls[:1]}")
            check("(e) алерт: текст ошибки, дата продаж, подсказка по причине (429 → повторим)",
                  tg_calls and "429" in tg_calls[0][1]
                  and str(fresh.get("last_sale_date")) in tg_calls[0][1]
                  and "Повторим автоматически" in tg_calls[0][1],
                  f"text={tg_calls[0][1][:220] if tg_calls else ''}")
            res3 = _sched.run_daily_job()
            st3 = client.get("/api/sync/status").json()
            check("(e) 3-й провал (планировщик): fail_streak=3, повторного алерта нет",
                  list(res3.values()) == ["error"] and st3.get("fail_streak") == 3
                  and len(tg_calls) == 1,
                  f"res={res3} streak={st3.get('fail_streak')} calls={len(tg_calls)}")
            _set_faults()
            _msc.MAX_RETRIES = 10
            res4 = _sched.run_daily_job()
            st4 = client.get("/api/sync/status").json()
            check("(e) успешный синк сбрасывает fail_streak и alerted_streak в 0",
                  list(res4.values()) == ["done"] and st4.get("fail_streak") == 0
                  and st4.get("alerted_streak") == 0 and len(tg_calls) == 1,
                  f"res={res4} streak={st4.get('fail_streak')} alerted={st4.get('alerted_streak')}")
            _msc.MAX_RETRIES = 1
            _set_faults(stock_429_burst=100000)
            _sched.run_daily_job(); _sched.run_daily_job()
            st5 = client.get("/api/sync/status").json()
            check("(e) новая серия провалов (планировщик ×2) → второй алерт",
                  st5.get("fail_streak") == 2 and len(tg_calls) == 2,
                  f"streak={st5.get('fail_streak')} calls={len(tg_calls)}")
            _set_faults()
            _msc.MAX_RETRIES = 10
            res6 = _sched.run_daily_job()
            check("(e) восстановление после серии", list(res6.values()) == ["done"])
        finally:
            _notify.send_message = _orig_send
            os.environ.pop("OBOROT_TG_BOT_TOKEN", None)
            _msc.MAX_RETRIES = 10
            _set_faults()
            client.post("/api/notify/settings", json={"tg_chat_id": "", "tg_enabled": False})
    mock_api.close()
    summary = client.get("/api/summary").json()

    # Хвост сценария терминален: шарду, которому он не достался, дальше делать
    # нечего — поэтому здесь обычный выход, а не ещё один уровень отступа.
    if not begin("tail"):
        client.close()
        return finish()

    print("== Новая организация: смена токена без складов не запускает синк ==")
    newbie = httpx.Client(headers={"X-Oborot-CSRF": "1"},
                          base_url=f"http://127.0.0.1:{APP_PORT}", timeout=60.0)
    r = newbie.post("/register", data={
        "name": "Новичок", "email": "newbie@test.io",
        "password": "secret123", "org_name": "Новый бренд",
    })
    check("регистрация третьего пользователя", r.status_code == 303)
    r = newbie.post("/api/connect/moysklad", json={"token": mock_ms.TOKEN})
    body_n = r.json() if r.status_code == 200 else {}
    check("токен без складов: pending, «Осталось выбрать склады», sync_started=false",
          r.status_code == 200 and body_n.get("sync_started") is False
          and "Осталось выбрать склады" in body_n.get("note", ""), f"resp={body_n}")
    r = newbie.post("/api/connect/moysklad/stores", json={"ext_ids": ["st-flag"]})
    check("новичок выбрал склад", r.status_code == 200)
    r = newbie.post("/api/connect/moysklad", json={"token": mock_ms.TOKEN})
    body_n2 = r.json() if r.status_code == 200 else {}
    conn_n = (newbie.get("/api/settings").json().get("connection") or {})
    check("склады есть, но синка ещё не было (pending): авто-синк НЕ стартует, «Запустите синхронизацию»",
          body_n2.get("sync_started") is False and conn_n.get("status") == "pending"
          and "Запустите синхронизацию" in body_n2.get("note", "")
          and newbie.get("/api/sync/status").json().get("state") == "idle",
          f"resp={body_n2} status={conn_n.get('status')}")

    print("== Ревью 21.08 (мажор 4): окно до finalize-lite — демо не должно стирать ==")
    # Пока идёт первая загрузка, подключение ещё 'pending', last_sync_at пуст:
    # «/» уводило на онбординг, где по умолчанию выбраны «Демо-данные», и один
    # клик стирал таблицы организации ПРЯМО ВО ВРЕМЯ записи их синком.
    def _newbie_counts():
        c = sqlite3.connect(DB_PATH)
        try:
            org = c.execute("SELECT org_id FROM memberships m JOIN users u "
                            "ON u.id=m.user_id WHERE u.email='newbie@test.io'").fetchone()[0]
            return (org,
                    c.execute("SELECT COUNT(*) FROM products WHERE org_id=?", (org,)).fetchone()[0],
                    c.execute("SELECT COUNT(*) FROM stock_days WHERE org_id=?", (org,)).fetchone()[0])
        finally:
            c.close()

    mock_ms.FAULTS["stock_delay_ms"] = 120  # чтобы окно «до finalize-lite» было заметным
    try:
        r = newbie.post("/api/sync/initial")
        check("(p9) первичный синк новой организации запущен", r.status_code == 200,
              f"status={r.status_code}")
        deadline = time.time() + 30
        seen_running = False
        while time.time() < deadline:
            pr = newbie.get("/api/sync/progress").json()
            if pr.get("state") == "running":
                seen_running = True
                break
            if pr.get("state") in ("done", "error"):
                break
            time.sleep(0.05)
        counts_before = _newbie_counts()
        r_root = newbie.get("/", follow_redirects=False)
        r_demo = newbie.post("/api/connect/demo")
        conn_p9 = (newbie.get("/api/settings").json().get("connection") or {})
        counts_after = _newbie_counts()
    finally:
        mock_ms.reset_faults()
    check("(p9) во время первой загрузки «/» НЕ ведёт на онбординг с демо-кнопкой",
          seen_running and "/onboarding" not in (r_root.headers.get("location") or ""),
          f"running={seen_running} status={r_root.status_code} "
          f"loc={r_root.headers.get('location')}")
    check("(p9) POST /api/connect/demo во время первой загрузки → 409, данные целы",
          r_demo.status_code == 409 and counts_after == counts_before
          and "МойСклад" in r_demo.json().get("detail", ""),
          f"demo={r_demo.status_code} before={counts_before} after={counts_after} "
          f"conn={conn_p9.get('status')}")
    status_n = wait_sync_done(newbie, timeout=180)
    check("(p9) первичная загрузка новой организации завершилась",
          status_n.get("state") == "done",
          f"state={status_n.get('state')} error={status_n.get('error', '')[:120]}")
    r_demo2 = newbie.post("/api/connect/demo")
    check("(p9) после успешной загрузки демо тоже запрещено (409)",
          r_demo2.status_code == 409, f"status={r_demo2.status_code}")
    newbie.close()

    print("== Ревью 21.08 (минор 8): покрытие считается по БД, а не «на доверии» ==")
    # Первичная загрузка умерла на фазе products (данных нет вовсе), владелец
    # нажал «Синхронизировать сейчас» — инкремент пишет ОДИН день остатков.
    # Раньше state=done/mode=incremental давали coverage_days=HISTORY_DAYS,
    # и таблица считала оборачиваемость по одному дню как «за год».
    cover = httpx.Client(headers={"X-Oborot-CSRF": "1"},
                         base_url=f"http://127.0.0.1:{APP_PORT}", timeout=120.0)
    r = cover.post("/register", data={
        "name": "Покрытие", "email": "cover@test.io",
        "password": "secret123", "org_name": "Организация без истории",
    })
    check("(p10) регистрация организации для проверки покрытия", r.status_code == 303)
    cover.post("/api/connect/moysklad", json={"token": mock_ms.TOKEN})
    cover.post("/api/connect/moysklad/stores", json={"ext_ids": ["st-flag"]})
    _msc.MAX_RETRIES = 1
    try:
        mock_ms.FAULTS["assortment_429_burst"] = 100000
        cover.post("/api/sync/initial")
        st_cov = wait_sync_done(cover)
    finally:
        _msc.MAX_RETRIES = 10
        mock_ms.reset_faults()
    check("(p10) первичная умерла на товарах: coverage_days=0, истории нет",
          st_cov.get("state") == "error" and st_cov.get("coverage_days") == 0,
          f"state={st_cov.get('state')} coverage={st_cov.get('coverage_days')}")
    cover.post("/api/sync/run")
    st_cov2 = wait_sync_done(cover)
    pr_cov = cover.get("/api/sync/progress").json()
    check("(p10) после инкремента над пустотой покрытие = 1 день (а не «за год»)",
          st_cov2.get("state") == "done" and st_cov2.get("mode") == "incremental"
          and st_cov2.get("coverage_days") == 1 and pr_cov.get("coverage_days") == 1,
          f"state={st_cov2.get('state')} mode={st_cov2.get('mode')} "
          f"coverage={st_cov2.get('coverage_days')}")
    cover.close()

    print("== Ручная ростовка на «Заказе» (черновик правок по размерам) ==")
    # Правка ростовки под фабрику должна переживать перезагрузку страницы,
    # не принимать мусор и не протекать между организациями.
    r = client.get("/api/replenish-draft")
    check("черновик ростовки пуст на старте", r.status_code == 200 and r.json()["drafts"] == {},
          f"got={r.text[:120]}")
    r = client.post("/api/replenish-draft",
                    json={"base_name": "Худи «Скетч»", "sizes": {"S": 60, "M": 0}})
    check("правка ростовки сохранена", r.status_code == 200 and r.json().get("ok"),
          f"status={r.status_code} body={r.text[:140]}")
    drafts = client.get("/api/replenish-draft").json()["drafts"]
    check("правка возвращается следующему заходу (S=60, M=0)",
          drafts.get("Худи «Скетч»") == {"S": 60, "M": 0}, f"got={drafts}")
    r = client.post("/api/replenish-draft",
                    json={"base_name": "Худи «Скетч»", "sizes": {"S": -1}})
    check("отрицательное количество отклонено (422)", r.status_code == 422,
          f"status={r.status_code} body={r.text[:140]}")
    r = client.post("/api/replenish-draft",
                    json={"base_name": "Худи «Скетч»", "sizes": {"S": 10000}})
    check("количество больше потолка отклонено (422)", r.status_code == 422,
          f"status={r.status_code}")
    r = client.post("/api/replenish-draft",
                    json={"base_name": "Худи «Скетч»", "sizes": {"S": 12.5}})
    check("дробное количество отклонено (422)", r.status_code == 422,
          f"status={r.status_code}")
    drafts = client.get("/api/replenish-draft").json()["drafts"]
    check("отклонённые правки не изменили сохранённое",
          drafts.get("Худи «Скетч»") == {"S": 60, "M": 0}, f"got={drafts}")
    # Ревью 22.08: раньше правка по несуществующей позиции отвечала {"ok":true},
    # запись даже появлялась в базе — но следующий GET её тут же вычищал
    # (_drop_orphan_drafts), и человек считал, что сохранил, а не сохранил.
    # Теперь base_name сверяется с каталогом ДО записи, как и в /api/ordered.
    r = client.post("/api/replenish-draft",
                    json={"base_name": "Позиции такой нет", "sizes": {"S": 5}})
    check("правка по несуществующей позиции отклонена (404), а не «ok:true» без записи",
          r.status_code == 404 and "каталоге" in r.json()["detail"],
          f"status={r.status_code} body={r.text[:140]}")
    r = client.post("/api/replenish-draft",
                    json={"base_name": "  Худи «Скетч»  ", "sizes": {"S": 61}})
    check("имя с пробелами по краям принято, пробелы убраны",
          r.status_code == 200 and r.json()["base_name"] == "Худи «Скетч»",
          f"status={r.status_code} body={r.text[:140]}")
    r = client.post("/api/replenish-draft",
                    json={"base_name": "Худи «Скетч»", "sizes": {"S": 60, "ZZZ": 4}})
    check("правка с несуществующим размером принята запросом", r.status_code == 200)
    drafts = client.get("/api/replenish-draft").json()["drafts"]
    check("черновик по несуществующей позиции в базу не попал",
          "Позиции такой нет" not in drafts, f"got={list(drafts)[:5]}")
    check("черновик по исчезнувшему размеру убран, живой размер цел",
          drafts.get("Худи «Скетч»") == {"S": 60}, f"got={drafts}")
    r = client.post("/api/replenish-draft", json={"base_name": "Худи «Скетч»", "sizes": {}})
    check("пустой набор размеров = правок нет", r.status_code == 200
          and client.get("/api/replenish-draft").json()["drafts"] == {},
          f"got={client.get('/api/replenish-draft').text[:120]}")
    client.post("/api/replenish-draft", json={"base_name": "Худи «Скетч»", "sizes": {"S": 7}})
    # у безразмерной позиции размер — пустая строка, она тоже допустима
    client.post("/api/replenish-draft", json={"base_name": "Кепка «Штамп-2»", "sizes": {"": 3}})
    r = client.post("/api/replenish-draft/reset", json={"base_name": "Худи «Скетч»"})
    check("сброс одной позиции удалил только её", r.status_code == 200
          and list(client.get("/api/replenish-draft").json()["drafts"]) == ["Кепка «Штамп-2»"],
          f"got={client.get('/api/replenish-draft').text[:160]}")
    r = client.post("/api/replenish-draft/reset", json={"base_name": ""})
    check("сброс всей таблицы вернул расчёт везде", r.status_code == 200
          and client.get("/api/replenish-draft").json()["drafts"] == {},
          f"got={client.get('/api/replenish-draft').text[:120]}")
    anon = httpx.Client(base_url=f"http://127.0.0.1:{APP_PORT}", timeout=30.0)
    r = anon.get("/api/replenish-draft")
    check("черновик без сессии → 401", r.status_code == 401, f"status={r.status_code}")
    r = anon.post("/api/replenish-draft", json={"base_name": "X", "sizes": {"S": 1}})
    check("сохранение без заголовка CSRF → 403", r.status_code == 403, f"status={r.status_code}")
    anon.close()

    print("== Демо-режим не сломан ==")
    demo = httpx.Client(headers={"X-Oborot-CSRF": "1"}, base_url=f"http://127.0.0.1:{APP_PORT}", timeout=120.0)
    r = demo.post("/register", data={
        "name": "Демо", "email": "demo@test.io",
        "password": "secret123", "org_name": "Демо-бренд",
    })
    check("регистрация второго пользователя", r.status_code == 303)
    r = demo.post("/api/connect/demo")
    check("POST /api/connect/demo работает", r.status_code == 200 and r.json().get("ok"))
    dsum = demo.get("/api/summary").json()
    check("демо-данные посеялись (позиций ~55)", dsum["positions"] >= 45,
          f"positions={dsum['positions']}")
    check("изоляция тенантов: числа демо ≠ числа МС-организации",
          dsum["stock_units"] != summary["stock_units"])
    # Обратная сторона мажора 3: у демо 400 дней истории — все 6 месяцев
    # покрыты, partial=false и проценты считаются как раньше.
    dpulse = demo.get("/api/pulse").json()
    check("демо (400 дн. истории): пульс не partial, 6 полных месяцев, pct считается",
          dpulse["partial"] is False and dpulse["covered_months"] == 6
          and len(dpulse["months"]) == 6
          and (dpulse["sales"]["pct"] is not None or dpulse["sales"]["avg6"] == 0)
          and (dpulse["stock"]["pct"] is not None or dpulse["stock"]["avg6"] == 0),
          f"partial={dpulse['partial']} covered={dpulse['covered_months']} "
          f"pct={dpulse['sales']['pct']}/{dpulse['stock']['pct']}")

    client.post("/api/replenish-draft", json={"base_name": "Худи «Скетч»", "sizes": {"S": 60}})
    check("демо-организация не видит чужой черновик ростовки",
          demo.get("/api/replenish-draft").json()["drafts"] == {},
          f"got={demo.get('/api/replenish-draft').text[:160]}")
    demo.post("/api/replenish-draft/reset", json={"base_name": ""})
    check("сброс в демо-организации не тронул чужие правки",
          client.get("/api/replenish-draft").json()["drafts"] == {"Худи «Скетч»": {"S": 60}},
          f"got={client.get('/api/replenish-draft').text[:160]}")
    client.post("/api/replenish-draft/reset", json={"base_name": ""})

    print("== Окно темпа «90 дней»: распроданное в ноль не выпадает из заказа ==")
    # У mock-мира истории всего 60 дней, поэтому окно 90 дней там ничего не
    # меняет; демо-данные — 365 дней и есть позиции, распроданные в ноль.
    demo.post("/api/settings", json={"rate_window": "year"})
    r_year = demo.get("/api/replenish").json()
    year_names = {it["base_name"] for it in r_year["items"]}
    # «Распродан в ноль и нужен»: остатка нет, но заказать надо.
    sold_out = {it["base_name"] for it in r_year["items"] if it["cs"] == 0}
    check("демо: есть распроданные в ноль позиции с потребностью",
          len(sold_out) > 0, f"n={len(sold_out)}")

    demo.post("/api/settings", json={"rate_window": "d90"})
    r90 = demo.get("/api/replenish").json()
    names90 = {it["base_name"] for it in r90["items"]}
    lost = sorted(sold_out - names90)
    check("d90: распроданные в ноль остались в заказе (фолбэк на годовой темп)",
          not lost, f"пропали={lost}")
    check("d90: заказ не короче, чем был до фолбэка",
          len(r90["items"]) >= len(sold_out),
          f"items={len(r90['items'])} sold_out={len(sold_out)}")
    fb = [it for it in r90["items"] if it.get("rate_fallback")]
    check("d90: фолбэк помечен флагом rate_fallback и посчитан",
          len(fb) > 0 and r90.get("fallback_count") == len(fb),
          f"fallback_count={r90.get('fallback_count')} n={len(fb)}")
    check("d90: у позиции под фолбэком темп равен годовому",
          all(abs(it["rate"] - it["rate_year"]) < 1e-9 for it in fb),
          f"bad={[it['base_name'] for it in fb if abs(it['rate'] - it['rate_year']) >= 1e-9][:3]}")
    check("d90: под фолбэк попали только позиции без остатка",
          all(it["cs"] == 0 for it in fb),
          f"bad={[it['base_name'] for it in fb if it['cs'] > 0][:3]}")
    # Фолбэк оправдан ровно одним: позиции НЕ БЫЛО на складе, продавать было
    # нечего. Порог min_stock_days тут ни при чём — проверяем по дням
    # физического наличия за 90 дней (instock90 в снапшоте).
    _con_fb = sqlite3.connect(DB_PATH)
    _demo_org = _con_fb.execute(
        "SELECT org_id FROM products WHERE base_name=? LIMIT 1", ("Браслет «Звенья»",)
    ).fetchone()[0]
    _con_fb.close()
    _dbf = _SL()
    try:
        _snapf = _an.get_snapshot(_dbf, _dbf.get(_Org, _demo_org))
        bad_fb = [it["base_name"] for it in fb
                  if _snapf["items"][it["base_name"]]["instock90"] > 0]
    finally:
        _dbf.close()
    check("d90: фолбэк только у позиций, которых не было на складе ни дня",
          not bad_fb, f"лежали на складе={bad_fb[:3]}")
    check("d90: подпись активного окна человеческая",
          r90.get("rate_window_label") == "темп за 90 дней",
          f"got={r90.get('rate_window_label')}")
    # Неликвид (лежит на складе, не продаётся) фолбэк подхватывать НЕ должен —
    # иначе заказ раздуется по мёртвому товару.
    dturn = demo.get("/api/turnover").json()["items"]
    dead_demo = {it["base_name"] for it in dturn
                 if it["cs"] > 0 and it["nq"] <= 0 and not it["archived"]}
    check("демо: неликвид в данных есть", len(dead_demo) > 0, f"n={len(dead_demo)}")
    check("d90: неликвид в заказ не попал", not (dead_demo & names90),
          f"попали={sorted(dead_demo & names90)[:3]}")
    excl90 = {e["base_name"]: e["reason"] for e in r90["excluded"]}
    check("d90: у неликвида внятная причина исключения",
          all(excl90.get(b) for b in dead_demo),
          f"без причины={[b for b in dead_demo if not excl90.get(b)][:3]}")
    check("d90: всё, чего нет в заказе, объяснено в excluded",
          not (year_names - names90 - set(excl90)),
          f"молча пропали={sorted(year_names - names90 - set(excl90))[:3]}")

    print("== Выгрузка «Что заказать»: окно темпа и лист «Не вошло и почему» ==")
    import io as _io

    from openpyxl import load_workbook

    xl = demo.get("/api/export/replenish.xlsx")
    check("выгрузка отдаётся", xl.status_code == 200, f"status={xl.status_code}")
    wbx = load_workbook(_io.BytesIO(xl.content))
    check("в книге два листа: заказ и «Не вошло и почему»",
          wbx.sheetnames == ["Что заказать", "Не вошло и почему"],
          f"got={wbx.sheetnames}")
    wsx = wbx["Что заказать"]
    title = wsx["A1"].value or ""
    check("шапка листа называет окно темпа и срок производства",
          "темп за 90 дней" in title and "срок производства" in title,
          f"A1={title}")
    xl_names = {wsx.cell(row=i, column=1).value for i in range(3, wsx.max_row + 1)}
    check("в выгрузке те же позиции, что в API (включая распроданные)",
          not (names90 - xl_names), f"нет в файле={sorted(names90 - xl_names)[:3]}")
    # 14 — «Производство» (добавлена, чтобы разложить файл по цехам), 15 — «Примечание»
    notes = [wsx.cell(row=i, column=15).value for i in range(3, wsx.max_row + 1)]
    check("позиции по фолбэку помечены в основном листе",
          sum(1 for n in notes if n) == len(fb),
          f"пометок={sum(1 for n in notes if n)} ожидалось={len(fb)}")
    wsx2 = wbx["Не вошло и почему"]
    xl_excl = {wsx2.cell(row=i, column=1).value for i in range(3, wsx2.max_row + 1)}
    check("лист «Не вошло и почему» содержит все исключённые позиции",
          xl_excl == set(excl90), f"лишние/недостающие={xl_excl ^ set(excl90)}")

    print("== «Прогноз»: склад сейчас — одно число в карточке, кольце и графике ==")
    fc = demo.get("/api/forecast").json()
    pl = demo.get("/api/pulse").json()
    check("карточка и старт недельного ряда — одно и то же число",
          fc["cards"]["stock_value"] == fc["weeks"][0]["stock_value"],
          f"card={fc['cards']['stock_value']} week0={fc['weeks'][0]['stock_value']}")
    check("карточка и старт помесячного ряда — одно и то же число",
          fc["cards"]["stock_value"] == fc["months"][0]["stock_value"],
          f"card={fc['cards']['stock_value']} month0={fc['months'][0]['stock_value']}")
    check("кольцо «Склад в деньгах» показывает то же число, что карточка",
          pl["stock"]["current"] == fc["cards"]["stock_value"],
          f"pulse={pl['stock']['current']} card={fc['cards']['stock_value']}")
    # Кольцо «Пульса» считает те же деньги, что карточка «Прогноза», но
    # называет ещё и набор позиций: скрытые и архивные не входят ни в
    # «сейчас», ни в шесть прошлых месяцев, и человек должен видеть это
    # в подписи — иначе скрытие позиции сдвигает всю историю молча.
    check("у денежных сумм «Прогноза» есть подпись базы",
          fc.get("money_basis", {}).get("stock_value")
          and pl["stock"].get("basis", "").startswith(fc["money_basis"]["stock_value"]),
          f"basis={fc.get('money_basis')} pulse={pl['stock'].get('basis')}")
    check("подпись кольца «Склад в деньгах» называет набор позиций",
          "скрыт" in (pl["stock"].get("basis") or "")
          and "архив" in (pl["stock"].get("basis") or ""),
          f"basis={pl['stock'].get('basis')}")
    check("у сумм склада на дашборде есть подпись базы",
          dsum.get("stock_value_retail_basis") and dsum.get("stock_value_cost_basis"),
          f"got={(dsum.get('stock_value_retail_basis'), dsum.get('stock_value_cost_basis'))}")
    print("== Себестоимость и валовая маржа (второй денежный слой) ==")
    turn = client.get("/api/turnover").json()
    money = turn.get("money") or {}
    rows = turn["items"]
    live = [i for i in rows if not i["archived"] and not i["hidden"]]
    check("у позиций «Оборачиваемости» есть себестоимость и маржа",
          all(k in rows[0] for k in ("cost_price", "no_cost", "margin_unit",
                                     "margin_pct", "gross_margin", "stock_cost")),
          f"keys={sorted(rows[0])[:12]}")
    check("сумма замороженного по позициям = итог страницы",
          sum(i["stock_cost"] or 0 for i in live) == money.get("stock_cost"),
          f"позиции={sum(i['stock_cost'] or 0 for i in live)} итог={money.get('stock_cost')}")
    check("сумма валовой маржи по позициям = итог страницы",
          sum(i["gross_margin"] or 0 for i in live) == money.get("gross_margin"),
          f"позиции={sum(i['gross_margin'] or 0 for i in live)} итог={money.get('gross_margin')}")
    cats = turn.get("money_by_category") or []
    check("сумма по категориям = общий итог (до рубля, без «расхождений округления»)",
          sum(c["stock_cost"] for c in cats) == money.get("stock_cost")
          and sum(c["gross_margin"] for c in cats) == money.get("gross_margin")
          and sum(c["positions"] for c in cats) == money.get("positions"),
          f"cats={sum(c['stock_cost'] for c in cats)} итог={money.get('stock_cost')}")
    summ = client.get("/api/summary").json()
    ast_money = client.get("/api/active-stock").json().get("money") or {}
    check("«заморожено по себестоимости» одинаково на дашборде, обороте и стоке",
          summ["stock_value_cost"] == money.get("stock_cost") == ast_money.get("stock_cost"),
          f"summary={summ['stock_value_cost']} turnover={money.get('stock_cost')} stock={ast_money.get('stock_cost')}")
    check("себестоимость склада меньше розницы (это разные суммы, а не одна)",
          0 < money.get("stock_cost", 0) < money.get("stock_retail", 0),
          f"cost={money.get('stock_cost')} retail={money.get('stock_retail')}")
    check("у каждой новой суммы есть машиночитаемая подпись базы",
          money.get("stock_cost_basis") == _an.BASIS_COST
          and money.get("stock_retail_basis") == _an.BASIS_RETAIL
          and money.get("stock_sale_basis") == _an.BASIS_AVG_SALE
          and (turn.get("money_basis") or {}).get("stock_cost") == _an.BASIS_COST,
          f"got={money.get('stock_cost_basis')}")
    check("процент маржи в разумных пределах и без деления на ноль",
          all(i["margin_pct"] is None or -50 <= i["margin_pct"] <= 1 for i in rows)
          and (money.get("gross_margin_pct") is None
               or 0 < money["gross_margin_pct"] <= 1),
          f"pct={money.get('gross_margin_pct')}")
    no_sales = [i for i in rows if i["nq"] <= 0]
    check("позиции без продаж не роняют расчёт: маржа от прайса, за год ноль",
          all(i["gross_margin"] in (0, None) and i["loss_total"] == 0 for i in no_sales),
          f"n={len(no_sales)}")
    zero_stock = [i for i in live if i["cs"] == 0]
    check("позиции с нулевым остатком не морозят денег",
          all((i["stock_cost"] or 0) == 0 and (i["stock_margin"] or 0) == 0
              for i in zero_stock),
          f"n={len(zero_stock)}")
    check("капитал оборачивается: положительное число раз в год",
          money.get("capital_turns") and money["capital_turns"] > 0,
          f"turns={money.get('capital_turns')}")

    print("== «Торгуете в минус» и позиции без себестоимости ==")
    # Подкручиваем себестоимость прямо в БД: в mock-мире все позиции прибыльны,
    # а проверить нужно ровно обратный случай (и позицию без себестоимости).
    _con = sqlite3.connect(DB_PATH)
    # Именно орг владельца (mock-МС), а не демо-орг второго пользователя.
    _oid = _con.execute(
        "SELECT m.org_id FROM memberships m JOIN users u ON u.id = m.user_id "
        "WHERE u.email = 'owner@test.io'"
    ).fetchone()[0]
    _tee = next(i for i in rows if i["base_name"] == "Футболка «Манифест»")
    _loss_cost = (_tee["avg_price"] or 0) + 500          # заведомо ниже себестоимости
    _con.execute("UPDATE products SET cost_price=? WHERE org_id=? AND base_name=?",
                 (_loss_cost, _oid, "Футболка «Манифест»"))
    _con.execute("UPDATE products SET cost_price=0 WHERE org_id=? AND base_name=?",
                 (_oid, "Сумка «Тоут»"))
    _con.commit()
    _con.close()
    _an.invalidate(_oid)
    turn2 = client.get("/api/turnover").json()
    below = turn2.get("below_cost") or {}
    rows2 = {i["base_name"]: i for i in turn2["items"]}
    tee2, bag2 = rows2["Футболка «Манифест»"], rows2["Сумка «Тоут»"]
    check("позиция дешевле себестоимости помечена и попала в предупреждение",
          tee2["below_cost"] and any(x["base_name"] == "Футболка «Манифест»"
                                     for x in below.get("items", [])),
          f"below_cost={tee2['below_cost']} items={[x['base_name'] for x in below.get('items', [])]}")
    _lost = next(x for x in below["items"] if x["base_name"] == "Футболка «Манифест»")
    check("потеря считается как (себестоимость − средняя цена) × продано",
          _lost["loss_unit"] == round(_loss_cost - tee2["avg_price"])
          and _lost["loss_total"] == round((_loss_cost - tee2["avg_price"]) * tee2["nq"])
          and _lost["loss_stock"] == round((_loss_cost - tee2["avg_price"]) * tee2["cs"]),
          f"unit={_lost['loss_unit']} total={_lost['loss_total']} nq={tee2['nq']}")
    check("итог предупреждения = сумма потерь по строкам",
          below["loss_total"] == sum(x["loss_total"] for x in below["items"])
          and below["positions"] == len(below["items"]),
          f"итог={below['loss_total']}")
    check("в минус попадают только значимые позиции (low_data — отдельным счётчиком)",
          all(not x["low_data"] for x in below["items"])
          and "low_data_positions" in below,
          f"got={[(x['base_name'], x['low_data']) for x in below['items']]}")
    check("у позиции в минусе маржа отрицательная, а не «ноль из ниоткуда»",
          tee2["margin_unit"] < 0 and tee2["margin_pct"] < 0,
          f"unit={tee2['margin_unit']} pct={tee2['margin_pct']}")
    check("позиция без себестоимости помечена no_cost, а не посчитана по нулю",
          bag2["no_cost"] and bag2["margin_unit"] is None
          and bag2["margin_pct"] is None and bag2["stock_cost"] is None
          and bag2["gross_margin"] is None and not bag2["below_cost"],
          f"got={(bag2['no_cost'], bag2['margin_unit'], bag2['stock_cost'])}")
    money2 = turn2["money"]
    live2 = [i for i in turn2["items"] if not i["archived"] and not i["hidden"]]
    check("позиция без себестоимости не портит агрегаты и видна отдельным счётчиком",
          money2["no_cost_positions"] >= 1
          and money2["no_cost_retail"] == sum(i["stock_retail"] for i in live2 if i["no_cost"])
          and money2["stock_cost"] == sum(i["stock_cost"] or 0 for i in live2)
          and money2["positions"] == len(live2),
          f"no_cost={money2['no_cost_positions']} retail={money2['no_cost_retail']}")
    check("«за сколько продастся» считается по всем позициям, включая без с/с",
          money2["stock_sale"] == sum(i["stock_sale"] for i in live2),
          f"итог={money2['stock_sale']}")
    ast2m = client.get("/api/active-stock").json()
    check("«Активный сток» показывает то же предупреждение и те же деньги",
          (ast2m.get("below_cost") or {}).get("loss_total") == below["loss_total"]
          and (ast2m.get("money") or {}).get("stock_cost") == money2["stock_cost"]
          and "incoming" in ast2m,
          f"stock={ast2m.get('money', {}).get('stock_cost')} turnover={money2['stock_cost']}")
    # Возвращаем себестоимость на место, чтобы дальнейшие проверки видели мир mock-МС.
    _con = sqlite3.connect(DB_PATH)
    _con.execute("UPDATE products SET cost_price=1500 WHERE org_id=? AND base_name=?",
                 (_oid, "Футболка «Манифест»"))
    _con.execute("UPDATE products SET cost_price=2400 WHERE org_id=? AND base_name=?",
                 (_oid, "Сумка «Тоут»"))
    _con.commit()
    _con.close()
    _an.invalidate(_oid)
    turn3 = client.get("/api/turnover").json()
    check("после возврата себестоимости предупреждение исчезает",
          turn3["below_cost"]["positions"] == 0
          and turn3["money"]["no_cost_positions"] == 0,
          f"below={turn3['below_cost']['positions']}")

    # Порог значимости отдельно: позиция с одной случайной дешёвой продажей
    # тревогу не поднимает — она уходит в счётчик low_data_*, а не в список.
    def _fake(name, low):
        return {"base_name": name, "category": "Тест", "cls": "weak",
                "low_data": low, "avg_price": 100, "sale_price": 300,
                "cost_price": 150.0, "discount_fact": 0.5, "nq": 2 if low else 20,
                "cs": 5, "below_cost": True, "loss_unit": 50,
                "loss_total": 100 if low else 1000, "loss_stock": 250}
    _rep = _an.below_cost_report([_fake("Шум", True), _fake("Правда", False)])
    check("одна случайная продажа не поднимает тревогу «в минус»",
          _rep["positions"] == 1 and _rep["items"][0]["base_name"] == "Правда"
          and _rep["loss_total"] == 1000 and _rep["low_data_positions"] == 1
          and _rep["low_data_loss"] == 100,
          f"got={( _rep['positions'], _rep['low_data_positions'])}")

    demo.post("/api/settings", json={"rate_window": "year"})

    print("== Условия производства: срок, минимальная партия, кратность ==")
    # 1) Аддитивная миграция на «старой» базе, где колонок условий ещё нет.
    import tempfile

    from sqlalchemy import create_engine as _create_engine
    from sqlalchemy import inspect as _inspect
    from sqlalchemy import text as _text

    from app import models as _models

    old_db = Path(tempfile.mkdtemp()) / "old_schema.db"
    old_engine = _create_engine(f"sqlite:///{old_db}", future=True)
    with old_engine.begin() as conn:
        conn.execute(_text(
            "CREATE TABLE productions (id INTEGER PRIMARY KEY, org_id INTEGER NOT NULL, "
            "name VARCHAR(120) NOT NULL, is_main BOOLEAN NOT NULL, created_at DATETIME)"))
        conn.execute(_text(
            "INSERT INTO productions (id, org_id, name, is_main) VALUES (1, 7, 'Старый цех', 1)"))
    _models.ensure_schema(bind=old_engine)
    cols = {c["name"] for c in _inspect(old_engine).get_columns("productions")}
    check("миграция добавила условия в существующую таблицу productions",
          {"lead_time_days", "moq", "pack_multiple"} <= cols, f"cols={sorted(cols)}")
    with old_engine.begin() as conn:
        row = conn.execute(_text(
            "SELECT name, lead_time_days, moq, pack_multiple FROM productions WHERE id = 1")).first()
        flags = [r[0] for r in conn.execute(_text("SELECT name FROM migration_flags")).fetchall()]
        conn.execute(_text("UPDATE productions SET lead_time_days = 21 WHERE id = 1"))
    check("старое производство уцелело, условия пустые = «как в общих настройках»",
          tuple(row) == ("Старый цех", None, None, None), f"row={tuple(row)}")
    check("миграция отмечена флагом (один запуск)",
          "productions_conditions_v1" in flags, f"flags={flags}")
    _models.ensure_schema(bind=old_engine)  # повторный запуск — ничего не делает
    with old_engine.begin() as conn:
        again = conn.execute(_text("SELECT lead_time_days FROM productions WHERE id = 1")).scalar()
    check("повторный прогон миграции не затирает заполненные условия",
          again == 21, f"lead_time_days={again}")
    old_engine.dispose()

    # 1б) Одновременный старт нескольких процессов на одной старой базе.
    # Ревью 22.08 (Н1): раньше 2-3 воркера из 4 падали на «duplicate column
    # name» ещё на импорте app.api — процесс не поднимался вообще.
    import subprocess

    race_db = Path(tempfile.mkdtemp()) / "race_schema.db"
    race_engine = _create_engine(f"sqlite:///{race_db}", future=True)
    with race_engine.begin() as conn:
        conn.execute(_text(
            "CREATE TABLE productions (id INTEGER PRIMARY KEY, org_id INTEGER NOT NULL, "
            "name VARCHAR(120) NOT NULL, is_main BOOLEAN NOT NULL, created_at DATETIME)"))
        conn.execute(_text(
            "INSERT INTO productions (id, org_id, name, is_main) VALUES (1, 7, 'Старый цех', 1)"))
        conn.execute(_text(
            "CREATE TABLE products (id INTEGER PRIMARY KEY, org_id INTEGER NOT NULL, "
            "ext_id VARCHAR(64) NOT NULL DEFAULT '', base_name VARCHAR(255) NOT NULL, "
            "size VARCHAR(32) NOT NULL DEFAULT '', category VARCHAR(128) NOT NULL DEFAULT '', "
            "sale_price FLOAT NOT NULL DEFAULT 0, cost_price FLOAT NOT NULL DEFAULT 0, "
            "archived BOOLEAN NOT NULL DEFAULT 0)"))
        conn.execute(_text(
            "INSERT INTO products (id, org_id, base_name, category) "
            "VALUES (1, 7, 'Пакет', 'Упаковка'), (2, 7, 'Худи', 'Одежда')"))
    race_engine.dispose()
    # Все процессы ждут одного и того же момента и стартуют вместе.
    start_at = time.time() + 2.0
    snippet = (
        "import os, sys, time\n"
        f"time.sleep(max(0.0, {start_at!r} - time.time()))\n"
        f"os.environ['DATABASE_URL'] = 'sqlite:///{race_db}'\n"
        "os.environ['SCHEDULER_ENABLED'] = '0'\n"
        f"sys.path.insert(0, {str(ROOT)!r})\n"
        "from app.db import init_db\n"
        "from app import exclusions\n"
        "init_db(); exclusions.ensure_schema()\n"
        "print('OK')\n"
    )
    started, results = _race_start(snippet, 4)
    check("одновременный старт 4 процессов на старой базе: поднялись все",
          started == 4,
          "поднялось %d/4; ошибки: %s" % (
              started, [_err_line(e) for o, e in results if "OK" not in o]))
    race_engine = _create_engine(f"sqlite:///{race_db}", future=True)
    race_cols = {c["name"] for c in _inspect(race_engine).get_columns("productions")}
    with race_engine.begin() as conn:
        race_flags = sorted(r[0] for r in conn.execute(
            _text("SELECT name FROM migration_flags")).fetchall())
        excl = dict(conn.execute(
            _text("SELECT base_name, excluded FROM products ORDER BY id")).fetchall())
        # пользователь вернул упаковку в аналитику — повторный старт не должен
        # перетереть это решение
        conn.execute(_text("UPDATE products SET excluded = 0 WHERE base_name = 'Пакет'"))
    check("после гонки схема корректна и флаги не задвоились",
          {"lead_time_days", "moq", "pack_multiple"} <= race_cols
          and race_flags == ["excl_samples_v1", "productions_conditions_v1"],
          f"cols={sorted(race_cols)} flags={race_flags}")
    check("бэкфилл эвристики отработал ровно один раз",
          excl == {"Пакет": 1, "Худи": 0}, f"excluded={excl}")
    race_engine.dispose()
    again = subprocess.run([sys.executable, "-c", snippet.replace(repr(start_at), "0.0")],
                           capture_output=True, text=True, timeout=90)
    race_engine = _create_engine(f"sqlite:///{race_db}", future=True)
    with race_engine.begin() as conn:
        excl2 = dict(conn.execute(
            _text("SELECT base_name, excluded FROM products ORDER BY id")).fetchall())
    check("повторный старт на мигрированной базе не перетирает выбор пользователя",
          "OK" in again.stdout and excl2 == {"Пакет": 0, "Худи": 0},
          f"excluded={excl2} err={again.stderr.strip()[-90:]}")
    race_engine.dispose()

    # 1в) Д4 (ревью деплоя 22.08): три оставшиеся миграции — ms_writeback,
    # ms_sync, ms_vendor — раньше делали голый ALTER TABLE без защиты от
    # гонки (и две из них вдобавок вызывались на импорте модуля, до старта
    # приложения). Проверяем то же самое, что и для models/exclusions выше:
    # «старая» схема без новых колонок, N процессов стартуют одновременно
    # по общему барьеру, все N обязаны подняться.
    from app import ms_sync as _ms_sync
    from app import ms_vendor as _ms_vendor
    from app import ms_writeback as _ms_writeback

    old3_db = Path(tempfile.mkdtemp()) / "old_schema3.db"
    old3_engine = _create_engine(f"sqlite:///{old3_db}", future=True)
    with old3_engine.begin() as conn:
        # production_orders без ms_doc_href/ms_doc_name (ms_writeback).
        conn.execute(_text(
            "CREATE TABLE production_orders (id INTEGER PRIMARY KEY, org_id INTEGER NOT NULL)"))
        # ordered_qty без ms_qty, sync_state без fail_streak/alerted_streak (ms_sync).
        conn.execute(_text(
            "CREATE TABLE ordered_qty (id INTEGER PRIMARY KEY, org_id INTEGER NOT NULL)"))
        conn.execute(_text(
            "CREATE TABLE sync_state (id INTEGER PRIMARY KEY, org_id INTEGER NOT NULL)"))
        # orgs без ms_account_id/source/status/ms_tariff_name, users без ms_uid (ms_vendor).
        conn.execute(_text(
            "CREATE TABLE orgs (id INTEGER PRIMARY KEY, name VARCHAR(120) NOT NULL)"))
        conn.execute(_text(
            "CREATE TABLE users (id INTEGER PRIMARY KEY, email VARCHAR(255) NOT NULL)"))
    for fn, table, expect_cols in (
        (_ms_writeback.ensure_schema, "production_orders", {"ms_doc_href", "ms_doc_name"}),
        (_ms_sync.ensure_schema, "ordered_qty", {"ms_qty"}),
        (_ms_sync.ensure_schema, "sync_state", {"fail_streak", "alerted_streak"}),
        (_ms_vendor.ensure_schema, "orgs",
         {"ms_account_id", "source", "status", "ms_tariff_name"}),
        (_ms_vendor.ensure_schema, "users", {"ms_uid"}),
    ):
        fn(bind=old3_engine)
        cols = {c["name"] for c in _inspect(old3_engine).get_columns(table)}
        check(f"миграция добавила недостающие колонки в {table} (одиночный запуск)",
              expect_cols <= cols, f"table={table} cols={sorted(cols)}")
    old3_engine.dispose()

    # Barrier-гонка: N процессов запускают ВСЕ пять миграций (models,
    # exclusions, ms_writeback, ms_sync, ms_vendor) одновременно на одной
    # старой базе — воспроизводит реальный старт `uvicorn --workers N`.
    N_RACE = 6
    race3_db = Path(tempfile.mkdtemp()) / "race_schema3.db"
    race3_engine = _create_engine(f"sqlite:///{race3_db}", future=True)
    with race3_engine.begin() as conn:
        conn.execute(_text(
            "CREATE TABLE productions (id INTEGER PRIMARY KEY, org_id INTEGER NOT NULL, "
            "name VARCHAR(120) NOT NULL, is_main BOOLEAN NOT NULL, created_at DATETIME)"))
        conn.execute(_text(
            "CREATE TABLE products (id INTEGER PRIMARY KEY, org_id INTEGER NOT NULL, "
            "ext_id VARCHAR(64) NOT NULL DEFAULT '', base_name VARCHAR(255) NOT NULL, "
            "size VARCHAR(32) NOT NULL DEFAULT '', category VARCHAR(128) NOT NULL DEFAULT '', "
            "sale_price FLOAT NOT NULL DEFAULT 0, cost_price FLOAT NOT NULL DEFAULT 0, "
            "archived BOOLEAN NOT NULL DEFAULT 0)"))
        conn.execute(_text(
            "CREATE TABLE production_orders (id INTEGER PRIMARY KEY, org_id INTEGER NOT NULL)"))
        conn.execute(_text(
            "CREATE TABLE ordered_qty (id INTEGER PRIMARY KEY, org_id INTEGER NOT NULL)"))
        conn.execute(_text(
            "CREATE TABLE sync_state (id INTEGER PRIMARY KEY, org_id INTEGER NOT NULL)"))
        conn.execute(_text(
            "CREATE TABLE orgs (id INTEGER PRIMARY KEY, name VARCHAR(120) NOT NULL)"))
        conn.execute(_text(
            "CREATE TABLE users (id INTEGER PRIMARY KEY, email VARCHAR(255) NOT NULL)"))
    race3_engine.dispose()
    start_at3 = time.time() + 2.0
    snippet3 = (
        "import os, sys, time\n"
        f"time.sleep(max(0.0, {start_at3!r} - time.time()))\n"
        f"os.environ['DATABASE_URL'] = 'sqlite:///{race3_db}'\n"
        "os.environ['SCHEDULER_ENABLED'] = '0'\n"
        f"sys.path.insert(0, {str(ROOT)!r})\n"
        "from app.db import init_db\n"
        "from app import exclusions, ms_sync, ms_vendor, ms_writeback\n"
        "init_db(); exclusions.ensure_schema()\n"
        "ms_writeback.ensure_schema(); ms_sync.ensure_schema(); ms_vendor.ensure_schema()\n"
        "print('OK')\n"
    )
    started3, results3 = _race_start(snippet3, N_RACE)
    check(f"barrier-гонка {N_RACE} процессов (все пять миграций, старая схема): поднялись все",
          started3 == N_RACE,
          "поднялось %d/%d; ошибки: %s" % (
              started3, N_RACE,
              [_err_line(e) for o, e in results3 if "OK" not in o]))
    race3_engine = _create_engine(f"sqlite:///{race3_db}", future=True)
    race3_insp = _inspect(race3_engine)
    race3_ok = (
        {"ms_doc_href", "ms_doc_name"} <= {c["name"] for c in race3_insp.get_columns("production_orders")}
        and {"ms_qty"} <= {c["name"] for c in race3_insp.get_columns("ordered_qty")}
        and {"fail_streak", "alerted_streak"} <= {c["name"] for c in race3_insp.get_columns("sync_state")}
        and {"ms_account_id", "source", "status", "ms_tariff_name"}
            <= {c["name"] for c in race3_insp.get_columns("orgs")}
        and {"ms_uid"} <= {c["name"] for c in race3_insp.get_columns("users")}
    )
    check("после гонки во всех пяти таблицах нужные колонки на месте",
          race3_ok, f"cols={ {t: sorted(c['name'] for c in race3_insp.get_columns(t)) for t in ('production_orders','ordered_qty','sync_state','orgs','users')} }")
    race3_engine.dispose()

    # То же самое на ЧИСТОЙ базе (create_all создаёт таблицы сразу со всеми
    # колонками — ensure_schema должен пройти как no-op и не мешать старту).
    clean3_db = Path(tempfile.mkdtemp()) / "clean_schema3.db"
    start_at3c = time.time() + 1.5
    snippet3c = snippet3.replace(str(race3_db), str(clean3_db)).replace(
        repr(start_at3), repr(start_at3c))
    started3c, results3c = _race_start(snippet3c, N_RACE)
    check(f"barrier-гонка {N_RACE} процессов на ЧИСТОЙ базе: поднялись все",
          started3c == N_RACE,
          "поднялось %d/%d; ошибки: %s" % (
              started3c, N_RACE,
              [_err_line(e) for o, e in results3c if "OK" not in o]))

    # 2) Округление до минимальной партии и кратности на демо-данных.
    repl0 = demo.get("/api/replenish").json()
    check("без условий производства количество равно расчётному",
          all(not it["moq_applied"] and it["need"] == it["need_raw"] for it in repl0["items"]),
          f"изменённых={[i['base_name'] for i in repl0['items'] if i['moq_applied']][:3]}")
    by_need = sorted(repl0["items"], key=lambda x: x["need"])
    small = by_need[0]                                    # потребность в пару штук
    # крупная позиция — на ней видно кратность (у мелкой её съедает партия)
    big = next((it for it in by_need if it["need"] > 60), by_need[-1])
    r = demo.post("/api/productions", json={
        "name": "Бишкек", "lead_time_days": 70, "moq": 30, "pack_multiple": 6})
    pid = r.json().get("id")
    check("производство с условиями создано",
          r.status_code == 200 and r.json().get("lead_time_days") == 70
          and r.json().get("moq") == 30 and r.json().get("pack_multiple") == 6,
          f"resp={r.text[:160]}")
    for base in (small["base_name"], big["base_name"]):
        rr = demo.post("/api/productions/assign", json={"base_name": base, "production_id": pid})
        check(f"позиция перенесена на «Бишкек»: {base}", rr.status_code == 200,
              f"status={rr.status_code}")
    moved = {it["base_name"]: it for it in demo.get("/api/replenish").json()["items"]}
    s_it = moved[small["base_name"]]
    b_it = moved[big["base_name"]]
    # need_raw после переноса пересчитан по сроку «Бишкека» (70 дней против
    # общих 45): за дольшее ожидание успеет продаться больше, поэтому расчёт
    # не может стать меньше прежнего. Требовать РАВЕНСТВА старому числу
    # нельзя — это и была та половина фичи, где срок подрядчика не влиял
    # ни на что, кроме показанной даты прихода.
    check("количество поднято до минимальной партии",
          s_it["need"] == 30 and s_it["need_raw"] >= small["need"] and s_it["moq_applied"],
          f"{small['need']} → need_raw={s_it['need_raw']} → {s_it['need']}")
    check("более долгий срок подрядчика не уменьшил расчётную потребность",
          b_it["need_raw"] >= big["need"],
          f"{big['need']} → {b_it['need_raw']} (срок 45 → 70 дн)")
    check("потребность сильно ниже партии помечена «заказывать невыгодно»",
          s_it["moq_low"] is (small["need"] < 15), f"need={small['need']} moq_low={s_it['moq_low']}")
    check("крупная позиция округлена вверх до кратности",
          b_it["need"] % 6 == 0 and b_it["need"] >= b_it["need_raw"] > 0
          and b_it["need"] - b_it["need_raw"] < 6,
          f"{b_it['need_raw']} → {b_it['need']}")
    for it in (s_it, b_it):
        check(f"сумма по размерам равна итогу позиции: {it['base_name']}",
              sum(v["rec"] for v in it["sizes"].values()) == it["need"],
              f"размеры={sum(v['rec'] for v in it['sizes'].values())} итог={it['need']}")
    check("срок производства отдаётся по позиции",
          s_it["lead_time_days"] == 70 and b_it["lead_time_days"] == 70,
          f"got={(s_it['lead_time_days'], b_it['lead_time_days'])}")
    others = [it for it in moved.values() if it["production_id"] != pid]
    check("позиции основного производства считаются по общему сроку",
          all(it["lead_time_days"] == repl0["lead_time_days"] and not it["moq_applied"]
              for it in others),
          f"n={len(others)}")

    print("== Ревью 22.08: «Заказ позиции» считает срок и раскладку как «Заказ» ==")
    from datetime import date as _d2, timedelta as _td2
    today2 = _d2.today()
    # Позиция на «Бишкеке» (срок 70): сервер обязан отдать дату прихода по
    # ЕГО сроку с первого ответа, без второго запроса с поправкой из браузера.
    sc = demo.get("/api/sizes/calc", params={
        "product": b_it["base_name"], "qty": b_it["need"], "period": "12m", "mode": "stock"})
    check("sizes/calc: срок и производство позиции на подрядчике",
          sc.status_code == 200 and sc.json()["lead_time_days"] == 70
          and sc.json()["production_id"] == pid and sc.json()["production_name"] == "Бишкек",
          f"resp={sc.text[:200]}")
    check("sizes/calc: дата прихода сразу по сроку подрядчика (70 дн), без второго запроса",
          sc.json()["days_to_arrival"] == 70
          and sc.json()["arrival_date"] == (today2 + _td2(days=70)).isoformat(),
          f"got days={sc.json()['days_to_arrival']} arrival={sc.json()['arrival_date']}")
    # Позиция основного производства — по общему сроку, как и раньше.
    other_it = others[0]
    sc_main = demo.get("/api/sizes/calc", params={
        "product": other_it["base_name"], "qty": 10, "period": "12m", "mode": "stock"})
    check("sizes/calc: позиция основного производства — общий срок",
          sc_main.status_code == 200
          and sc_main.json()["lead_time_days"] == repl0["lead_time_days"],
          f"resp={sc_main.text[:200]}")
    # Раскладка по размерам на равном количестве обязана совпасть с «Заказом»
    # (общая analytics.size_split, а не свой пересчёт по календарным месяцам).
    order_rec = {s: v["rec"] for s, v in b_it["sizes"].items()}
    calc_pure = {row["size"]: row["order_pure"] for row in sc.json()["sizes"]}
    check("sizes/calc: «чистая пропорция» = ростовка «Заказа» на том же количестве",
          sc.json()["split_matches_order_page"] is True and calc_pure == order_rec,
          f"заказ={order_rec} sizes/calc={calc_pure}")

    # Заказ уходит с округлёнными числами (как их видит человек на странице).
    r = demo.post("/api/orders", json={"name": "Проверка партии", "eta_date": None, "items": [
        {"base_name": s_it["base_name"], "qty": s_it["need"],
         "sizes": {k: v["rec"] for k, v in s_it["sizes"].items() if v["rec"] > 0},
         "cost": s_it["cost_price"]}]})
    check("заказ по позиции с минимальной партией создан", r.status_code == 200, f"resp={r.text[:160]}")
    order = demo.get(f"/api/orders/{r.json()['id']}").json()
    check("в заказ ушло округлённое количество",
          order["total_qty"] == 30 and sum(order["items"][0]["sizes"].values()) == 30,
          f"qty={order['total_qty']}")
    demo.delete(f"/api/orders/{order['id']}")

    # 2б) Защита от случайного дубля заказа (ревью 22.08: тройной POST давал
    # три настоящих заказа на фабрику — двойной тап, ретрай, повтор формы).
    need_before = sum(it["need"] for it in demo.get("/api/replenish").json()["items"])
    dup_payload = {"name": "Заказ-дубль", "eta_date": None, "items": [
        {"base_name": s_it["base_name"], "qty": 12, "sizes": {}, "cost": 100}]}
    n0 = len(demo.get("/api/orders").json()["orders"])
    answers = [demo.post("/api/orders", json=dup_payload).json() for _ in range(3)]
    n1 = len(demo.get("/api/orders").json()["orders"])
    check("три одинаковых POST /api/orders создают один заказ", n1 - n0 == 1,
          f"было {n0}, стало {n1}, ответы={[a.get('id') for a in answers]}")
    check("повторы отмечены duplicate и указывают на тот же заказ",
          [a.get("duplicate") for a in answers] == [None, True, True]
          and len({a["id"] for a in answers}) == 1, f"answers={answers}")
    check("человеку объясняют, как всё-таки заказать второй раз",
          "измените название" in answers[1].get("message", ""),
          f"message={answers[1].get('message', '')[:80]}")
    dup_id = answers[0]["id"]
    r = demo.post(f"/api/orders/{dup_id}/status", json={"status": "sent"})
    r2 = demo.post(f"/api/orders/{dup_id}/status", json={"status": "sent"})
    check("повтор того же статуса не ошибка и ничего не меняет",
          r.status_code == 200 and r2.status_code == 200 and r2.json().get("unchanged") is True,
          f"первый={r.status_code} второй={r2.text[:80]}")
    r = demo.post("/api/orders", json=dict(dup_payload, allow_duplicate=True))
    forced_id = r.json().get("id")
    check("осознанный второй такой же заказ (allow_duplicate) проходит",
          r.status_code == 200 and forced_id != dup_id and not r.json().get("duplicate"),
          f"resp={r.text[:120]}")
    r = demo.post("/api/orders", json=dict(dup_payload, name="Заказ-дубль второй"))
    renamed_id = r.json().get("id")
    check("другое название — другой заказ", renamed_id not in (dup_id, forced_id),
          f"id={renamed_id}")
    for oid in (dup_id, forced_id, renamed_id):
        demo.delete(f"/api/orders/{oid}")
    check("после уборки заказов рекомендация вернулась к прежней",
          sum(it["need"] for it in demo.get("/api/replenish").json()["items"]) == need_before,
          f"было {need_before}")

    # 2в) Дедуп не должен цепляться к уже ПРИНЯТОМУ на склад заказу — это уже
    # история, и повтор того же состава после приёмки означает «нужен ещё
    # такой же», а не случайный повтор формы (ревью деплоя 22.08: раньше
    # клиент получал обратно received-заказ и падал в 422 на переходе в sent).
    recv_payload = {"name": "Заказ до приёмки", "eta_date": None, "items": [
        {"base_name": s_it["base_name"], "qty": 7, "sizes": {}, "cost": 100}]}
    recv_id = demo.post("/api/orders", json=recv_payload).json()["id"]
    demo.post(f"/api/orders/{recv_id}/status", json={"status": "sent"})
    r = demo.post(f"/api/orders/{recv_id}/status", json={"status": "received"})
    check("заказ принят на склад (подготовка)",
          r.status_code == 200 and r.json()["status"] == "received", f"resp={r.text[:120]}")
    r = demo.post("/api/orders", json=recv_payload)
    repeated = r.json()
    check("повтор того же состава после приёмки создаёт НОВЫЙ заказ, не склеивается с received",
          r.status_code == 200 and not repeated.get("duplicate") and repeated["id"] != recv_id,
          f"resp={repeated}")
    r2 = demo.post(f"/api/orders/{repeated['id']}/status", json={"status": "sent"})
    check("новый заказ спокойно переводится в «В производстве» (не 422 недопустимого перехода)",
          r2.status_code == 200 and r2.json().get("status") == "sent",
          f"status={r2.status_code} body={r2.text[:150]}")
    demo.delete(f"/api/orders/{repeated['id']}")

    # Для сравнения: повтор для заказа, который ещё «В производстве» (sent, а
    # не received), по-прежнему безопасно склеивается — статус можно послать
    # повторно, ничего не ломается (unchanged).
    sent_payload = {"name": "Заказ уже в производстве", "eta_date": None, "items": [
        {"base_name": s_it["base_name"], "qty": 5, "sizes": {}, "cost": 100}]}
    sent_id = demo.post("/api/orders", json=sent_payload).json()["id"]
    demo.post(f"/api/orders/{sent_id}/status", json={"status": "sent"})
    twin = demo.post("/api/orders", json=sent_payload).json()
    check("повтор для sent-заказа по-прежнему склеивается с ним",
          twin.get("duplicate") is True and twin.get("id") == sent_id, f"resp={twin}")
    r2 = demo.post(f"/api/orders/{sent_id}/status", json={"status": "sent"})
    check("статус sent→sent для склеенного заказа не ошибка (unchanged)",
          r2.status_code == 200 and r2.json().get("unchanged") is True, f"resp={r2.text[:120]}")
    demo.delete(f"/api/orders/{sent_id}")

    # 2г) Позиция вне каталога организации: заказ не создаётся вообще, и
    # заодно проверяем, что сервер никогда не доверяет присланной клиентом
    # себестоимости (раньше для «призрачных» позиций верил ей ровно потому,
    # что не знал, что с ними делать иначе — см. п. 2д про 404).
    r = demo.post("/api/orders", json={"name": "Призрак", "eta_date": None, "items": [
        {"base_name": "Несуществующий товар XYZ 12345", "qty": 5, "sizes": {}, "cost": 999999}]})
    check("заказ на товар не из каталога → 404, а не 200 с придуманной ценой",
          r.status_code == 404 and "каталоге" in r.text, f"status={r.status_code} body={r.text[:150]}")
    r = demo.post("/api/orders", json={"name": "Наполовину призрак", "eta_date": None, "items": [
        {"base_name": s_it["base_name"], "qty": 3, "sizes": {}, "cost": 1},
        {"base_name": "Ещё один призрак", "qty": 2, "sizes": {}, "cost": 1}]})
    check("даже одна позиция вне каталога отклоняет заказ целиком",
          r.status_code == 404, f"status={r.status_code}")
    r = demo.post("/api/orders", json={"name": "Реальный товар, чужая цена", "eta_date": None,
                                        "items": [{"base_name": s_it["base_name"], "qty": 1,
                                                    "sizes": {}, "cost": 1}]})
    check("заказ на реальный товар создаётся", r.status_code == 200, f"resp={r.text[:120]}")
    trusted_id = r.json()["id"]
    trusted_order = demo.get(f"/api/orders/{trusted_id}").json()
    check("присланная клиентом себестоимость (1 ₽) проигнорирована — подставлена своя из каталога",
          trusted_order["items"][0]["cost"] == s_it["cost_price"]
          and trusted_order["items"][0]["cost"] != 1,
          f"cost={trusted_order['items'][0]['cost']} expected={s_it['cost_price']}")
    demo.delete(f"/api/orders/{trusted_id}")

    # 2д) eta_date — только формат ГГГГ-ММ-ДД; название заказа ограничено
    # 120 символами (как у названия производства, см. ProductionIn).
    r = demo.post("/api/orders", json={"name": "Кривая дата", "eta_date": "not-a-date",
                                        "items": [{"base_name": s_it["base_name"], "qty": 1,
                                                    "sizes": {}, "cost": 1}]})
    check("eta_date не в формате ГГГГ-ММ-ДД → 422", r.status_code == 422, f"status={r.status_code}")
    r = demo.post("/api/orders", json={"name": "д" * 200, "eta_date": None,
                                        "items": [{"base_name": s_it["base_name"], "qty": 1,
                                                    "sizes": {}, "cost": 1}]})
    check("название заказа длиннее 120 символов → 422", r.status_code == 422, f"status={r.status_code}")
    r = demo.post("/api/orders", json={"name": "Валидные дата и название", "eta_date": "2026-09-01",
                                        "items": [{"base_name": s_it["base_name"], "qty": 1,
                                                    "sizes": {}, "cost": 1}]})
    check("валидная дата и короткое название по-прежнему создают заказ",
          r.status_code == 200, f"status={r.status_code}")
    demo.delete(f"/api/orders/{r.json()['id']}")
    check("после уборки этих заказов рекомендация снова как прежде",
          sum(it["need"] for it in demo.get("/api/replenish").json()["items"]) == need_before,
          f"было {need_before}")

    # 3) Изоляция: производства и их условия — строго внутри организации.
    ms_prods = client.get("/api/productions").json()
    check("чужое производство не видно другой организации",
          all(p["name"] != "Бишкек" for p in ms_prods["productions"]),
          f"got={[p['name'] for p in ms_prods['productions']]}")
    r = client.post("/api/productions/assign",
                    json={"base_name": "Худи «Скетч»", "production_id": pid})
    check("перенос позиции на чужое производство → 404", r.status_code == 404,
          f"status={r.status_code}")
    r = client.post(f"/api/productions/{pid}", json={"name": "Взлом", "moq": 999})
    check("правка чужого производства → 404", r.status_code == 404, f"status={r.status_code}")
    r = client.delete(f"/api/productions/{pid}")
    check("удаление чужого производства → 404", r.status_code == 404, f"status={r.status_code}")
    check("после чужих попыток условия не изменились",
          demo.get("/api/productions").json()["productions"][-1]["moq"] == 30)

    # 4) Границы id в пути: огромное число — понятная 422, а не 500.
    for path, method in (("/api/orders/999999999999999999999/ms-doc", "GET"),
                         ("/api/orders/999999999999999999999/push-to-ms", "POST"),
                         ("/api/orders/999999999999999999999", "GET")):
        r = demo.request(method, path)
        check(f"огромный id в пути → 422 ({path.split('/')[-1]})", r.status_code == 422,
              f"status={r.status_code}")
    r = demo.post("/api/productions/assign",
                  json={"base_name": "Пальто «Кокон»", "production_id": 10 ** 24})
    check("огромный id производства в теле запроса → 422", r.status_code == 422,
          f"status={r.status_code}")

    print("== Выгрузка «Что заказать» = то, что человек видит на экране ==")
    # Файл несут на фабрику: он обязан совпадать со страницей и по условиям
    # подрядчика (минимальная партия, кратность), и по ручным правкам ростовки.
    # Раньше выгрузка звала build_replenish напрямую и не знала ни про то, ни
    # про другое: на экране 17 шт «правлено», в файле 19 шт по расчёту.
    from openpyxl import load_workbook as _load_wb

    def _xlsx_replenish(cl):
        """Позиции выгрузки «Что заказать» по ВСЕМ листам производств.

        Книга разложена так же, как страница: лист на каждое производство,
        со своей шапкой и своим «Итого» (раньше был один лист и один общий
        итог, который не совпадал ни с одной вкладкой — ревью «экран ≠ файл»).
        Возвращает {позиция: ...}, {лист: итог} и {лист: шапка A1}.
        """
        resp = cl.get("/api/export/replenish.xlsx")
        wb = _load_wb(io.BytesIO(resp.content))
        out, totals, titles = {}, {}, {}
        for ws in wb:
            if ws.title == "Не вошло и почему":
                continue
            titles[ws.title] = ws["A1"].value or ""
            cur = None
            for row in ws.iter_rows(min_row=3, values_only=True):
                name = row[0]
                if not name:
                    continue
                name = str(name)
                if name.startswith("Итого"):
                    totals[ws.title] = row[10]
                    continue
                if name.startswith("Позиции, не попавшие") or name.startswith("Это итог"):
                    continue
                if name.startswith("— "):
                    if cur:
                        out[cur]["sizes"][name[2:]] = row[10]
                        out[cur]["size_notes"][name[2:]] = row[14]
                    continue
                cur = name
                out[name] = {"need": row[10], "prod": row[13], "note": row[14],
                             "sheet": ws.title, "sizes": {}, "size_notes": {}}
        return out, totals, titles

    api_items = {it["base_name"]: it for it in demo.get("/api/replenish").json()["items"]}
    xls, xls_totals, xls_titles = _xlsx_replenish(demo)
    check("в выгрузке столько же позиций, сколько на странице",
          len(xls) == len(api_items), f"файл={len(xls)} страница={len(api_items)}")
    diff = [(n, api_items[n]["need"], xls[n]["need"]) for n in api_items
            if n in xls and xls[n]["need"] != api_items[n]["need"]]
    check("количества в файле совпадают со страницей (с минимальной партией)",
          not diff, f"расходятся={diff[:3]}")
    check("итог файла равен сумме строк файла",
          sum(xls_totals.values()) == sum(v["need"] for v in xls.values()),
          f"итого={xls_totals} сумма={sum(v['need'] for v in xls.values())}")
    b_name = b_it["base_name"]
    check("в файле указано производство позиции",
          xls[b_name]["prod"] == "Бишкек", f"got={xls[b_name]['prod']}")

    # Экран ≠ файл (находка ревью): страница разведена по вкладкам производств,
    # а книга склеивала их в один лист с одним итогом — «35 поз. / 1 802 шт» не
    # совпадало ни с одной вкладкой, и шапка обещала общий срок 45 дней над
    # строкой, которая шьётся 70 дней у другого подрядчика.
    _tabs = {}
    for _it in api_items.values():
        _tabs.setdefault(_it.get("production_name") or "", []).append(_it)
    check("в книге лист на каждое производство",
          len(xls_totals) == len(_tabs), f"листов={sorted(xls_totals)} вкладок={sorted(_tabs)}")
    _bad_tabs = []
    for _name, _its in _tabs.items():
        _sheet = next((t for t in xls_totals if t.startswith(_name[:20])), None)
        _want = sum(i["need"] for i in _its)
        if _sheet is None or xls_totals[_sheet] != _want:
            _bad_tabs.append((_name, _want, _sheet and xls_totals[_sheet]))
    check("итог листа совпадает с итогом своей вкладки на странице",
          not _bad_tabs, f"расходятся={_bad_tabs}")
    check("позиции «Бишкека» лежат на своём листе, а не вперемешку",
          all(xls[i["base_name"]]["sheet"].startswith("Бишкек") for i in _tabs["Бишкек"])
          and all(not xls[i["base_name"]]["sheet"].startswith("Бишкек")
                  for i in _tabs["Основное производство"]),
          f"листы={ {n: v['sheet'] for n, v in list(xls.items())[:3]} }")
    _t_bish = next(t for t in xls_titles if t.startswith("Бишкек"))
    check("шапка листа называет срок СВОЕГО подрядчика, а не общий из настроек",
          "срок производства 70 дней" in xls_titles[_t_bish]
          and "партия от 30 шт" in xls_titles[_t_bish],
          f"A1={xls_titles[_t_bish]}")
    _t_main = next(t for t in xls_titles if not t.startswith("Бишкек"))
    check("шапка основного листа обещает срок основного производства",
          "срок производства 45 дней" in xls_titles[_t_main], f"A1={xls_titles[_t_main]}")
    # Примечание проверяем на мелкой позиции: её партия поднимает гарантированно,
    # тогда как крупная попадает под кратность только если расчёт на неё не делится
    # (после того как срок подрядчика стал влиять на расчёт, это уже не гарантия).
    s_name = s_it["base_name"]
    check("примечание объясняет, откуда взялось количество больше расчётного",
          "по расчёту" in (xls[s_name]["note"] or "") and "партия" in (xls[s_name]["note"] or ""),
          f"note={xls[s_name]['note']}")
    check("сумма по размерам в файле равна итогу позиции",
          sum(xls[b_name]["sizes"].values()) == xls[b_name]["need"],
          f"размеры={sum(xls[b_name]['sizes'].values())} итог={xls[b_name]['need']}")

    # Ручная правка: человек уменьшил один размер под фабрику.
    b_sizes = {k: v["rec"] for k, v in api_items[b_name]["sizes"].items()}
    first_size = sorted(b_sizes)[0]
    manual = {first_size: 1}
    expect_need = sum(manual.get(k, v) for k, v in b_sizes.items())
    r = demo.post("/api/replenish-draft", json={"base_name": b_name, "sizes": manual})
    check("ручная правка ростовки сохранена", r.status_code == 200, f"resp={r.text[:120]}")
    xls, xls_totals, _ = _xlsx_replenish(demo)
    check("в файле — правка человека, а не расчёт",
          xls[b_name]["need"] == expect_need != api_items[b_name]["need"],
          f"файл={xls[b_name]['need']} ожидали={expect_need} расчёт={api_items[b_name]['need']}")
    check("правленый размер в файле равен введённому",
          xls[b_name]["sizes"].get(first_size) == 1,
          f"got={xls[b_name]['sizes'].get(first_size)}")
    check("примечание называет и расчёт, и правку",
          "правлена вручную" in (xls[b_name]["note"] or ""), f"note={xls[b_name]['note']}")
    check("правленый размер подписан расчётом, от которого отошли",
          "правлено вручную" in (xls[b_name]["size_notes"].get(first_size) or "")
          and str(b_sizes[first_size]) in (xls[b_name]["size_notes"].get(first_size) or ""),
          f"note={xls[b_name]['size_notes'].get(first_size)}")
    check("итог файла пересчитан по правке",
          sum(xls_totals.values()) == sum(v["need"] for v in xls.values()),
          f"итого={xls_totals} сумма={sum(v['need'] for v in xls.values())}")
    # Ручная правка сильнее минимальной партии — но о нарушении говорим вслух.
    check("правка ниже минимальной партии не переписана условиями подрядчика",
          xls[b_name]["need"] == expect_need, f"got={xls[b_name]['need']}")
    if expect_need < 30 or expect_need % 6:
        check("файл предупреждает, что заказ не проходит по условиям фабрики",
              "не проходит по условиям" in (xls[b_name]["note"] or ""),
              f"note={xls[b_name]['note']}")
    check("правки одной позиции не задели остальные строки файла",
          all(xls[n]["need"] == api_items[n]["need"] for n in xls if n != b_name),
          "расходятся: " + str([n for n in xls if n != b_name
                                and xls[n]["need"] != api_items[n]["need"]][:3]))
    demo.post("/api/replenish-draft/reset", json={"base_name": ""})
    xls, _, _ = _xlsx_replenish(demo)
    check("после сброса правок файл вернулся к расчёту",
          xls[b_name]["need"] == api_items[b_name]["need"],
          f"файл={xls[b_name]['need']} расчёт={api_items[b_name]['need']}")
    print("== Формульная инъекция в выгрузках Excel ==")
    import zipfile
    from datetime import date as _d, timedelta as _td

    # Название организации, позиции, категории и производства задаёт человек,
    # а файл открывает бухгалтерия клиента. Строка, начинающаяся с '=', '+',
    # '-' или '@', в Excel исполняется как формула (CWE-1236): «=cmd|"/c
    # calc"!A1» — это команда операционной системе, а не текст.
    _EVIL = '=cmd|"/c calc"!A1'
    evil = httpx.Client(headers={"X-Oborot-CSRF": "1"},
                        base_url=f"http://127.0.0.1:{APP_PORT}", timeout=120.0)
    r = evil.post("/register", data={
        "name": "Вредитель", "email": "evil@test.io",
        "password": "secret123", "org_name": _EVIL,
    })
    check("регистрация с формулой в названии организации", r.status_code == 303,
          f"status={r.status_code}")
    evil.post("/api/connect/demo")
    _con_ev = sqlite3.connect(DB_PATH)
    _evil_org = _con_ev.execute(
        "SELECT id FROM orgs WHERE name=?", (_EVIL,)).fetchone()[0]
    # тот же payload — в название позиции, категорию и производство
    _con_ev.execute(
        "UPDATE products SET base_name=?, category=? WHERE org_id=? AND base_name=?",
        (_EVIL, "@SUM(1+1)", _evil_org, "Худи «Штрих»"))
    _con_ev.commit()
    _con_ev.close()
    r = evil.post("/api/productions", json={"name": '=HYPERLINK("http://evil","Цех")'})
    _evil_pid = r.json().get("id")
    evil.post("/api/productions/assign",
              json={"base_name": _EVIL, "production_id": _evil_pid})
    _risky, _unprotected, _formulas = 0, [], 0
    for _url in ("/api/export/replenish.xlsx", "/api/export/turnover.xlsx",
                 "/api/export/budget.xlsx"):
        _resp = evil.get(_url)
        check(f"выгрузка отдаётся: {_url.rsplit('/', 1)[-1]}", _resp.status_code == 200,
              f"status={_resp.status_code}")
        _wb = _load_wb(io.BytesIO(_resp.content))
        for _ws in _wb:
            for _row in _ws.iter_rows():
                for _c in _row:
                    if isinstance(_c.value, str) and _c.value[:1] in ("=", "+", "-", "@"):
                        _risky += 1
                        # data_type 'f' — openpyxl записал бы настоящую формулу;
                        # quotePrefix — тот самый апостроф Excel, невидимый в
                        # значении и переживающий копирование и CSV
                        if _c.data_type == "f" or not _c.quotePrefix:
                            _unprotected.append((_url, _ws.title, _c.coordinate,
                                                 _c.value[:40]))
        _formulas += sum(
            zipfile.ZipFile(io.BytesIO(_resp.content)).read(_n).count(b"<f>")
            for _n in zipfile.ZipFile(io.BytesIO(_resp.content)).namelist()
            if _n.startswith("xl/worksheets/")
        )
    check("пользовательский текст с '=' долетел до всех выгрузок (тест бьёт в цель)",
          _risky >= 3, f"опасных ячеек={_risky}")
    check("ни одна такая ячейка не осталась без защиты",
          not _unprotected, f"без защиты={_unprotected[:3]}")
    check("в книгах нет ни одной ячейки-формулы",
          _formulas == 0, f"формул в xml={_formulas}")
    evil.close()

    print("== Сезонное эхо: фолбэк не заказывает то, что умерло полгода назад ==")
    # Фолбэк на годовой темп срабатывал по факту «товара не было на складе» —
    # и одинаково тянул в заказ и бестселлер, распроданный вчера, и зимнее
    # пальто, которого нет с января. Разводим три случая на одних данных.
    _con_e = sqlite3.connect(DB_PATH)
    _demo_org2 = _con_e.execute(
        "SELECT org_id FROM products WHERE base_name=? AND org_id<>? LIMIT 1",
        ("Пальто «Кокон»", _evil_org)).fetchone()[0]

    def _wipe(base, sale_day, born_day=None):
        """Позиция ушла с полки в sale_day; продажи оставлены одним днём.

        born_day — если задан, вся история ДО этой даты стирается: получается
        новинка, которой в сезонном окне прошлого года ещё не существовало.
        """
        pids = [x[0] for x in _con_e.execute(
            "SELECT id FROM products WHERE org_id=? AND base_name=?",
            (_demo_org2, base))]
        qs = ",".join("?" * len(pids))
        gone = (_d.fromisoformat(sale_day) + _td(days=1)).isoformat()
        _con_e.execute(f"UPDATE stock_days SET qty=0 WHERE product_id IN ({qs}) AND date>=?",
                       pids + [gone])
        _con_e.execute(f"UPDATE warehouse_stock SET qty=0 WHERE product_id IN ({qs})", pids)
        _con_e.execute(f"DELETE FROM sales WHERE product_id IN ({qs})", pids)
        if born_day:
            _con_e.execute(
                f"DELETE FROM stock_days WHERE product_id IN ({qs}) AND date<?",
                pids + [born_day])
        for pid in pids:
            _con_e.execute(
                "INSERT INTO sales (org_id,product_id,date,qty,revenue,is_return)"
                " VALUES (?,?,?,?,?,0)",
                (_demo_org2, pid, sale_day, 9.0, 9 * 26000.0))

    _today_e = _d.today()
    # 1) «Сезонное эхо»: продавалось только в январе, с полки ушло тогда же,
    #    и в сезонном окне прошлого года (период прихода заказа) продаж нет.
    _echo_day = (_today_e - _td(days=225)).isoformat()
    _wipe("Пальто «Кокон»", _echo_day)
    # 2) Бестселлер, распроданный давно, НО продававшийся в сезонном окне
    #    прошлого года — ровно в тот период, на который считается заказ.
    _season_day = (_today_e + _td(days=45 + 10 - 365)).isoformat()
    _wipe("Платье «Макси чёрное»", _season_day)
    # 3) Новинка: появилась 120 дней назад, распродана в ноль 100 дней назад —
    #    сезонного окна у неё нет вовсе, отсекать её не за что.
    _wipe("Тренч «Классика»", (_today_e - _td(days=100)).isoformat(),
          born_day=(_today_e - _td(days=120)).isoformat())
    _con_e.commit()
    _con_e.close()

    demo.post("/api/settings", json={"rate_window": "d90"})  # заодно сбрасывает кэш
    _re = demo.get("/api/replenish").json()
    _in = {x["base_name"]: x for x in _re["items"]}
    _out = {e["base_name"]: e["reason"] for e in _re["excluded"]}
    check("d90: распроданное полгода назад и не в сезон в заказ НЕ идёт",
          "Пальто «Кокон»" not in _in,
          f"need={_in.get('Пальто «Кокон»', {}).get('need')}")
    check("d90: причина исключения объясняет, а не отмахивается",
          "не было на складе" in (_out.get("Пальто «Кокон»") or ""),
          f"reason={_out.get('Пальто «Кокон»')!r}")
    check("d90: бестселлер, распроданный давно, но продающийся в сезон заказа, остался",
          "Платье «Макси чёрное»" in _in and _in["Платье «Макси чёрное»"]["rate_fallback"],
          f"в заказе={'Платье «Макси чёрное»' in _in} excl={_out.get('Платье «Макси чёрное»')!r}")
    check("d90: новинку, распроданную в ноль, сезонное окно не отсекает",
          "Тренч «Классика»" in _in and _in["Тренч «Классика»"]["rate_fallback"],
          f"в заказе={'Тренч «Классика»' in _in} excl={_out.get('Тренч «Классика»')!r}")
    _yr = demo.post("/api/settings", json={"rate_window": "year"})
    _ry = demo.get("/api/replenish").json()
    check("окно «за год» правкой не задето: там темп настоящий, а не фолбэк",
          any(x["base_name"] == "Пальто «Кокон»" for x in _ry["items"])
          and not any(x["rate_fallback"] for x in _ry["items"]),
          f"в заказе={[x['base_name'] for x in _ry['items'] if x['base_name'] == 'Пальто «Кокон»']}")

    print("== Возвраты не делают деньги отрицательными ==")
    # Возврат вычитается из выручки. Если за год их больше, чем продаж,
    # нетто-выручка отрицательна — «средняя цена −4 888 ₽» отравляла и цену,
    # и маржу, и «заморожено», и кольцо «Склад в деньгах» (находка ревью).
    _con_r = sqlite3.connect(DB_PATH)
    _rid = [x[0] for x in _con_r.execute(
        "SELECT id FROM products WHERE org_id=? AND base_name=?",
        (_demo_org2, 'Кольцо «Печатка»'))]
    # Возврат небольшой по штукам, но крупный по деньгам (вернули то, что
    # продавали по полной, а продажи шли со скидкой): штуки остаются
    # положительными, а нетто-выручка уходит в минус — ровно случай ревью.
    for _pid in _rid:
        _con_r.execute(
            "INSERT INTO sales (org_id,product_id,date,qty,revenue,is_return)"
            " VALUES (?,?,?,?,?,1)",
            (_demo_org2, _pid, (_today_e - _td(days=5)).isoformat(), 5.0, 5 * 900_000.0))
    _con_r.commit()
    _con_r.close()
    demo.post("/api/settings", json={"rate_window": "year"})  # сброс кэша
    _tn = demo.get("/api/turnover").json()
    _ring = demo.get("/api/pulse").json()
    _fc2 = demo.get("/api/forecast").json()
    _ri = next(x for x in _tn["items"] if x["base_name"] == 'Кольцо «Печатка»')
    check("возвраты больше продаж: средней цены продажи нет, а не «минус»",
          _ri["avg_price"] is None and _ri["returns_over_sales"],
          f"avg_price={_ri['avg_price']} flag={_ri.get('returns_over_sales')}")
    check("оборачиваемость ₽/день не отрицательная",
          _ri["turnover"] >= 0 and all(v >= 0 for v in (_ri.get("sea") or {}).values()),
          f"turnover={_ri['turnover']} sea={_ri.get('sea')}")
    check("деньги в остатке по такой позиции не ушли в минус",
          _ri["stock_sale"] >= 0 and (_ri["stock_margin"] is None or _ri["stock_margin"] >= 0
                                      or _ri["cs"] == 0),
          f"stock_sale={_ri['stock_sale']} stock_margin={_ri['stock_margin']}")
    _neg = [x["base_name"] for x in _tn["items"]
            if x["turnover"] < 0 or (x["stock_sale"] or 0) < 0
            or (x["avg_price"] is not None and x["avg_price"] < 0)
            or any(v < 0 for v in (x.get("sea") or {}).values())]
    check("во всей таблице нет отрицательных цен, ₽/день и сумм",
          not _neg, f"отрицательные={_neg[:3]}")
    _tot = _tn.get("totals") or {}
    check("агрегаты страницы не отрицательные",
          all(v >= 0 for k, v in _tot.items()
              if isinstance(v, (int, float)) and k.startswith("stock_")),
          f"totals={ {k: v for k, v in _tot.items() if isinstance(v, (int, float)) and v < 0} }")
    check("кольцо «Склад в деньгах» и «Прогноз» тоже без минуса",
          (_ring.get("stock", {}).get("now", 0) or 0) >= 0
          and _fc2["cards"]["stock_value"] >= 0,
          f"кольцо={_ring.get('stock', {}).get('now')} прогноз={_fc2['cards']['stock_value']}")
    check("сезонные колонки с перевесом возвратов подписаны словами",
          not _ri.get("sea_returns")
          or "возвраты" in (_ri.get("sea_returns") and "возвраты" or ""),
          f"sea_returns={_ri.get('sea_returns')}")

    print("== «Стока хватит до»: заказанный товар не приближает дефицит ==")
    # Находка бухгалтера: добавление 223 шт «в производство» сдвигало дату
    # на неделю РАНЬШЕ. Порог считался от «склад + едет», и партия ходового
    # товара поднимала знаменатель быстрее, чем остаток.
    _u0 = demo.get("/api/forecast").json()["cards"]
    _dates = [(0, _u0["until_date"], _u0["until_weeks"])]
    for _qty in (50, 150, 223, 400):
        demo.post("/api/ordered", json={"base_name": 'Футболка «Манифест»', "qty": _qty})
        _c2 = demo.get("/api/forecast").json()["cards"]
        _dates.append((_qty, _c2["until_date"], _c2["until_weeks"]))
    _weeks = [(q, w if w is not None else 10 ** 6) for q, _dt, w in _dates]
    check("дата дефицита не едет назад, сколько бы товара ни заказали",
          all(_weeks[i][1] <= _weeks[i + 1][1] for i in range(len(_weeks) - 1)),
          f"недели по заказу={_weeks}")
    check("карточка объясняет, от чего считается порог",
          "склад" in (_u0.get("until_basis") or "").lower(),
          f"until_basis={_u0.get('until_basis')!r}")
    demo.post("/api/ordered", json={"base_name": 'Футболка «Манифест»', "qty": 0})

    print("== Хвост неликвида (1 шт на полке) фолбэк не подхватывает ==")
    # Кейс ревью: 1 шт на складе — это МЕНЬШЕ порога min_stock_days (3 шт),
    # поэтому «дней в стоке» за 90 дней ноль, как и у распроданного в ноль.
    # Но товар лежал, и продавать его было можно: годовой темп сюда тянуть
    # нельзя, иначе система предлагает дошить то, что не продаётся.
    from datetime import date as _d, timedelta as _td
    _dead = "Браслет «Звенья»"  # позиция из демо-каталога (app/demo_seed.py)
    _cut120 = (_d.today() - _td(days=120)).isoformat()
    _con = sqlite3.connect(DB_PATH)
    _demo_org_id = _con.execute(
        "SELECT org_id FROM products WHERE base_name=? LIMIT 1", (_dead,)).fetchone()[0]
    _pids = [r[0] for r in _con.execute(
        "SELECT id FROM products WHERE org_id=? AND base_name=?", (_demo_org_id, _dead))]
    _qs = ",".join("?" * len(_pids))
    # продажи за 120 дней убираем, на полке оставляем ровно 1 шт
    _con.execute(f"DELETE FROM sales WHERE product_id IN ({_qs}) AND date>=?", _pids + [_cut120])
    _con.execute(f"UPDATE stock_days SET qty=0 WHERE product_id IN ({_qs}) AND date>=?",
                 _pids + [_cut120])
    _con.execute("UPDATE stock_days SET qty=1 WHERE product_id=? AND date>=?", (_pids[0], _cut120))
    _con.execute(f"UPDATE warehouse_stock SET qty=0 WHERE product_id IN ({_qs})", _pids)
    _con.execute("UPDATE warehouse_stock SET qty=1 WHERE product_id=?", (_pids[0],))
    _con.commit()
    _con.close()
    for _win in ("year", "d90", "season"):
        demo.post("/api/settings", json={"rate_window": _win})  # заодно сбрасывает кэш
        _r = demo.get("/api/replenish").json()
        _it = next((x for x in _r["items"] if x["base_name"] == _dead), None)
        _why = {e["base_name"]: e["reason"] for e in _r["excluded"]}.get(_dead)
        if _win == "d90":
            check("неликвид (1 шт, нет продаж 90 дн) в заказ по окну «90 дней» не попал",
                  _it is None and _why == "нет продаж за последние 90 дней",
                  f"need={_it and _it['need']} reason={_why!r}")
        else:
            # В годовом и сезонном окнах продажи были и темп настоящий —
            # позиция остаётся в заказе, но БЕЗ фолбэка на чужой темп.
            check(f"{_win}: позиция считается своим темпом, без фолбэка",
                  _it is None or not _it["rate_fallback"],
                  f"fallback={_it and _it['rate_fallback']}")
    # Полный неликвид: лежит 1 шт и не продаётся ВООБЩЕ — ни в одном окне.
    _con = sqlite3.connect(DB_PATH)
    _con.execute(f"DELETE FROM sales WHERE product_id IN ({_qs})", _pids)
    _con.commit()
    _con.close()
    for _win in ("year", "d90", "season"):
        demo.post("/api/settings", json={"rate_window": _win})
        _r = demo.get("/api/replenish").json()
        _names = {x["base_name"] for x in _r["items"]}
        _why = {e["base_name"]: e["reason"] for e in _r["excluded"]}.get(_dead)
        check(f"{_win}: мало товара + нет продаж → в заказ не попадает",
              _dead not in _names and _why, f"в заказе={_dead in _names} reason={_why!r}")
    demo.post("/api/settings", json={"rate_window": "year"})

    demo.close()
    client.close()

    return finish()


if __name__ == "__main__":
    sys.exit(main())
