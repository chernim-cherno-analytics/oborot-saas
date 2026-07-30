"""Проверка экспорта в Excel: register → demo → 4 файла → сверка с JSON API."""
import io
import re

import time

import httpx
from openpyxl import load_workbook

BASE = "http://127.0.0.1:8002"

c = httpx.Client(base_url=BASE, follow_redirects=True, timeout=120)

# register
r = c.post("/register", data={
    "name": "Тест", "email": f"export-test-{int(time.time())}@example.com",
    "password": "secret123", "org_name": "Тестовый бренд",
})
assert r.status_code == 200, r.status_code

# demo seed
r = c.post("/api/connect/demo")
assert r.status_code == 200 and r.json().get("ok"), r.text

api = {
    "replenish": c.get("/api/replenish").json(),
    "turnover": c.get("/api/turnover").json(),
    "discounts": c.get("/api/discounts").json(),
    "budget": c.get("/api/budget", params={"amount": 2000000, "max_share": 30}).json(),
}

urls = {
    "replenish": "/api/export/replenish.xlsx",
    "turnover": "/api/export/turnover.xlsx",
    "discounts": "/api/export/discounts.xlsx",
    "budget": "/api/export/budget.xlsx?amount=2000000&max_share=30",
}

results = {}
for key, url in urls.items():
    r = c.get(url)
    assert r.status_code == 200, (key, r.status_code, r.text[:200])
    ct = r.headers["content-type"]
    cd = r.headers["content-disposition"]
    assert "spreadsheetml" in ct, ct
    assert "filename*=UTF-8''" in cd, cd
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    results[key] = (r.content, wb, ws, cd)
    print(f"{key}: {len(r.content)} bytes, sheet='{ws.title}', rows={ws.max_row}, "
          f"cols={ws.max_column}, disposition={cd}")

def cell(ws, r, c_):
    return ws.cell(row=r, column=c_).value

# ── replenish ────────────────────────────────────────────────────────────────
_, wb, ws, _ = results["replenish"]
assert ws.title == "Что заказать"
assert "Тестовый бренд" in str(cell(ws, 1, 1))
hdr = [cell(ws, 2, i) for i in range(1, 14)]
assert hdr[0] == "Позиция" and hdr[10] == "Заказать, шт" and hdr[12] == "Сумма заказа, ₽", hdr
assert ws.freeze_panes == "A3"
n_data = ws.max_row - 3  # минус титул, шапка, итог
assert n_data > 10, n_data
# сумма «Сумма заказа» из ячеек (без итоговой строки) vs API
xl_sum = 0
n_size_rows = 0
classes = set()
for r_ in range(3, ws.max_row):
    name = str(cell(ws, r_, 1) or "")
    if name.startswith("— "):
        n_size_rows += 1
        continue
    xl_sum += cell(ws, r_, 13) or 0
    classes.add(cell(ws, r_, 3))
api_sum = round(sum(it["need"] * float(it.get("cost_price") or 0)
                    for it in api["replenish"]["items"]))
total_row_val = cell(ws, ws.max_row, 13)
print(f"replenish: xl_sum={xl_sum}, api Σ need×cost={api_sum}, итоговая строка={total_row_val}, "
      f"строк размеров={n_size_rows}, классы={classes}")
assert xl_sum == api_sum == total_row_val
assert n_size_rows > 0
assert classes <= {"Бестселлер", "Хороший", "Унылый", "Слабый"}
xl_need = sum(cell(ws, r_, 11) or 0 for r_ in range(3, ws.max_row)
              if not str(cell(ws, r_, 1) or "").startswith("— "))
api_need = sum(it["need"] for it in api["replenish"]["items"])
assert xl_need == api_need == cell(ws, ws.max_row, 11), (xl_need, api_need)

# ── turnover ─────────────────────────────────────────────────────────────────
_, wb, ws, _ = results["turnover"]
assert ws.title == "Оборачиваемость"
assert cell(ws, 2, 1) == "Позиция" and cell(ws, 2, 7) == "Оборачиваемость, ₽/день"
n_items = len(api["turnover"]["items"])
assert ws.max_row == 2 + n_items + 1, (ws.max_row, n_items)
assert n_items > 10
xl_nr = sum(cell(ws, r_, 6) or 0 for r_ in range(3, ws.max_row))
api_nr = sum(it["nr"] for it in api["turnover"]["items"])
xl_cs = sum(cell(ws, r_, 11) or 0 for r_ in range(3, ws.max_row))
api_cs = sum(it["cs"] for it in api["turnover"]["items"])
print(f"turnover: строк={n_items}, Σвыручка xl={xl_nr} api={api_nr}, Σостаток xl={xl_cs} api={api_cs}")
assert xl_nr == api_nr == cell(ws, ws.max_row, 6)
assert xl_cs == api_cs == cell(ws, ws.max_row, 11)

# ── discounts ────────────────────────────────────────────────────────────────
_, wb, ws, _ = results["discounts"]
assert ws.title == "Уценка"
assert cell(ws, 2, 4) == "Заморожено, ₽" and cell(ws, 2, 9) == "Причина"
n_items = len(api["discounts"]["items"])
assert n_items > 10
xl_frozen = sum(cell(ws, r_, 4) or 0 for r_ in range(3, 3 + n_items))
api_frozen = api["discounts"]["cards"]["frozen_total"]
xl_rec = sum(cell(ws, r_, 8) or 0 for r_ in range(3, 3 + n_items))
api_rec = api["discounts"]["cards"]["expected_recovery"]
print(f"discounts: строк={n_items}, Σзаморожено xl={xl_frozen} api={api_frozen}, "
      f"Σвернёт xl={xl_rec} api={api_rec}")
assert xl_frozen == api_frozen == cell(ws, 3 + n_items, 4)
assert xl_rec == api_rec == cell(ws, 3 + n_items, 8)
# скидка — доля для формата 0%
pcts = {cell(ws, r_, 5) for r_ in range(3, 3 + n_items)}
assert all(0 < p <= 0.6 for p in pcts), pcts

# ── budget ───────────────────────────────────────────────────────────────────
_, wb, ws, _ = results["budget"]
assert ws.title == "Бюджет закупки"
assert cell(ws, 2, 9) == "Заказать, шт" and cell(ws, 2, 11) == "Сумма, ₽"
n_items = len(api["budget"]["items"])
assert n_items > 10
total_row = 2 + n_items + 1
xl_total = sum(cell(ws, r_, 11) or 0 for r_ in range(3, total_row))
api_used = api["budget"]["used"]
xl_qty = sum(cell(ws, r_, 9) or 0 for r_ in range(3, total_row))
api_qty = api["budget"]["totals"]["units"]
print(f"budget: строк={n_items}, Σсумма xl={xl_total} api used={api_used}, "
      f"Σшт xl={xl_qty} api={api_qty}")
assert xl_total == api_used == cell(ws, total_row, 11)
assert xl_qty == api_qty == cell(ws, total_row, 9)
assert "бюджет 2 000 000 ₽" in str(cell(ws, 1, 1))

# ── кнопки в DOM ─────────────────────────────────────────────────────────────
for page, needle in [
    ("/replenish", "/api/export/replenish.xlsx"),
    ("/turnover", "/api/export/turnover.xlsx"),
    ("/discounts", "/api/export/discounts.xlsx"),
    ("/budget", "/api/export/budget.xlsx"),
]:
    html = c.get(page).text
    assert needle in html and "Скачать Excel" in html, page
    print(f"{page}: кнопка «Скачать Excel» в DOM — ок")

# без сессии — 401
r = httpx.get(BASE + "/api/export/turnover.xlsx")
print(f"без сессии: {r.status_code}")
assert r.status_code == 401, r.status_code

print("\nALL CHECKS PASSED")
