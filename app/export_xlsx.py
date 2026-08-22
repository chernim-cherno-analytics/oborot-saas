"""Экспорт ключевых таблиц в Excel (.xlsx) — openpyxl, без pandas.

Целевая аудитория живёт в Excel: выгрузка — мостик доверия («могу проверить
и доработать у себя»). Данные НЕ пересчитываются — билдеры получают готовые
ответы analytics.build_replenish / build_turnover, analytics_extra.build_budget,
analytics_markdown.build_discounts и только раскладывают их по ячейкам.

Оформление едино для всех листов:
- A1 (объединённая через все колонки, серым) — организация + дата выгрузки;
- строка 2 — жирная шапка с заливкой, заморожена (freeze_panes = A3);
- автоширина колонок по содержимому, максимум 45 символов;
- числа: # ##0 (в xlsx — канонический код #,##0: русский Excel показывает
  пробелы-разряды), деньги — ₽ без копеек, проценты — 0%;
- итоговая строка жирным с верхней границей, где уместно.

ВСЁ, что попадает в ячейку, пишется через _cell(): там и только там стоит
защита от формульной инъекции (названия организации, позиций, категорий,
производств и примечания — пользовательский текст, а файл открывает
бухгалтерия клиента). Новая выгрузка получает защиту автоматически, если
пишет через _cell/_write_row/_write_total/_new_sheet и не зовёт ws.cell сама.

Ручки отдают файл через xlsx_response(): StreamingResponse с русским именем
файла по RFC 5987 (filename* = UTF-8''...) и ASCII-fallback в filename=.
"""
import io
from datetime import date
from urllib.parse import quote

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Числовые форматы (#,##0 — канонический код «# ##0» русского Excel).
FMT_INT = "#,##0"
FMT_MONEY = '#,##0" ₽"'
FMT_PCT = "0%"
FMT_NUM1 = "#,##0.0"
FMT_NUM2 = "#,##0.00"

CLS_RU = {"weak": "Слабый", "dull": "Медленный", "good": "Хороший", "best": "Бестселлер"}
_EXCLUDED_SHEET = "Не вошло и почему"  # второй лист книги «Что заказать»

# Подпись окна темпа для шапки листа (fallback, если ответ API её не отдал).
_RATE_WINDOW_RU = {
    "year": "темп за год",
    "d90": "темп за 90 дней",
    "season": "сезонный темп",
}

_TITLE_FONT = Font(color="808080", size=9)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill("solid", fgColor="1F2B3D")  # тёмно-синий, как сайдбар
_SIZE_FONT = Font(color="6B7280", size=10)             # строки размеров — серым
_TOTAL_FONT = Font(bold=True)
_TOTAL_BORDER = Border(top=Side(style="thin", color="9CA3AF"))
_WRAP = Alignment(vertical="center", wrap_text=True)


def _rate_window_ru(data: dict) -> str:
    """Человеческая подпись активного окна темпа из ответа API."""
    return data.get("rate_window_label") or _RATE_WINDOW_RU.get(
        data.get("rate_window") or "year", _RATE_WINDOW_RU["year"]
    )


def _fmt_date_ru(iso: str | None) -> str:
    """ISO-дата → дд.мм.гггг ('' если None)."""
    if not iso:
        return ""
    try:
        return date.fromisoformat(iso).strftime("%d.%m.%Y")
    except ValueError:
        return iso


# Символы, с которых Excel начинает читать содержимое ячейки как формулу
# (в том числе DDE-команду «=cmd|...»). Табуляция и возврат каретки — потому
# что при копировании и сохранении в CSV они обрезаются, и следующим символом
# оказывается «=».
_FORMULA_STARTS = ("=", "+", "-", "@", "\t", "\r")


def _cell(ws: Worksheet, row: int, col: int, value,
          fmt: str | None = None, font: Font | None = None,
          fill: PatternFill | None = None, alignment: Alignment | None = None,
          border: Border | None = None):
    """ЕДИНСТВЕННАЯ точка записи значения в книгу — пишем только через неё.

    Почти весь текст выгрузки приходит от человека: название организации,
    названия позиций и категорий, категории, примечания, названия производств.
    Excel исполняет ячейку, которая начинается с '=', '+', '-', '@' или
    табуляции, как формулу — это формульная инъекция (CWE-1236), а файл у
    клиента открывает бухгалтерия. Здесь текст остаётся текстом:

    - data_type='s' — openpyxl сам классифицирует строку с '=' как ФОРМУЛУ
      и записывает её в <f>; принудительный тип оставляет обычную строку;
    - quotePrefix — тот самый ведущий апостроф Excel, но живущий в стиле
      ячейки, а не в значении: на экране его не видно, а при копировании
      ячейки и при сохранении в CSV защита едет вместе с текстом.

    Писать в ячейку мимо _cell() нельзя: тогда новая выгрузка окажется
    незащищённой молча. Числа и пустые значения проходят как есть.
    """
    if value is None:
        value = ""
    cell = ws.cell(row=row, column=col, value=value)
    if isinstance(value, str):
        if cell.data_type == "f":
            cell.data_type = "s"
        if value[:1] in _FORMULA_STARTS:
            cell.quotePrefix = True
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if fmt and isinstance(value, (int, float)) and not isinstance(value, bool):
        cell.number_format = fmt
    return cell


def _sheet_title(wb: Workbook, name: str) -> str:
    """Название листа из пользовательского текста: Excel не всё разрешает.

    Запрещённые символы (:\\/?*[]) меняем на дефис, режем до 31 знака и
    разводим совпадения номером — иначе openpyxl молча падает или теряет лист.
    """
    clean = "".join("-" if ch in ':\\/?*[]' else ch for ch in (name or "").strip())
    clean = clean[:31].strip() or "Производство"
    # «Не вошло и почему» добавляется в книгу последним: производство с таким
    # же названием увело бы этот лист под чужое имя.
    if clean not in wb.sheetnames and clean != _EXCLUDED_SHEET:
        return clean
    for n in range(2, 100):
        suffix = f" ({n})"
        candidate = clean[: 31 - len(suffix)] + suffix
        if candidate not in wb.sheetnames and candidate != _EXCLUDED_SHEET:
            return candidate
    return clean[:28] + "..."


def _new_sheet(wb: Workbook, sheet_title: str, org_name: str, ncols: int,
               subtitle: str = "", first: bool = True) -> Worksheet:
    """Лист с русским названием и серой строкой A1: организация + дата выгрузки.

    first=False — второй и следующие листы книги (создаются, а не берут active).
    """
    ws = wb.active if first else wb.create_sheet()
    ws.title = sheet_title
    text = f"{org_name} · выгрузка от {date.today().strftime('%d.%m.%Y')}"
    if subtitle:
        text += f" · {subtitle}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    # org_name — пользовательский ввод; защиту ставит _cell (см. её докстринг)
    _cell(ws, 1, 1, text, font=_TITLE_FONT, alignment=Alignment(vertical="center"))
    return ws


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    """Строка 2: жирная шапка с заливкой; заморозка шапки (freeze A3)."""
    for i, title in enumerate(headers, 1):
        _cell(ws, 2, i, title, font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_WRAP)
    ws.freeze_panes = "A3"


def _write_row(ws: Worksheet, row_idx: int, values: list, formats: dict[int, str],
               font: Font | None = None) -> None:
    """Строка данных: values по колонкам (1-based formats), None → пустая ячейка."""
    for i, v in enumerate(values, 1):
        _cell(ws, row_idx, i, v, fmt=formats.get(i), font=font)


def _write_total(ws: Worksheet, row_idx: int, ncols: int, values: dict[int, object],
                 formats: dict[int, str], label: str = "Итого") -> None:
    """Итоговая строка: жирным, с верхней границей по всей ширине."""
    for i in range(1, ncols + 1):
        _cell(ws, row_idx, i, values.get(i, ""), fmt=formats.get(i),
              font=_TOTAL_FONT, border=_TOTAL_BORDER)
    _cell(ws, row_idx, 1, label, font=_TOTAL_FONT, border=_TOTAL_BORDER)


def _autofit(ws: Worksheet, min_width: int = 6, max_width: int = 45) -> None:
    """Автоширина колонок по содержимому (строка A1-титула не учитывается)."""
    widths: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            v = cell.value
            if v is None or v == "":
                continue
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                # длина с разрядами-пробелами и возможным « ₽»
                length = len(f"{v:,.0f}") + 2
            else:
                length = max(len(part) for part in str(v).split("\n"))
            col = cell.column_letter
            widths[col] = max(widths.get(col, 0), length)
    for col, w in widths.items():
        ws.column_dimensions[col].width = min(max_width, max(min_width, w + 2))


def xlsx_response(wb: Workbook, filename_ru: str, filename_ascii: str) -> StreamingResponse:
    """Workbook → StreamingResponse с русским именем файла (RFC 5987 filename*)."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    disposition = (
        f'attachment; filename="{filename_ascii}"; '
        f"filename*=UTF-8''{quote(filename_ru)}"
    )
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": disposition, "Cache-Control": "no-store"},
    )


# ── «Что заказать» ────────────────────────────────────────────────────────────

def _replenish_note(it: dict, window_ru: str) -> str:
    """Примечание строки позиции: откуда взялось количество и что с ним не так.

    Человек, который несёт файл на фабрику, должен прочитать в нём ту же
    историю числа, что видел на экране: расчёт → минимальная партия → правка
    руками, и предупреждение, если правка условиям подрядчика не отвечает.
    """
    parts = []
    need = int(it.get("need") or 0)
    need_raw = int(it.get("need_raw", need) or 0)
    need_calc = int(it.get("need_calc", need) or 0)
    if it.get("moq_applied"):
        why = []
        if it.get("moq"):
            why.append(f"минимальная партия {int(it['moq'])} шт")
        if int(it.get("pack_multiple") or 0) > 1:
            why.append(f"кратность {int(it['pack_multiple'])} шт")
        parts.append(
            f"по расчёту {need_raw} шт → {need_calc} шт"
            + (f" ({' и '.join(why)})" if why else "")
        )
    if it.get("manual"):
        parts.append(f"ростовка правлена вручную: расчёт {need_calc} шт → ваш заказ {need} шт")
    if it.get("moq_note"):
        parts.append("⚠ " + it["moq_note"])
    if it.get("moq_low"):
        parts.append("потребность заметно ниже минимальной партии — решение за вами")
    # Темп по годовым продажам вместо активного окна: позиции не было на
    # складе всё окно (распродана в ноль) — считать по окну не из чего.
    if it.get("rate_fallback"):
        parts.append(f"темп посчитан за год: в окне «{window_ru}» позиции не было на складе")
    return "; ".join(parts)


_REPLENISH_HEADERS = [
    "Позиция", "Категория", "Класс", "Оборачиваемость, ₽/день", "Темп, шт/день",
    "Продано за год, шт", "Остаток, шт", "Едет, шт", "Покрытие, нед",
    "Дата стокаута", "Заказать, шт", "Себестоимость за шт, ₽", "Сумма заказа, ₽",
    "Производство", "Примечание",
]
_REPLENISH_FORMATS = {
    4: FMT_MONEY, 5: FMT_NUM2, 6: FMT_INT, 7: FMT_INT, 8: FMT_INT,
    9: FMT_NUM1, 11: FMT_INT, 12: FMT_MONEY, 13: FMT_MONEY,
}


def _group_by_production(data: dict) -> list[tuple[str, list[dict]]]:
    """Позиции заказа, разложенные по производствам В ПОРЯДКЕ их появления.

    Тот же состав, что на вкладках страницы «Заказ»: у каждой вкладки свои
    позиции, свой итог и свой срок. Позиции без производства (условия ещё не
    применялись) собираются в одну группу с пустым именем.
    """
    groups: dict[str, list[dict]] = {}
    for it in data.get("items", []):
        groups.setdefault(it.get("production_name") or "", []).append(it)
    return list(groups.items())


def _production_subtitle(name: str, items: list[dict], data: dict, window_ru: str) -> str:
    """Шапка листа производства: его СОБСТВЕННЫЕ срок, партия и кратность.

    Раньше в шапке стоял общий срок из Настроек, а строки считались по сроку
    подрядчика — файл обещал «45 дней» над позицией, которая шьётся 70.
    """
    parts = [f"производство «{name}»"] if name else []
    leads = {int(it.get("lead_time_days") or 0) for it in items if it.get("lead_time_days")}
    if len(leads) == 1:
        parts.append(f"срок производства {leads.pop()} дней")
    elif leads:
        parts.append("срок производства: у позиций разный, см. примечания")
    else:
        parts.append(f"срок производства {data.get('lead_time_days', 45)} дней")
    moq = {int(it.get("moq") or 0) for it in items}
    if len(moq) == 1 and (m := moq.pop()):
        parts.append(f"партия от {m} шт")
    step = {int(it.get("pack_multiple") or 0) for it in items}
    if len(step) == 1 and (st := step.pop()) > 1:
        parts.append(f"кратно {st} шт")
    parts.append(f"горизонт {data.get('horizon_days', 90)} дней")
    parts.append(window_ru)
    where = f"вкладке «{name}»" if name else "странице «Заказ»"
    # Раньше шапка честно предупреждала, что галочки на состав файла не
    # влияют, — это было объяснение проблемы, а не решение. Теперь выгрузка
    # умеет выгружать именно отмеченное (см. app.routes_extra), и шапка
    # говорит правду о том, что реально попало в лист.
    included = len(items)
    total_here = (data.get("selection_totals_by_production") or {}).get(name, included)
    if data.get("selection_partial") and total_here != included:
        parts.append(
            f"количества — как на {where}: с условиями производства и вашими правками"
            f" ростовки. В лист входят отмеченные позиции: {included} из {total_here}"
            f" на вкладке — остальные сняты галочкой вручную, они на листе «{_EXCLUDED_SHEET}»"
        )
    else:
        parts.append(
            f"количества — как на {where}: с условиями производства и вашими правками"
            " ростовки. В лист входят все позиции вкладки"
        )
    return " · ".join(parts)


def _replenish_sheet(wb: Workbook, org_name: str, sheet_title: str, name: str,
                     items: list[dict], data: dict, window_ru: str,
                     others: list[str], first: bool) -> Worksheet:
    """Лист одного производства: его позиции, его шапка, его «Итого»."""
    headers = _REPLENISH_HEADERS
    formats = _REPLENISH_FORMATS
    ws = _new_sheet(wb, sheet_title, org_name, len(headers), first=first,
                    subtitle=_production_subtitle(name, items, data, window_ru))
    _write_header(ws, headers)

    row = 3
    total_need = 0
    total_sum = 0.0
    for it in items:
        cost = float(it.get("cost_price") or 0)
        order_sum = round(it["need"] * cost)
        sold_year = sum(round(s.get("sold365") or 0) for s in it.get("sizes", {}).values())
        total_need += it["need"]
        total_sum += order_sum
        _write_row(ws, row, [
            it["base_name"], it.get("category") or "", CLS_RU.get(it["cls"], it["cls"]),
            it["turnover"], it["rate"], sold_year, it["cs"], it.get("ordered") or 0,
            it.get("wos"), _fmt_date_ru(it.get("stockout_date")), it["need"],
            cost if cost > 0 else None, order_sum,
            it.get("production_name") or "", _replenish_note(it, window_ru),
        ], formats)
        row += 1
        # Размерная сетка: сток / продано за год / рекомендация — отступом.
        # Правленый вручную размер подписан расчётом, от которого отошли.
        for size, s in it.get("sizes", {}).items():
            size_note = (
                f"правлено вручную · расчёт {int(s.get('rec_calc') or 0)} шт"
                if s.get("manual") else ""
            )
            _write_row(ws, row, [
                f"— {size}", None, None, None, None,
                round(s.get("sold365") or 0), s.get("stock") or 0, None, None, None,
                s.get("rec") or 0, None, None, None, size_note,
            ], formats, font=_SIZE_FONT)
            row += 1

    label = f"Итого по «{name}»: {len(items)} поз." if name else f"Итого: {len(items)} поз."
    _write_total(ws, row, len(headers),
                 {11: total_need, 13: round(total_sum)}, formats, label=label)
    note_row = row + 2
    if others:
        _cell(ws, note_row, 1,
              "Это итог одного производства. Остальные — на своих листах: "
              + ", ".join(f"«{o}»" for o in others)
              + ". Общего итога по всем производствам в файле нет намеренно:"
                " заказ вы отдаёте каждому подрядчику отдельно.",
              font=_TITLE_FONT)
        note_row += 1
    excluded = data.get("excluded") or []
    _cell(ws, note_row, 1,
          f"Позиции, не попавшие в заказ, — на листе «{_EXCLUDED_SHEET}» "
          f"({len(excluded)} шт). Числа посчитаны по темпу «{window_ru}».",
          font=_TITLE_FONT)
    _autofit(ws)
    return ws


def replenish_workbook(org_name: str, data: dict) -> Workbook:
    """Книга «Что заказать»: лист на каждое производство + «Не вошло и почему».

    Человек несёт этот файл на фабрику и должен получить ровно то, что видел
    на экране. На странице «Заказ» производства разведены по вкладкам: у
    каждой свой состав, свой итог и свой срок. Поэтому и в книге на каждое
    производство — свой лист со своей шапкой (срок, минимальная партия,
    кратность ЭТОГО подрядчика) и своим «Итого». Раньше книга была одна и
    склеивала все вкладки: «Итого: 35 поз. / 1 802 шт» не совпадало ни с
    одной вкладкой, а шапка обещала «срок производства 45 дней» над строкой,
    которая шьётся 70 дней у другого подрядчика.

    Пока производство одно, книга выглядит как раньше: лист «Что заказать» и
    лист «Не вошло и почему».

    Колонка «Заказать» — итоговое количество, ровно как на странице: расчёт,
    поднятый до минимальной партии подрядчика, а поверх — ручная правка
    ростовки, если человек её делал. Откуда взялось число, написано в
    «Примечании»; там же — предупреждение, если правка расходится с условиями
    производства.
    """
    window_ru = _rate_window_ru(data)
    wb = Workbook()
    groups = _group_by_production(data)
    if len(groups) <= 1:
        name, items = groups[0] if groups else ("", [])
        _replenish_sheet(wb, org_name, "Что заказать", name, items, data,
                         window_ru, [], first=True)
    else:
        names = [n for n, _ in groups]
        for i, (name, items) in enumerate(groups):
            title = _sheet_title(wb, name or "Без производства")
            others = [n or "Без производства" for n in names if n != name]
            _replenish_sheet(wb, org_name, title, name, items, data,
                             window_ru, others, first=(i == 0))
    _excluded_sheet(wb, org_name, data, window_ru)
    return wb


def _excluded_sheet(wb: Workbook, org_name: str, data: dict, window_ru: str) -> Worksheet:
    """Второй лист книги «Что заказать»: позиции, которые в заказ не попали.

    Товаровед должна видеть не только то, что заказываем, но и что осталось за
    бортом и почему — иначе «33 позиции вместо 42» по файлу не поймать.
    """
    headers = ["Позиция", "Почему не вошло"]
    ws = _new_sheet(
        wb, _EXCLUDED_SHEET, org_name, len(headers), first=False,
        subtitle=f"{window_ru} · горизонт {data.get('horizon_days', 90)} дней",
    )
    _write_header(ws, headers)
    row = 3
    for it in data.get("excluded", []):
        _write_row(ws, row, [it.get("base_name") or "", it.get("reason") or ""], {})
        row += 1
    if row == 3:
        _cell(ws, 3, 1, "Все позиции вошли в заказ", font=_TITLE_FONT)
    _autofit(ws)
    return ws


# ── «Оборачиваемость» ─────────────────────────────────────────────────────────

_SEASON_RU = {"winter": "зима", "spring": "весна", "summer": "лето", "autumn": "осень"}


def _turnover_note(it: dict) -> str:
    """Примечание строки: почему в деньгах стоит не то, чего человек ожидал.

    Возвраты вычитаются из выручки. Если за период их оказалось больше, чем
    продаж, нетто-выручка отрицательна: «₽/день» в такие периоды показываем
    нулём (отрицательной скорости не бывает), а средней цены продажи у базы
    нет вовсе. Молчать об этом нельзя — иначе ноль читается как «не считали».
    """
    parts = []
    if it.get("returns_over_sales"):
        parts.append(
            "возвраты за год превысили продажи: средней цены продажи нет, "
            "оборачиваемость считаем нулевой"
        )
    seasons = [_SEASON_RU.get(s, s) for s in (it.get("sea_returns") or [])]
    if seasons:
        parts.append(
            f"возвраты превысили продажи в сезонах: {', '.join(seasons)} — "
            "₽/день по ним показаны нулём"
        )
    return "; ".join(parts)


def turnover_workbook(org_name: str, data: dict) -> Workbook:
    """Лист «Оборачиваемость»: все колонки страницы /turnover."""
    headers = [
        "Позиция", "Категория", "Класс", "Дней в стоке", "Продано, шт",
        "Выручка, ₽", "Оборачиваемость, ₽/день",
        "Зима, ₽/день", "Весна, ₽/день", "Лето, ₽/день", "Осень, ₽/день",
        "Ср. цена, ₽", "Номинал, ₽",
        "Скидка факт.", "Остаток, шт", "Покрытие, нед", "Дата стокаута", "Архив",
        "Примечание",
    ]
    formats = {
        4: FMT_INT, 5: FMT_INT, 6: FMT_MONEY, 7: FMT_MONEY,
        8: FMT_MONEY, 9: FMT_MONEY, 10: FMT_MONEY, 11: FMT_MONEY,
        12: FMT_MONEY, 13: FMT_MONEY, 14: FMT_PCT, 15: FMT_INT, 16: FMT_NUM1,
    }
    wb = Workbook()
    # Покрытие и дата стокаута считаются по активному окну темпа — подписываем,
    # как и на «Что заказать»: без этого два файла с разными числами не различить.
    ws = _new_sheet(wb, "Оборачиваемость", org_name, len(headers),
                    subtitle="метрики за скользящие 365 дней · покрытие и стокаут: "
                             + _rate_window_ru(data))
    _write_header(ws, headers)

    row = 3
    t_nq = t_nr = t_turn = t_cs = 0
    for it in data.get("items", []):
        group = it.get("group") or "rank"
        # Класс — как на странице: шумовые группы вместо класса получают пометку.
        cls_label = (
            "мало данных" if group == "low_data"
            else "без продаж" if group == "no_sales"
            else CLS_RU.get(it["cls"], it["cls"])
        )
        t_nq += it["nq"]
        t_nr += it["nr"]
        if group == "rank":  # сумма шумовых оборачиваемостей не имеет смысла
            t_turn += it["turnover"]
        t_cs += it["cs"]
        sea = it.get("sea") or {}
        _write_row(ws, row, [
            it["base_name"], it.get("category") or "", cls_label,
            it["dis"], it["nq"], it["nr"],
            it["turnover"] if group == "rank" else None,
            sea.get("winter") if group == "rank" else None,
            sea.get("spring") if group == "rank" else None,
            sea.get("summer") if group == "rank" else None,
            sea.get("autumn") if group == "rank" else None,
            it.get("avg_price"),
            it.get("sale_price"), it.get("discount_fact"), it["cs"], it.get("wos"),
            _fmt_date_ru(it.get("stockout_date")), "да" if it.get("archived") else "",
            _turnover_note(it),
        ], formats)
        row += 1

    _write_total(
        ws, row, len(headers),
        {5: t_nq, 6: t_nr, 7: t_turn, 15: t_cs}, formats,
        label=f"Итого: {len(data.get('items', []))} поз.",
    )
    _autofit(ws)
    return wb


# ── «Уценка» ──────────────────────────────────────────────────────────────────

def discounts_workbook(org_name: str, data: dict) -> Workbook:
    """Лист «Уценка»: прайс уценки со скидками и причинами."""
    headers = [
        "Позиция", "Категория", "Остаток, шт", "Заморожено, ₽", "Скидка, %",
        "Старая цена, ₽", "Новая цена, ₽", "Вернёт, ₽", "Причина",
    ]
    formats = {
        3: FMT_INT, 4: FMT_MONEY, 5: FMT_PCT, 6: FMT_MONEY, 7: FMT_MONEY, 8: FMT_MONEY,
    }
    wb = Workbook()
    ws = _new_sheet(wb, "Уценка", org_name, len(headers), subtitle="прайс уценки")
    _write_header(ws, headers)

    row = 3
    t_cs = t_frozen = t_rec = 0
    for it in data.get("items", []):
        t_cs += it["cs"]
        t_frozen += it["frozen"]
        t_rec += it["expected_recovery"]
        star = " *" if it.get("no_cost") else ""
        _write_row(ws, row, [
            it["base_name"] + star, it.get("category") or "", it["cs"], it["frozen"],
            it["discount_pct"] / 100.0, it["avg_price"], it["new_price"],
            it["expected_recovery"], it.get("reason") or "",
        ], formats)
        row += 1

    _write_total(
        ws, row, len(headers),
        {3: t_cs, 4: t_frozen, 8: t_rec}, formats,
        label=f"Итого: {len(data.get('items', []))} поз.",
    )
    if any(it.get("no_cost") for it in data.get("items", [])):
        _cell(ws, row + 2, 1,
              "* — нет себестоимости, заморозка посчитана по средней цене продажи",
              font=_TITLE_FONT)
    _autofit(ws)
    return wb


# ── «Бюджет закупки» ──────────────────────────────────────────────────────────

def budget_workbook(org_name: str, data: dict) -> Workbook:
    """Лист «Бюджет закупки»: распределение бюджета по оборачиваемости."""
    headers = [
        "Позиция", "Категория", "Класс", "Оборачиваемость, ₽/день", "Темп, шт/мес",
        "Остаток+едет, шт", "Запас, дн", "Потребность, шт", "Заказать, шт",
        "Себестоимость за шт, ₽", "Сумма, ₽", "Ожидаемая прибыль, ₽", "Примечание",
    ]
    formats = {
        4: FMT_MONEY, 5: FMT_NUM1, 6: FMT_INT, 7: FMT_INT, 8: FMT_INT,
        9: FMT_INT, 10: FMT_MONEY, 11: FMT_MONEY, 12: FMT_MONEY,
    }
    subtitle = (
        f"бюджет {data.get('amount', 0):,.0f} ₽".replace(",", " ")
        + f", лимит на позицию {data.get('max_share', 0)}%"
        + f" · горизонт {data.get('horizon_days', 90)} дней"
        + f" · {_rate_window_ru(data)}"
    )
    wb = Workbook()
    ws = _new_sheet(wb, "Бюджет закупки", org_name, len(headers), subtitle=subtitle)
    _write_header(ws, headers)

    row = 3
    for it in data.get("items", []):
        if it.get("over_limit"):
            note = "дорогая позиция — 1 шт сверх лимита доли"
        elif it.get("capped") and it.get("cap_reason") == "share":
            note = f"лимит доли; потребность {it['need']} шт"
        elif it.get("capped"):
            note = f"бюджет исчерпан; потребность {it['need']} шт"
        else:
            note = ""
        _write_row(ws, row, [
            it["base_name"], it.get("category") or "", CLS_RU.get(it["cls"], it["cls"]),
            it["turnover"], round(it["rate"] * 30, 1), it["stock_eff"],
            it.get("days_left"), it["need"], it["qty"], it["cost_price"], it["total"],
            it["expected_profit"], note,
        ], formats)
        row += 1

    totals = data.get("totals", {})
    _write_total(
        ws, row, len(headers),
        {
            9: totals.get("units", 0),
            11: data.get("used", 0),
            12: totals.get("expected_profit", 0),
        },
        formats,
        label=f"Итого: {totals.get('positions', 0)} поз.",
    )
    _cell(
        ws, row + 2, 1,
        f"Использовано {data.get('used', 0):,.0f} ₽ из {data.get('amount', 0):,.0f} ₽, "
        f"остаток {data.get('rest', 0):,.0f} ₽".replace(",", " "),
        font=_TITLE_FONT,
    )
    # «Потребность» здесь намеренно другая, чем «Заказать» на «Что заказать»:
    # бюджет делит деньги и почти всегда берёт МЕНЬШЕ потребности, поэтому
    # поднимать её до минимальной партии тут нечестно — партию всё равно не
    # выкупить. Подписываем, чтобы два файла с разными числами не путали.
    _cell(
        ws, row + 3, 1,
        "«Потребность» — расчёт без условий производства: бюджет делит деньги "
        "и берёт столько, на сколько их хватает. Минимальная партия и кратность "
        "применяются на странице «Заказ» — её выгрузка и есть заказ для фабрики.",
        font=_TITLE_FONT,
    )
    _autofit(ws)
    return wb
